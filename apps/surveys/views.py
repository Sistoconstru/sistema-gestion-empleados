from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import TemplateView, ListView, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.db import models
from django.db.models import Count, Q, Avg, Max
from django.utils import timezone
from django.http import JsonResponse
from django.views import View
from datetime import datetime, timedelta

from .models import (
    Encuesta, TipoEncuesta, ParticipacionEncuesta,
    PreguntaEncuesta, RespuestaEncuesta, OpcionEncuesta
)


class DashboardView(LoginRequiredMixin, TemplateView):
    """Dashboard principal del módulo de encuestas.

    Para staff/superuser: métricas GLOBALES de toda la organización.
    Para empleado normal: métricas PERSONALES.
    """
    template_name = 'surveys/index.html'

    def get_context_data(self, **kwargs):
        from datetime import date
        context = super().get_context_data(**kwargs)
        user = self.request.user
        es_admin = user.is_staff or user.is_superuser
        hoy = date.today()

        context['es_admin'] = es_admin
        context['total_encuestas'] = Encuesta.objects.filter(activa=True).count()
        context['tipos_encuesta'] = TipoEncuesta.objects.filter(activo=True).count()

        encuestas_vigentes = Encuesta.objects.filter(
            activa=True, fecha_inicio__lte=hoy, fecha_fin__gte=hoy,
        )
        context['encuestas_recientes'] = encuestas_vigentes.order_by('-fecha_creacion')[:5]

        if es_admin:
            # Métricas GLOBALES de la organización
            participaciones = ParticipacionEncuesta.objects.filter(
                encuesta__in=encuestas_vigentes,
            )
            completadas_globales = participaciones.filter(completada=True).count()
            pendientes_globales = participaciones.filter(completada=False).count()
            total_part_vigentes = completadas_globales + pendientes_globales
            pct = round(100 * completadas_globales / total_part_vigentes, 1) if total_part_vigentes else 0

            context['encuestas_pendientes'] = pendientes_globales
            context['encuestas_completadas'] = completadas_globales
            context['participacion_stats'] = {
                'total_participaciones': total_part_vigentes,
                'participaciones_completadas': completadas_globales,
                'porcentaje_completado': pct,
            }
            # Vista global no muestra "Encuestas disponibles" por usuario
            context['encuestas_disponibles'] = []
            # Métricas propias (para el bloque "Mi actividad")
            context['mis_pendientes'] = self._mis_pendientes(hoy)
            context['mis_completadas'] = ParticipacionEncuesta.objects.filter(
                empleado__usuario=user, completada=True,
            ).count()
        else:
            # Métricas PERSONALES del empleado
            context['encuestas_pendientes'] = self._mis_pendientes(hoy)
            context['encuestas_completadas'] = ParticipacionEncuesta.objects.filter(
                empleado__usuario=user, completada=True,
            ).count()
            context['encuestas_disponibles'] = self._encuestas_disponibles(hoy)
            # Participación global sigue mostrándose como referencia informativa
            total_participaciones = ParticipacionEncuesta.objects.count()
            comp = ParticipacionEncuesta.objects.filter(completada=True).count()
            pct = round(100 * comp / total_participaciones, 1) if total_participaciones else 0
            context['participacion_stats'] = {
                'total_participaciones': total_participaciones,
                'participaciones_completadas': comp,
                'porcentaje_completado': pct,
            }
        return context

    def _mis_pendientes(self, hoy):
        vigentes = Encuesta.objects.filter(
            activa=True, fecha_inicio__lte=hoy, fecha_fin__gte=hoy,
        )
        completadas_ids = ParticipacionEncuesta.objects.filter(
            empleado__usuario=self.request.user, completada=True,
        ).values_list('encuesta_id', flat=True)
        return vigentes.exclude(id__in=completadas_ids).count()

    def _encuestas_disponibles(self, hoy):
        vigentes = Encuesta.objects.filter(
            activa=True, fecha_inicio__lte=hoy, fecha_fin__gte=hoy,
        )
        completadas_ids = ParticipacionEncuesta.objects.filter(
            empleado__usuario=self.request.user, completada=True,
        ).values_list('encuesta_id', flat=True)
        return vigentes.exclude(id__in=completadas_ids)[:6]


class EncuestaListView(LoginRequiredMixin, ListView):
    """Lista de todas las encuestas disponibles"""
    model = Encuesta
    template_name = 'surveys/encuesta_list.html'
    context_object_name = 'encuestas'
    paginate_by = 10
    
    def get_queryset(self):
        return Encuesta.objects.filter(
            activa=True,
            fecha_inicio__lte=timezone.now().date(),
            fecha_fin__gte=timezone.now().date()
        ).order_by('-fecha_creacion')


class ResponderEncuestaView(LoginRequiredMixin, View):
    """Vista para responder una encuesta específica"""
    template_name = 'surveys/responder_encuesta.html'

    def get(self, request, pk):
        """Mostrar formulario de encuesta"""
        encuesta = get_object_or_404(Encuesta, pk=pk, activa=True)

        # Verificar si ya completó la encuesta
        participacion = ParticipacionEncuesta.objects.filter(
            empleado__usuario=request.user,
            encuesta=encuesta
        ).first()

        if participacion and participacion.completada:
            messages.warning(request, 'Ya has completado esta encuesta.')
            return redirect('surveys:index')

        # Obtener o crear participación
        if not participacion:
            participacion = ParticipacionEncuesta.objects.create(
                empleado=request.user.empleado if hasattr(request.user, 'empleado') else None,
                encuesta=encuesta,
                ip_address=self.get_client_ip(request),
                dispositivo=request.META.get('HTTP_USER_AGENT', '')[:200]
            )

        # Obtener preguntas
        preguntas = PreguntaEncuesta.objects.filter(
            encuesta=encuesta,
            activa=True
        ).prefetch_related('opcionencuesta_set').order_by('orden')

        # Obtener respuestas existentes (para modo borrador)
        respuestas_existentes = {}
        if participacion:
            for respuesta in RespuestaEncuesta.objects.filter(participacion=participacion):
                if respuesta.opcion_seleccionada:
                    respuestas_existentes[str(respuesta.pregunta.id)] = respuesta.opcion_seleccionada.id
                elif respuesta.respuesta_texto:
                    respuestas_existentes[str(respuesta.pregunta.id)] = respuesta.respuesta_texto

        context = {
            'encuesta': encuesta,
            'participacion': participacion,
            'preguntas': preguntas,
            'respuestas_existentes': respuestas_existentes
        }

        return render(request, self.template_name, context)

    def post(self, request, pk):
        """Guardar respuestas de encuesta"""
        encuesta = get_object_or_404(Encuesta, pk=pk, activa=True)

        # Obtener participación existente
        participacion = ParticipacionEncuesta.objects.filter(
            empleado__usuario=request.user,
            encuesta=encuesta
        ).first()

        if not participacion:
            return JsonResponse({'success': False, 'message': 'Participación no encontrada'}, status=400)

        if participacion.completada:
            return JsonResponse({'success': False, 'message': 'Ya completaste esta encuesta'}, status=400)

        # Verificar si es guardar borrador o finalizar
        es_borrador = request.POST.get('guardar_borrador') == 'true'
        finalizar = request.POST.get('finalizar_encuesta') == 'true'

        # Procesar respuestas
        preguntas = PreguntaEncuesta.objects.filter(encuesta=encuesta, activa=True)
        respuestas_guardadas = 0

        # OpcionEncuesta.pk es UUID — validamos que el valor recibido
        # corresponda efectivamente a una opción de la pregunta antes de
        # guardarlo en opcion_seleccionada_id. Si no matchea, cae a texto.
        tipos_con_opciones_unica = {
            'multiple_choice', 'select', 'rating',
            'PREGMULT', 'ESCALA_5', 'SI_NO', 'ESCALA_3',
        }
        tipos_con_opciones_multiple = {'checkbox', 'CHECKMULT'}

        def _guardar_unica(pregunta, valor):
            if pregunta.tipo_pregunta.codigo in tipos_con_opciones_unica:
                opcion = pregunta.opcionencuesta_set.filter(pk=valor).first()
                if opcion:
                    # Reemplazar cualquier respuesta previa a esta pregunta
                    RespuestaEncuesta.objects.filter(
                        participacion=participacion, pregunta=pregunta,
                    ).delete()
                    RespuestaEncuesta.objects.create(
                        participacion=participacion, pregunta=pregunta,
                        opcion_seleccionada=opcion, respuesta_texto='',
                    )
                    return True
                # No matchea ninguna opción → guardamos texto libre ("Otro")
                RespuestaEncuesta.objects.filter(
                    participacion=participacion, pregunta=pregunta,
                ).delete()
                RespuestaEncuesta.objects.create(
                    participacion=participacion, pregunta=pregunta,
                    opcion_seleccionada=None, respuesta_texto=str(valor),
                )
                return True
            # Texto libre / texto corto
            RespuestaEncuesta.objects.update_or_create(
                participacion=participacion, pregunta=pregunta,
                opcion_seleccionada=None,
                defaults={'respuesta_texto': str(valor)},
            )
            return True

        def _guardar_multiple(pregunta, valores):
            """Checkbox: N filas, una por opción marcada. Reemplaza previas."""
            RespuestaEncuesta.objects.filter(
                participacion=participacion, pregunta=pregunta,
            ).delete()
            opciones = pregunta.opcionencuesta_set.filter(pk__in=valores)
            for opcion in opciones:
                RespuestaEncuesta.objects.create(
                    participacion=participacion, pregunta=pregunta,
                    opcion_seleccionada=opcion, respuesta_texto='',
                )
            return opciones.count() > 0

        for pregunta in preguntas:
            campo_nombre = f'pregunta_{pregunta.id}'

            if pregunta.tipo_pregunta.codigo in tipos_con_opciones_multiple:
                valores = request.POST.getlist(f'{campo_nombre}[]')
                if valores and _guardar_multiple(pregunta, valores):
                    respuestas_guardadas += 1
            else:
                valor = request.POST.get(campo_nombre)
                if valor:
                    _guardar_unica(pregunta, valor)
                    respuestas_guardadas += 1

        # Calcular porcentaje de completado (cuenta preguntas distintas con
        # al menos una respuesta — importante para checkbox que persiste N filas)
        total_preguntas = preguntas.count()
        preguntas_respondidas = (
            RespuestaEncuesta.objects.filter(participacion=participacion)
            .values('pregunta_id').distinct().count()
        )
        porcentaje = round((preguntas_respondidas / total_preguntas * 100)) if total_preguntas > 0 else 0

        # Actualizar participación
        participacion.porcentaje_completado = porcentaje

        if finalizar:
            participacion.completada = True
            participacion.fecha_completada = timezone.now()

            # Calcular tiempo empleado
            tiempo_delta = participacion.fecha_completada - participacion.fecha_inicio
            participacion.tiempo_empleado_minutos = int(tiempo_delta.total_seconds() / 60)

        participacion.save()

        if es_borrador:
            return JsonResponse({
                'success': True,
                'message': 'Borrador guardado correctamente',
                'porcentaje_completado': porcentaje
            })
        elif finalizar:
            messages.success(request, '¡Gracias por completar la encuesta!')
            return JsonResponse({
                'success': True,
                'message': 'Encuesta completada exitosamente',
                'redirect': '/encuestas/'
            })
        else:
            return JsonResponse({'success': True, 'message': 'Respuestas guardadas'})

    def get_client_ip(self, request):
        """Obtener IP del cliente"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip


class MisEncuestasView(LoginRequiredMixin, ListView):
    """Vista de encuestas del usuario actual"""
    template_name = 'surveys/mis_encuestas.html'
    context_object_name = 'participaciones'
    paginate_by = 10

    def get_queryset(self):
        return ParticipacionEncuesta.objects.filter(
            empleado__usuario=self.request.user
        ).order_by('-fecha_inicio')


# =============================================================================
# VISTAS DE ADMINISTRACIÓN DE ENCUESTAS (solo staff/superuser)
# =============================================================================

class EncuestaAdminListView(LoginRequiredMixin, ListView):
    """Lista de encuestas para administración"""
    model = Encuesta
    template_name = 'surveys/admin/encuesta_admin_list.html'
    context_object_name = 'encuestas'
    paginate_by = 15

    def dispatch(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser):
            messages.error(request, 'No tienes permisos para acceder a esta sección')
            return redirect('surveys:index')
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        from datetime import date
        qs = Encuesta.objects.all().select_related('tipo_encuesta', 'creada_por').order_by('-fecha_creacion')
        hoy = date.today()
        estado = self.request.GET.get('estado', 'todas')
        if estado == 'activa':
            qs = qs.filter(activa=True, fecha_fin__gte=hoy)
        elif estado == 'inactiva':
            qs = qs.filter(activa=False)
        elif estado == 'vencida':
            qs = qs.filter(fecha_fin__lt=hoy)
        elif estado == 'borrador':
            qs = qs.filter(activa=False, fecha_inicio__gt=hoy)
        # 'todas' → sin filtro adicional
        buscar = self.request.GET.get('q', '').strip()
        if buscar:
            qs = qs.filter(Q(nombre__icontains=buscar) | Q(codigo__icontains=buscar))
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['estado_filtro'] = self.request.GET.get('estado', 'todas')
        ctx['buscar'] = self.request.GET.get('q', '')
        # Conteos por estado (para badges en botones de filtro)
        from datetime import date
        hoy = date.today()
        todas = Encuesta.objects.all()
        ctx['count_todas'] = todas.count()
        ctx['count_activa'] = todas.filter(activa=True, fecha_fin__gte=hoy).count()
        ctx['count_inactiva'] = todas.filter(activa=False).count()
        ctx['count_vencida'] = todas.filter(fecha_fin__lt=hoy).count()
        return ctx


class CrearEncuestaView(LoginRequiredMixin, View):
    """Vista para crear una nueva encuesta"""
    template_name = 'surveys/admin/crear_encuesta.html'

    def dispatch(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser):
            messages.error(request, 'No tienes permisos para crear encuestas')
            return redirect('surveys:index')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        """Mostrar formulario de creación"""
        tipos_encuesta = TipoEncuesta.objects.filter(activo=True)
        context = {
            'tipos_encuesta': tipos_encuesta
        }
        return render(request, self.template_name, context)

    def post(self, request):
        """Guardar nueva encuesta"""
        try:
            # Validar datos requeridos
            codigo = request.POST.get('codigo')
            nombre = request.POST.get('nombre')
            descripcion = request.POST.get('descripcion', '')
            instrucciones = request.POST.get('instrucciones', '')
            tipo_encuesta_id = request.POST.get('tipo_encuesta')
            fecha_inicio = request.POST.get('fecha_inicio')
            fecha_fin = request.POST.get('fecha_fin')

            if not all([codigo, nombre, tipo_encuesta_id, fecha_inicio, fecha_fin]):
                messages.error(request, 'Todos los campos obligatorios deben ser completados')
                return redirect('surveys:crear_encuesta')

            # Verificar que el código sea único
            if Encuesta.objects.filter(codigo=codigo).exists():
                messages.error(request, f'El código "{codigo}" ya existe. Usa otro código.')
                return redirect('surveys:crear_encuesta')

            # Crear encuesta
            encuesta = Encuesta.objects.create(
                codigo=codigo,
                nombre=nombre,
                descripcion=descripcion,
                instrucciones=instrucciones,
                tipo_encuesta_id=tipo_encuesta_id,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                creada_por=request.user,
                activa=True
            )

            messages.success(request, f'Encuesta "{nombre}" creada exitosamente. Ahora agrega las preguntas.')
            return redirect('surveys:editar_preguntas', pk=encuesta.pk)

        except Exception as e:
            messages.error(request, f'Error al crear encuesta: {str(e)}')
            return redirect('surveys:crear_encuesta')


class EditarPreguntasView(LoginRequiredMixin, View):
    """Vista para agregar/editar preguntas de una encuesta"""
    template_name = 'surveys/admin/editar_preguntas.html'

    def dispatch(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser):
            messages.error(request, 'No tienes permisos para editar encuestas')
            return redirect('surveys:index')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, pk):
        """Mostrar formulario de preguntas"""
        from apps.evaluations.models import TipoPregunta

        encuesta = get_object_or_404(Encuesta, pk=pk)
        preguntas = PreguntaEncuesta.objects.filter(encuesta=encuesta).order_by('orden')
        tipos_pregunta = TipoPregunta.objects.all()

        context = {
            'encuesta': encuesta,
            'preguntas': preguntas,
            'tipos_pregunta': tipos_pregunta
        }
        return render(request, self.template_name, context)

    def post(self, request, pk):
        """Agregar nueva pregunta"""
        encuesta = get_object_or_404(Encuesta, pk=pk)

        try:
            pregunta_texto = request.POST.get('pregunta')
            tipo_pregunta_id = request.POST.get('tipo_pregunta')
            categoria = request.POST.get('categoria', '')
            descripcion = request.POST.get('descripcion', '')
            obligatoria = request.POST.get('obligatoria') == 'on'

            # Obtener siguiente orden
            ultimo_orden = PreguntaEncuesta.objects.filter(encuesta=encuesta).aggregate(
                max_orden=models.Max('orden')
            )['max_orden'] or 0

            # Crear pregunta
            pregunta = PreguntaEncuesta.objects.create(
                encuesta=encuesta,
                tipo_pregunta_id=tipo_pregunta_id,
                categoria=categoria,
                pregunta=pregunta_texto,
                descripcion=descripcion,
                obligatoria=obligatoria,
                orden=ultimo_orden + 1,
                activa=True
            )

            # Si tiene opciones, crearlas
            opciones = request.POST.getlist('opciones[]')
            valores = request.POST.getlist('valores[]')

            for idx, opcion_texto in enumerate(opciones):
                if opcion_texto.strip():
                    OpcionEncuesta.objects.create(
                        pregunta=pregunta,
                        opcion=opcion_texto,
                        valor_numerico=valores[idx] if idx < len(valores) and valores[idx].isdigit() else None,
                        orden=idx + 1,
                        activa=True
                    )

            messages.success(request, 'Pregunta agregada exitosamente')
            return redirect('surveys:editar_preguntas', pk=encuesta.pk)

        except Exception as e:
            messages.error(request, f'Error al agregar pregunta: {str(e)}')
            return redirect('surveys:editar_preguntas', pk=encuesta.pk)


class AsignarEncuestaView(LoginRequiredMixin, View):
    """Vista para asignar encuesta a empleados/áreas"""
    template_name = 'surveys/admin/asignar_encuesta.html'

    def dispatch(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser):
            messages.error(request, 'No tienes permisos para asignar encuestas')
            return redirect('surveys:index')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, pk):
        """Mostrar formulario de asignación"""
        from apps.employees.models import Empleado
        from apps.organizational.models import AreaEmpresa, Cargo

        encuesta = get_object_or_404(Encuesta, pk=pk)

        # Obtener empleados, áreas y cargos
        empleados = Empleado.objects.filter(estado__codigo='999').select_related('usuario', 'estado').order_by('apellidos', 'nombres')
        areas = AreaEmpresa.objects.all().order_by('nombre')
        cargos = Cargo.objects.all().order_by('nombre')

        # Obtener asignaciones existentes
        from .models import EncuestaArea, EncuestaCargo
        areas_asignadas = EncuestaArea.objects.filter(encuesta=encuesta).values_list('area_id', flat=True)
        cargos_asignados = EncuestaCargo.objects.filter(encuesta=encuesta).values_list('cargo_id', flat=True)

        # Obtener empleados que ya tienen participación
        participaciones_existentes = ParticipacionEncuesta.objects.filter(
            encuesta=encuesta
        ).values_list('empleado_id', flat=True)

        context = {
            'encuesta': encuesta,
            'empleados': empleados,
            'areas': areas,
            'cargos': cargos,
            'areas_asignadas': list(areas_asignadas),
            'cargos_asignados': list(cargos_asignados),
            'participaciones_existentes': list(participaciones_existentes)
        }
        return render(request, self.template_name, context)

    def post(self, request, pk):
        """Procesar asignación"""
        from apps.employees.models import Empleado
        from apps.organizational.models import AreaEmpresa, Cargo
        from .models import EncuestaArea, EncuestaCargo

        encuesta = get_object_or_404(Encuesta, pk=pk)

        try:
            asignar_a = request.POST.get('asignar_a')

            if asignar_a == 'todos':
                # Asignar a todos los empleados activos
                empleados = Empleado.objects.filter(estado__codigo='999')
                creados = 0

                for empleado in empleados:
                    _, created = ParticipacionEncuesta.objects.get_or_create(
                        empleado=empleado,
                        encuesta=encuesta
                    )
                    if created:
                        creados += 1

                messages.success(request, f'Encuesta asignada a {creados} empleados')

            elif asignar_a == 'areas':
                # Asignar por áreas
                areas_ids = request.POST.getlist('areas[]')

                # Guardar relaciones
                EncuestaArea.objects.filter(encuesta=encuesta).delete()
                for area_id in areas_ids:
                    EncuestaArea.objects.create(encuesta=encuesta, area_id=area_id)

                # Crear participaciones para empleados de esas áreas
                creados = 0
                for area_id in areas_ids:
                    empleados = Empleado.objects.filter(
                        historialcargo__cargo__area_id=area_id,
                        historialcargo__activo=True,
                        estado__codigo='999'
                    ).distinct()

                    for empleado in empleados:
                        _, created = ParticipacionEncuesta.objects.get_or_create(
                            empleado=empleado,
                            encuesta=encuesta
                        )
                        if created:
                            creados += 1

                messages.success(request, f'Encuesta asignada a {creados} empleados de las áreas seleccionadas')

            elif asignar_a == 'cargos':
                # Asignar por cargos
                cargos_ids = request.POST.getlist('cargos[]')

                # Guardar relaciones
                EncuestaCargo.objects.filter(encuesta=encuesta).delete()
                for cargo_id in cargos_ids:
                    EncuestaCargo.objects.create(encuesta=encuesta, cargo_id=cargo_id)

                # Crear participaciones para empleados con esos cargos
                creados = 0
                for cargo_id in cargos_ids:
                    empleados = Empleado.objects.filter(
                        historialcargo__cargo_id=cargo_id,
                        historialcargo__activo=True,
                        estado__codigo='999'
                    ).distinct()

                    for empleado in empleados:
                        _, created = ParticipacionEncuesta.objects.get_or_create(
                            empleado=empleado,
                            encuesta=encuesta
                        )
                        if created:
                            creados += 1

                messages.success(request, f'Encuesta asignada a {creados} empleados de los cargos seleccionados')

            elif asignar_a == 'empleados':
                # Asignar a empleados específicos
                empleados_ids = request.POST.getlist('empleados[]')
                creados = 0

                for empleado_id in empleados_ids:
                    _, created = ParticipacionEncuesta.objects.get_or_create(
                        empleado_id=empleado_id,
                        encuesta=encuesta
                    )
                    if created:
                        creados += 1

                messages.success(request, f'Encuesta asignada a {creados} empleados seleccionados')

            return redirect('surveys:asignar_encuesta', pk=encuesta.pk)

        except Exception as e:
            messages.error(request, f'Error al asignar encuesta: {str(e)}')
            return redirect('surveys:asignar_encuesta', pk=encuesta.pk)

class ResultadosEncuestaView(LoginRequiredMixin, View):
    """Dashboard de resultados de una encuesta específica. Solo staff.

    Muestra por pregunta:
    - Con opciones: distribución (conteos + %) para cada OpcionEncuesta,
      más una fila "Otro (texto libre)" cuando hay respuestas sueltas.
    - Texto libre: lista de respuestas paginadas por respuesta_texto.
    """
    template_name = 'surveys/admin/resultados_encuesta.html'

    def dispatch(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser):
            messages.error(request, 'No tienes permisos para ver resultados.')
            return redirect('surveys:index')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, pk):
        encuesta = get_object_or_404(Encuesta, pk=pk)

        # === KPIs ===
        participaciones = ParticipacionEncuesta.objects.filter(encuesta=encuesta)
        total_part = participaciones.count()
        completadas = participaciones.filter(completada=True).count()
        en_progreso = total_part - completadas
        tiempo_promedio = participaciones.filter(
            completada=True, tiempo_empleado_minutos__isnull=False,
        ).aggregate(m=Avg('tiempo_empleado_minutos'))['m']

        # === Preguntas y distribuciones ===
        preguntas = list(
            PreguntaEncuesta.objects.filter(encuesta=encuesta, activa=True)
            .select_related('tipo_pregunta').order_by('orden')
        )

        # IDs de participaciones completadas — solo estas cuentan para agregación
        part_ids_completas = list(
            participaciones.filter(completada=True).values_list('id', flat=True)
        )

        preguntas_data = []
        tipos_con_opciones = {
            'multiple_choice', 'PREGMULT', 'select', 'rating',
            'ESCALA_5', 'ESCALA_3', 'SI_NO', 'checkbox', 'CHECKMULT',
        }
        tipos_multiples = {'checkbox', 'CHECKMULT'}
        for p in preguntas:
            respuestas_p = RespuestaEncuesta.objects.filter(
                pregunta=p, participacion_id__in=part_ids_completas,
            )
            # Personas que respondieron (denominador para %)
            personas_resp = respuestas_p.values('participacion_id').distinct().count()
            es_multiple = p.tipo_pregunta.codigo in tipos_multiples
            # Total mostrado: en múltiple es personas; en única es filas (=personas)
            total_resp = personas_resp

            item = {
                'pregunta': p,
                'total_respuestas': total_resp,
                'tiene_opciones': p.tipo_pregunta.codigo in tipos_con_opciones,
                'es_multiple': es_multiple,
                'opciones': [],
                'respuestas_texto': [],
                'otras_texto_count': 0,
            }

            if item['tiene_opciones']:
                opciones = list(p.opcionencuesta_set.filter(activa=True).order_by('orden'))
                conteos = dict(
                    respuestas_p.exclude(opcion_seleccionada__isnull=True)
                    .values_list('opcion_seleccionada_id')
                    .annotate(n=Count('id')).values_list('opcion_seleccionada_id', 'n')
                )
                for o in opciones:
                    n = conteos.get(o.id, 0)
                    # Siempre % sobre personas que respondieron la pregunta.
                    # En múltiple, la suma de % puede superar 100 (esperado).
                    pct = round(100 * n / personas_resp, 1) if personas_resp else 0
                    item['opciones'].append({'opcion': o, 'n': n, 'pct': pct})
                # Texto libre ("Otro" o mismatch): respuestas sin opción pero con texto
                otras = respuestas_p.filter(
                    opcion_seleccionada__isnull=True,
                ).exclude(respuesta_texto='')
                item['otras_texto_count'] = otras.count()
                item['otras_texto'] = list(otras.values_list('respuesta_texto', flat=True)[:20])
            else:
                # Texto libre — muestro las últimas 50 respuestas
                item['respuestas_texto'] = list(
                    respuestas_p.exclude(respuesta_texto='')
                    .order_by('-fecha_respuesta')
                    .values_list('respuesta_texto', flat=True)[:50]
                )

            preguntas_data.append(item)

        # Cobertura: cuántos empleados fueron elegibles y cuántos respondieron.
        # Reutilizamos _destinatarios_encuesta para consistencia con el signal
        # de publicación y el cron: asignación por cargo/área → esa lista;
        # sin asignación → toda la empresa (encuesta general).
        cobertura_total = _destinatarios_encuesta(encuesta).count()
        pct_respuesta = (
            round(100 * completadas / cobertura_total, 1)
            if cobertura_total else None
        )

        return render(request, self.template_name, {
            'encuesta': encuesta,
            'total_participaciones': total_part,
            'completadas': completadas,
            'en_progreso': en_progreso,
            'tiempo_promedio': round(tiempo_promedio, 1) if tiempo_promedio else None,
            'cobertura_total': cobertura_total,
            'pct_respuesta': pct_respuesta,
            'preguntas_data': preguntas_data,
        })


def _destinatarios_encuesta(encuesta):
    """Retorna queryset de Empleados destinatarios de la encuesta.

    - Si hay EncuestaCargo o EncuestaArea: los empleados con cargo activo
      en esa lista (unión).
    - Si NO hay ninguna asignación → encuesta GENERAL: todos los empleados
      activos con usuario. Es lo esperado por el negocio: una encuesta sin
      cargo/área es para toda la empresa.
    - Filtra siempre por estado activo/prueba y usuario activo.
    """
    from apps.employees.models import Empleado
    from .models import EncuestaCargo, EncuestaArea

    cargo_ids = list(
        EncuestaCargo.objects.filter(encuesta=encuesta).values_list('cargo_id', flat=True)
    )
    area_ids = list(
        EncuestaArea.objects.filter(encuesta=encuesta).values_list('area_id', flat=True)
    )

    base = (
        Empleado.objects
        .filter(estado__codigo__in=['999', 'p-prue'])
        .filter(usuario__isnull=False, usuario__is_active=True)
    )
    if not (cargo_ids or area_ids):
        return base.distinct()

    filtro = Q()
    if cargo_ids:
        filtro |= Q(historialcargo__cargo_id__in=cargo_ids)
    if area_ids:
        filtro |= Q(historialcargo__cargo__area_id__in=area_ids)

    return base.filter(filtro, historialcargo__activo=True).distinct()


class ExportarRespuestasEncuestaView(LoginRequiredMixin, View):
    """Descarga XLSX con las respuestas de una encuesta. Solo staff.

    Formato: una hoja "Respuestas" con una fila por (participación x pregunta).
    Columnas: participación, empleado (o "Anónimo"), documento, cargo, área,
    pregunta, tipo, opción elegida, texto libre, fecha respuesta.
    """

    def dispatch(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser):
            messages.error(request, 'No tienes permisos.')
            return redirect('surveys:index')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, pk):
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        encuesta = get_object_or_404(Encuesta, pk=pk)

        respuestas = (
            RespuestaEncuesta.objects
            .filter(participacion__encuesta=encuesta, participacion__completada=True)
            .select_related(
                'participacion__empleado',
                'pregunta__tipo_pregunta',
                'opcion_seleccionada',
            )
            .order_by('participacion__fecha_completada', 'pregunta__orden')
        )

        wb = Workbook()
        ws = wb.active
        ws.title = 'Respuestas'

        headers = ['Participación', 'Empleado', 'Documento', 'Cargo actual', 'Área',
                   'Pregunta', 'Tipo pregunta', 'Opción elegida', 'Texto libre',
                   'Fecha respuesta']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='0E5F3F')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for r in respuestas:
            emp = r.participacion.empleado
            if emp:
                hc = emp.historialcargo_set.filter(activo=True).select_related('cargo__area').first()
                cargo = hc.cargo.nombre if hc and hc.cargo else ''
                area = hc.cargo.area.nombre if hc and hc.cargo and hc.cargo.area_id else ''
                nombre = emp.nombre_completo
                documento = emp.numero_documento
            else:
                nombre = 'Anónimo'
                documento = cargo = area = ''
            ws.append([
                str(r.participacion.id),
                nombre,
                documento,
                cargo,
                area,
                r.pregunta.pregunta,
                r.pregunta.tipo_pregunta.nombre,
                r.opcion_seleccionada.opcion if r.opcion_seleccionada else '',
                r.respuesta_texto or '',
                r.fecha_respuesta.strftime('%d/%m/%Y %H:%M'),
            ])

        anchos = [12, 32, 14, 28, 24, 45, 22, 30, 40, 16]
        for i, w in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'

        resp = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f'encuesta_{encuesta.codigo}_respuestas.xlsx'
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(resp)
        return resp


# =============================================================================
# EDICIÓN DE ENCUESTA, PREGUNTAS Y OPCIONES (Bloque A)
# =============================================================================


class _StaffRequiredMixin:
    def dispatch(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser):
            messages.error(request, 'No tienes permisos.')
            return redirect('surveys:index')
        return super().dispatch(request, *args, **kwargs)


class EditarEncuestaView(LoginRequiredMixin, _StaffRequiredMixin, View):
    """Editar metadata de la encuesta (nombre, descripción, fechas, tipo)."""
    template_name = 'surveys/admin/editar_encuesta.html'

    def get(self, request, pk):
        encuesta = get_object_or_404(Encuesta, pk=pk)
        tipos_encuesta = TipoEncuesta.objects.filter(activo=True).order_by('nombre')
        return render(request, self.template_name, {
            'encuesta': encuesta, 'tipos_encuesta': tipos_encuesta,
        })

    def post(self, request, pk):
        encuesta = get_object_or_404(Encuesta, pk=pk)
        try:
            nombre = request.POST.get('nombre', '').strip()
            descripcion = request.POST.get('descripcion', '').strip()
            instrucciones = request.POST.get('instrucciones', '').strip()
            tipo_id = request.POST.get('tipo_encuesta')
            fecha_inicio = request.POST.get('fecha_inicio')
            fecha_fin = request.POST.get('fecha_fin')
            if not all([nombre, tipo_id, fecha_inicio, fecha_fin]):
                messages.error(request, 'Todos los campos obligatorios deben ser completados.')
                return redirect('surveys:editar_encuesta', pk=pk)
            encuesta.nombre = nombre
            encuesta.descripcion = descripcion
            encuesta.instrucciones = instrucciones
            encuesta.tipo_encuesta_id = tipo_id
            encuesta.fecha_inicio = fecha_inicio
            encuesta.fecha_fin = fecha_fin
            encuesta.save()
            messages.success(request, 'Encuesta actualizada.')
            return redirect('surveys:encuesta_admin_list')
        except Exception as e:
            messages.error(request, f'Error al actualizar: {e}')
            return redirect('surveys:editar_encuesta', pk=pk)


class ToggleActivaEncuestaView(LoginRequiredMixin, _StaffRequiredMixin, View):
    """Cerrar (activa=False) o reactivar una encuesta manualmente."""

    def post(self, request, pk):
        encuesta = get_object_or_404(Encuesta, pk=pk)
        encuesta.activa = not encuesta.activa
        encuesta.save()
        estado = 'reactivada' if encuesta.activa else 'cerrada'
        messages.success(request, f'Encuesta {estado}.')
        # Volver a la lista o al referer
        return redirect(request.META.get('HTTP_REFERER', 'surveys:encuesta_admin_list'))


class EditarPreguntaView(LoginRequiredMixin, _StaffRequiredMixin, View):
    """Editar los campos de una pregunta existente."""

    def post(self, request, pk, pregunta_id):
        pregunta = get_object_or_404(PreguntaEncuesta, pk=pregunta_id, encuesta_id=pk)
        try:
            texto = request.POST.get('pregunta', '').strip()
            tipo_id = request.POST.get('tipo_pregunta')
            categoria = request.POST.get('categoria', '')
            descripcion = request.POST.get('descripcion', '')
            obligatoria = request.POST.get('obligatoria') == 'on'
            if not texto or not tipo_id:
                messages.error(request, 'Texto y tipo son obligatorios.')
                return redirect('surveys:editar_preguntas', pk=pk)

            # Si cambia el tipo y ya hay respuestas: advertir pero permitir
            if str(pregunta.tipo_pregunta_id) != str(tipo_id):
                if RespuestaEncuesta.objects.filter(pregunta=pregunta).exists():
                    messages.warning(request,
                        'Cambiaste el tipo de una pregunta que ya tenía respuestas. '
                        'Las respuestas antiguas quedaron en el histórico pero pueden '
                        'no verse coherentes en el nuevo tipo.')
            pregunta.pregunta = texto
            pregunta.tipo_pregunta_id = tipo_id
            pregunta.categoria = categoria
            pregunta.descripcion = descripcion
            pregunta.obligatoria = obligatoria
            pregunta.save()
            messages.success(request, 'Pregunta actualizada.')
        except Exception as e:
            messages.error(request, f'Error: {e}')
        return redirect('surveys:editar_preguntas', pk=pk)


class EliminarPreguntaView(LoginRequiredMixin, _StaffRequiredMixin, View):
    """Elimina una pregunta. Si tiene respuestas, la desactiva en su lugar."""

    def post(self, request, pk, pregunta_id):
        pregunta = get_object_or_404(PreguntaEncuesta, pk=pregunta_id, encuesta_id=pk)
        if RespuestaEncuesta.objects.filter(pregunta=pregunta).exists():
            pregunta.activa = False
            pregunta.save()
            messages.info(request,
                f'La pregunta tenía respuestas: se desactivó (no se elimina para '
                f'preservar el histórico).')
        else:
            pregunta.delete()
            messages.success(request, 'Pregunta eliminada.')
        return redirect('surveys:editar_preguntas', pk=pk)


class MoverPreguntaView(LoginRequiredMixin, _StaffRequiredMixin, View):
    """Sube o baja una pregunta en el orden. Intercambia con la vecina.

    El unique_together (encuesta, orden) impide el swap directo, así que
    usamos un orden puente fuera del rango (max+1) para uno de los dos saves.
    """

    def post(self, request, pk, pregunta_id):
        from django.db import transaction
        direccion = request.POST.get('direccion')
        pregunta = get_object_or_404(PreguntaEncuesta, pk=pregunta_id, encuesta_id=pk)
        qs = PreguntaEncuesta.objects.filter(encuesta_id=pk)
        if direccion == 'up':
            vecina = qs.filter(orden__lt=pregunta.orden).order_by('-orden').first()
        else:
            vecina = qs.filter(orden__gt=pregunta.orden).order_by('orden').first()
        if vecina:
            puente = (qs.aggregate(m=models.Max('orden'))['m'] or 0) + 1
            orden_p, orden_v = pregunta.orden, vecina.orden
            with transaction.atomic():
                pregunta.orden = puente
                pregunta.save(update_fields=['orden'])
                vecina.orden = orden_p
                vecina.save(update_fields=['orden'])
                pregunta.orden = orden_v
                pregunta.save(update_fields=['orden'])
        return redirect('surveys:editar_preguntas', pk=pk)


class CrearOpcionView(LoginRequiredMixin, _StaffRequiredMixin, View):
    """Agrega una opción a una pregunta existente."""

    def post(self, request, pk, pregunta_id):
        pregunta = get_object_or_404(PreguntaEncuesta, pk=pregunta_id, encuesta_id=pk)
        texto = request.POST.get('opcion', '').strip()
        valor_raw = request.POST.get('valor_numerico', '').strip()
        if not texto:
            messages.error(request, 'La opción no puede estar vacía.')
            return redirect('surveys:editar_preguntas', pk=pk)
        ultimo = pregunta.opcionencuesta_set.aggregate(m=models.Max('orden'))['m'] or 0
        OpcionEncuesta.objects.create(
            pregunta=pregunta, opcion=texto,
            valor_numerico=int(valor_raw) if valor_raw.isdigit() else None,
            orden=ultimo + 1, activa=True,
        )
        messages.success(request, 'Opción agregada.')
        return redirect('surveys:editar_preguntas', pk=pk)


class EditarOpcionView(LoginRequiredMixin, _StaffRequiredMixin, View):
    def post(self, request, pk, pregunta_id, opcion_id):
        opcion = get_object_or_404(
            OpcionEncuesta, pk=opcion_id, pregunta_id=pregunta_id,
            pregunta__encuesta_id=pk,
        )
        texto = request.POST.get('opcion', '').strip()
        valor_raw = request.POST.get('valor_numerico', '').strip()
        if not texto:
            messages.error(request, 'La opción no puede estar vacía.')
            return redirect('surveys:editar_preguntas', pk=pk)
        opcion.opcion = texto
        opcion.valor_numerico = int(valor_raw) if valor_raw.isdigit() else None
        opcion.save()
        messages.success(request, 'Opción actualizada.')
        return redirect('surveys:editar_preguntas', pk=pk)


class EliminarOpcionView(LoginRequiredMixin, _StaffRequiredMixin, View):
    """Elimina opción. Si tiene respuestas, la desactiva."""

    def post(self, request, pk, pregunta_id, opcion_id):
        opcion = get_object_or_404(
            OpcionEncuesta, pk=opcion_id, pregunta_id=pregunta_id,
            pregunta__encuesta_id=pk,
        )
        if RespuestaEncuesta.objects.filter(opcion_seleccionada=opcion).exists():
            opcion.activa = False
            opcion.save()
            messages.info(request,
                'La opción tenía respuestas: se desactivó (no se elimina para '
                'preservar el histórico).')
        else:
            opcion.delete()
            messages.success(request, 'Opción eliminada.')
        return redirect('surveys:editar_preguntas', pk=pk)


class PreviewEncuestaView(LoginRequiredMixin, _StaffRequiredMixin, View):
    """Vista previa (staff). No requiere activa=True ni crea participación.

    Reutiliza el mismo template de respuesta pero en modo readonly.
    """
    template_name = 'surveys/responder_encuesta.html'

    def get(self, request, pk):
        encuesta = get_object_or_404(Encuesta, pk=pk)
        preguntas = PreguntaEncuesta.objects.filter(
            encuesta=encuesta, activa=True,
        ).prefetch_related('opcionencuesta_set').order_by('orden')
        return render(request, self.template_name, {
            'encuesta': encuesta,
            'participacion': None,
            'preguntas': preguntas,
            'respuestas_existentes': {},
            'preview': True,
        })


class DuplicarEncuestaView(LoginRequiredMixin, _StaffRequiredMixin, View):
    """Clona una encuesta con todas sus preguntas, opciones y asignaciones.

    La copia queda inactiva y con código {orig}-COPIA. Sin participaciones ni
    respuestas. Redirige al editor de la nueva.
    """

    def post(self, request, pk):
        from django.db import transaction
        from .models import EncuestaCargo, EncuestaArea

        original = get_object_or_404(Encuesta, pk=pk)
        # Buscar un código único
        base = f'{original.codigo}-COPIA'
        codigo = base
        n = 2
        while Encuesta.objects.filter(codigo=codigo).exists():
            codigo = f'{base}-{n}'
            n += 1

        with transaction.atomic():
            nueva = Encuesta.objects.create(
                codigo=codigo,
                nombre=f'{original.nombre} (copia)',
                descripcion=original.descripcion,
                instrucciones=original.instrucciones,
                tipo_encuesta=original.tipo_encuesta,
                fecha_inicio=original.fecha_inicio,
                fecha_fin=original.fecha_fin,
                creada_por=request.user,
                activa=False,
            )
            # Clonar preguntas y opciones
            for p in PreguntaEncuesta.objects.filter(encuesta=original).order_by('orden'):
                nueva_p = PreguntaEncuesta.objects.create(
                    encuesta=nueva,
                    tipo_pregunta=p.tipo_pregunta,
                    pregunta=p.pregunta,
                    descripcion=p.descripcion,
                    categoria=p.categoria,
                    obligatoria=p.obligatoria,
                    orden=p.orden,
                    activa=p.activa,
                )
                for o in p.opcionencuesta_set.order_by('orden'):
                    OpcionEncuesta.objects.create(
                        pregunta=nueva_p,
                        opcion=o.opcion,
                        valor_numerico=o.valor_numerico,
                        orden=o.orden,
                        activa=o.activa,
                    )
            # Clonar asignaciones
            for ec in EncuestaCargo.objects.filter(encuesta=original):
                EncuestaCargo.objects.create(encuesta=nueva, cargo=ec.cargo)
            for ea in EncuestaArea.objects.filter(encuesta=original):
                EncuestaArea.objects.create(encuesta=nueva, area=ea.area)

        messages.success(request,
            f'Encuesta duplicada como "{nueva.codigo}". Queda inactiva; '
            f'revisa y publícala cuando esté lista.')
        return redirect('surveys:editar_encuesta', pk=nueva.pk)
