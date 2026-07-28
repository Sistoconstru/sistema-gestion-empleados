"""Generación del reporte Excel de Novedades de nómina.

Un solo formato se usa desde dos scopes:
  - RRHH: recibe un queryset con todas las novedades filtradas.
  - Jefe: recibe solo las novedades del equipo del jefe.
"""

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# --- Estilos ---
COLOR_CORPORATIVO = '1E3A5F'  # azul oscuro Construinmuniza
FONT_TITULO = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
FONT_META = Font(name='Calibri', size=10, italic=True, color='555555')
FONT_HEADER = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
FONT_CELDA = Font(name='Calibri', size=10)
FONT_TOTAL = Font(name='Calibri', size=11, bold=True)

FILL_TITULO = PatternFill('solid', fgColor=COLOR_CORPORATIVO)
FILL_HEADER = PatternFill('solid', fgColor='2C5282')
FILL_TOTAL = PatternFill('solid', fgColor='EDF2F7')
FILL_APROBADA = PatternFill('solid', fgColor='C6F6D5')
FILL_RECHAZADA = PatternFill('solid', fgColor='FED7D7')
FILL_PENDIENTE = PatternFill('solid', fgColor='FEEBC8')

BORDE_FINO = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')

# Columnas del reporte (index 1-based en openpyxl)
COLUMNAS = [
    ('Fecha', 12),
    ('Empleado', 32),
    ('Documento', 14),
    ('Tipo', 22),
    ('Hora inicio', 11),
    ('Hora fin', 11),
    ('Total horas', 11),
    ('Motivo', 40),
    ('Observaciones', 30),
    ('Registrada por', 25),
    ('Estado', 12),
    ('Aprobada por', 20),
    ('Fecha aprobación', 16),
    ('Motivo rechazo', 30),
]


def _fmt_hora(t):
    return t.strftime('%H:%M') if t else ''


def _fmt_fecha(d):
    return d.strftime('%d/%m/%Y') if d else ''


def _fmt_datetime(dt):
    return dt.strftime('%d/%m/%Y %H:%M') if dt else ''


def _get_fill_por_estado(estado):
    return {
        'aprobada': FILL_APROBADA,
        'rechazada': FILL_RECHAZADA,
        'pendiente': FILL_PENDIENTE,
    }.get(estado)


def generar_excel_novedades(
    novedades,
    titulo='Reporte de Novedades',
    fecha_desde=None,
    fecha_hasta=None,
    filtros_desc=None,
    contexto_scope='',
):
    """Genera un archivo Excel con las novedades y lo retorna como bytes.

    Args:
        novedades: queryset o iterable de NovedadNomina. Se itera 1 vez.
        titulo: título en el encabezado (ej. 'Reporte de Novedades — RRHH').
        fecha_desde, fecha_hasta: rango del reporte (para el header).
        filtros_desc: str libre con descripción de filtros ('Área: X · Sede: Y').
        contexto_scope: str libre para identificar el generador
            (ej. 'Equipo de Fernando Gómez' o 'RRHH — Todas las áreas').

    Returns:
        bytes del .xlsx.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Novedades'

    # -------- 1) Encabezado --------
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNAS))
    c = ws.cell(row=1, column=1, value=titulo)
    c.font = FONT_TITULO
    c.fill = FILL_TITULO
    c.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 28

    # Metadata
    fila_meta = 2
    meta_lineas = []
    if fecha_desde and fecha_hasta:
        meta_lineas.append(f'Período: {_fmt_fecha(fecha_desde)} — {_fmt_fecha(fecha_hasta)}')
    if contexto_scope:
        meta_lineas.append(f'Alcance: {contexto_scope}')
    if filtros_desc:
        meta_lineas.append(f'Filtros: {filtros_desc}')
    meta_lineas.append(f'Generado: {_fmt_datetime(datetime.now())}')

    for linea in meta_lineas:
        ws.merge_cells(start_row=fila_meta, start_column=1, end_row=fila_meta, end_column=len(COLUMNAS))
        c = ws.cell(row=fila_meta, column=1, value=linea)
        c.font = FONT_META
        c.alignment = ALIGN_LEFT
        fila_meta += 1

    # Fila vacía separadora
    fila_meta += 1

    # -------- 2) Header de columnas --------
    fila_header = fila_meta
    for idx, (nombre, ancho) in enumerate(COLUMNAS, start=1):
        c = ws.cell(row=fila_header, column=idx, value=nombre)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDE_FINO
        ws.column_dimensions[get_column_letter(idx)].width = ancho
    ws.row_dimensions[fila_header].height = 24

    # -------- 3) Filas de datos --------
    fila = fila_header + 1
    total_horas = Decimal('0')
    horas_por_estado = {'pendiente': Decimal('0'), 'aprobada': Decimal('0'), 'rechazada': Decimal('0')}
    horas_por_tipo = {}
    n_registros = 0

    # Necesitamos el mapa de tipos para labels legibles
    from apps.employees.models import NovedadNomina
    tipo_labels = dict(NovedadNomina.TIPO_CHOICES)

    for n in novedades:
        n_registros += 1
        total_horas += n.total_horas or Decimal('0')
        horas_por_estado[n.estado_aprobacion] = horas_por_estado.get(n.estado_aprobacion, Decimal('0')) + (n.total_horas or Decimal('0'))
        horas_por_tipo[n.tipo] = horas_por_tipo.get(n.tipo, Decimal('0')) + (n.total_horas or Decimal('0'))

        registrado_por = n.registrado_por.nombre_completo if n.registrado_por_id else ''
        aprobado_por = n.aprobado_por_rrhh.get_full_name() if n.aprobado_por_rrhh_id else ''
        if aprobado_por.strip() == '' and n.aprobado_por_rrhh_id:
            aprobado_por = n.aprobado_por_rrhh.username

        valores = [
            _fmt_fecha(n.fecha),
            n.empleado.nombre_completo,
            n.empleado.numero_documento,
            tipo_labels.get(n.tipo, n.tipo),
            _fmt_hora(n.hora_inicio),
            _fmt_hora(n.hora_fin),
            float(n.total_horas or 0),
            n.motivo or '',
            n.observaciones or '',
            registrado_por,
            n.get_estado_aprobacion_display(),
            aprobado_por,
            _fmt_datetime(n.fecha_aprobacion),
            n.motivo_rechazo or '',
        ]

        fill = _get_fill_por_estado(n.estado_aprobacion)
        for col, val in enumerate(valores, start=1):
            c = ws.cell(row=fila, column=col, value=val)
            c.font = FONT_CELDA
            c.border = BORDE_FINO
            if col in (5, 6, 11, 13):  # horas, estado, fecha_aprobación
                c.alignment = ALIGN_CENTER
            elif col == 7:  # total horas
                c.alignment = ALIGN_RIGHT
                c.number_format = '0.00'
            else:
                c.alignment = ALIGN_LEFT
            if fill and col == 11:  # solo la columna estado con color de fondo
                c.fill = fill
        fila += 1

    # -------- 4) Totales al final --------
    fila_totales = fila + 1
    ws.cell(row=fila_totales, column=1, value='RESUMEN').font = FONT_TOTAL
    ws.merge_cells(start_row=fila_totales, start_column=1, end_row=fila_totales, end_column=len(COLUMNAS))
    ws.cell(row=fila_totales, column=1).fill = FILL_TOTAL
    ws.cell(row=fila_totales, column=1).alignment = ALIGN_LEFT

    fila_totales += 1

    def _linea_resumen(etiqueta, valor):
        nonlocal fila_totales
        ws.cell(row=fila_totales, column=1, value=etiqueta).font = FONT_TOTAL
        ws.merge_cells(start_row=fila_totales, start_column=1, end_row=fila_totales, end_column=6)
        c = ws.cell(row=fila_totales, column=7, value=valor)
        c.font = FONT_TOTAL
        c.alignment = ALIGN_RIGHT
        c.number_format = '0.00'
        fila_totales += 1

    _linea_resumen(f'Total de registros', n_registros)
    _linea_resumen('Total horas', float(total_horas))
    _linea_resumen('Horas pendientes de aprobación', float(horas_por_estado.get('pendiente', 0)))
    _linea_resumen('Horas aprobadas', float(horas_por_estado.get('aprobada', 0)))
    _linea_resumen('Horas rechazadas', float(horas_por_estado.get('rechazada', 0)))

    # Desglose por tipo (si hay más de un tipo con datos)
    tipos_con_datos = [(t, h) for t, h in horas_por_tipo.items() if h > 0]
    if len(tipos_con_datos) > 1:
        fila_totales += 1
        ws.cell(row=fila_totales, column=1, value='Horas por tipo:').font = FONT_TOTAL
        fila_totales += 1
        for tipo, horas in sorted(tipos_con_datos, key=lambda x: -x[1]):
            ws.cell(row=fila_totales, column=2, value=tipo_labels.get(tipo, tipo)).font = FONT_CELDA
            c = ws.cell(row=fila_totales, column=7, value=float(horas))
            c.font = FONT_CELDA
            c.alignment = ALIGN_RIGHT
            c.number_format = '0.00'
            fila_totales += 1

    # Freeze panes para dejar visible el header al scrollear
    ws.freeze_panes = f'A{fila_header + 1}'

    # Ajustar altura de filas para que el wrap se vea bien
    for r in range(fila_header + 1, fila):
        ws.row_dimensions[r].height = None  # auto

    # -------- 5) Serializar a bytes --------
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def nombre_archivo_novedades(prefijo='novedades', fecha_desde=None, fecha_hasta=None):
    """Retorna nombre de archivo tipo 'novedades_2026-07-01_al_2026-07-31.xlsx'."""
    from datetime import date as _date
    if fecha_desde and fecha_hasta:
        return f'{prefijo}_{fecha_desde.isoformat()}_al_{fecha_hasta.isoformat()}.xlsx'
    return f'{prefijo}_{_date.today().isoformat()}.xlsx'
