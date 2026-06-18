# =============================================================================
# apps/employees/admin.py
# =============================================================================

from django.contrib import admin
from django.utils.html import format_html
from django.urls import reverse
from django.utils.safestring import mark_safe
from django.db.models import Q
from datetime import date

from .models import (
    TipoDocumento, Escolaridad, EstadoEmpleado,
    Empleado, HistorialCargo,
    # Familia
    Familiar, DocumentoFamiliar,
    # Marketplace
    Categoria, Producto, Venta, Subasta, PujaSubasta, Regalo,
    # Messaging
    Conversacion, Mensaje, LecturaConversacion,
    # Feed/Publicaciones
    Publicacion, Comentario,
    # Polla Mundial
    PartidoMundial, PrediccionMundial
)

# Registro del modelo TipoDocumento en el admin de Django
@admin.register(TipoDocumento)
class TipoDocumentoAdmin(admin.ModelAdmin):
    list_display = ('codigo', 'nombre', 'longitud_minima', 'longitud_maxima', 'activo')
    list_filter = ('activo', 'requiere_numero')
    search_fields = ('codigo', 'nombre')
    ordering = ('codigo',)

# Registro del modelo Escolaridad en el admin de Django
@admin.register(Escolaridad)
class EscolaridadAdmin(admin.ModelAdmin):
    list_display = ('orden', 'codigo', 'nivel')
    ordering = ('orden',)

# Registro del modelo EstadoEmpleado en el admin de Django
@admin.register(EstadoEmpleado)
class EstadoEmpleadoAdmin(admin.ModelAdmin):
    list_display = ('codigo', 'nombre', 'permite_acceso_sistema')
    list_filter = ('permite_acceso_sistema',)
    ordering = ('codigo',)

# Inline para historial de cargos en el formulario de empleado
class HistorialCargoInline(admin.TabularInline):
    model = HistorialCargo
    fk_name = 'empleado'  # Especificar cuál ForeignKey usar (hay dos: empleado y jefe_directo)
    extra = 1
    fields = ('cargo', 'jefe_directo', 'fecha_inicio', 'fecha_fin', 'activo', 'motivo_cambio')
    readonly_fields = ('fecha_creacion', 'creado_por')
    ordering = ('-fecha_inicio',)

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('cargo__area', 'jefe_directo')

# Registro del modelo Empleado en el admin de Django
@admin.register(Empleado)
class EmpleadoAdmin(admin.ModelAdmin):
    list_display = (
        'nombre_completo', 'numero_documento', 'get_cargo_actual', 
        'get_area_actual', 'estado', 'fecha_ingreso', 'get_estado_badge'
    )
    list_filter = (
        'estado', 'tipo_documento', 'sede', 'fecha_ingreso',
        'historialcargo__cargo__area', 'escolaridad'
    )
    search_fields = (
        'nombres', 'apellidos', 'numero_documento', 
        'correo_electronico', 'telefono_contacto'
    )
    readonly_fields = (
        'id', 'fecha_creacion', 'fecha_actualizacion', 'creado_por',
        'get_antiguedad', 'get_cargo_actual', 'get_area_actual',
    )
    exclude = ('usuario',)  # Excluir campo usuario del formulario
    
    # Agrupación de campos en el formulario de edición
    fieldsets = (
        ('Información Básica', {
            'fields': (
                'tipo_documento', 'numero_documento', 'nombres', 'apellidos',
                'telefono_contacto', 'fecha_ingreso', 'sede', 'estado'
            )
        }),
        ('Información Personal', {
            'fields': (
                'fecha_nacimiento', 'ciudad_nacimiento', 'escolaridad',
                'contacto_emergencia_nombre', 'contacto_emergencia_telefono',
                'correo_electronico'
            ),
            'classes': ('collapse',)
        }),
        ('Información del Sistema', {
            'fields': (
                'id', 'fecha_creacion', 'fecha_actualizacion', 
                'creado_por', 'get_antiguedad', 'get_cargo_actual', 'get_area_actual'
            ),
            'classes': ('collapse',),
            'description': 'El usuario se crea automáticamente al guardar el empleado .'
        }),
    )
    
    inlines = [HistorialCargoInline]
    list_per_page = 25
    list_max_show_all = 100
    date_hierarchy = 'fecha_ingreso'
    ordering = ('apellidos', 'nombres')
    actions = ['marcar_como_activo', 'marcar_como_inactivo', 'export_to_excel']
    
    def get_queryset(self, request):
        """Optimizar consultas"""
        return super().get_queryset(request).select_related(
            'tipo_documento', 'estado', 'sede', 'escolaridad', 'usuario', 'creado_por'
        ).prefetch_related('historialcargo_set__cargo__area')
    
    def get_cargo_actual(self, obj):
        """Obtener cargo actual del empleado"""
        try:
            historial = obj.historialcargo_set.filter(activo=True).first()
            if historial:
                return historial.cargo.nombre
            return "Sin cargo asignado"
        except:
            return "Sin cargo asignado"
    get_cargo_actual.short_description = 'Cargo Actual'
    get_cargo_actual.admin_order_field = 'historialcargo__cargo__nombre'
    
    def get_area_actual(self, obj):
        """Obtener área actual del empleado"""
        try:
            historial = obj.historialcargo_set.filter(activo=True).first()
            if historial:
                return historial.cargo.area.nombre
            return "Sin área asignada"
        except:
            return "Sin área asignada"
    get_area_actual.short_description = 'Área Actual'
    get_area_actual.admin_order_field = 'historialcargo__cargo__area__nombre'
    
    def get_estado_badge(self, obj):
        """Badge colorido para el estado"""
        colors = {
            '999': '#28a745',
            'p-prue': '#ffc107',
            'INACTIVO': '#dc3545',
            'VACACIONES': '#17a2b8',
            'LICENCIA': '#fd7e14',
            'SUSPENDIDO': '#6f42c1',
            'RETIRADO': '#343a40',
        }
        color = colors.get(obj.estado.codigo, '#6c757d')
        
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 8px; '
            'border-radius: 12px; font-size: 11px; font-weight: bold;">{}</span>',
            color, obj.estado.nombre
        )
    get_estado_badge.short_description = 'Estado'
    get_estado_badge.admin_order_field = 'estado__nombre'
    
    def get_antiguedad(self, obj):
        """Calcular antigüedad en la empresa"""
        if obj.fecha_ingreso:
            delta = date.today() - obj.fecha_ingreso
            years = delta.days // 365
            months = (delta.days % 365) // 30
            
            if years > 0:
                return f"{years} año{'s' if years > 1 else ''}, {months} mes{'es' if months != 1 else ''}"
            elif months > 0:
                return f"{months} mes{'es' if months != 1 else ''}"
            else:
                return f"{delta.days} días"
        return "No disponible"
    get_antiguedad.short_description = 'Antigüedad'
    
    def save_model(self, request, obj, form, change):
        """Configurar usuario creador y generar credenciales automáticamente"""
        if not change:  # Solo en creación
            obj.creado_por = request.user
            
            # Generar usuario del sistema automáticamente
            if not obj.usuario:
                obj.usuario = self.crear_usuario_automatico(request, obj)
        
        super().save_model(request, obj, form, change)
    
    def save_formset(self, request, form, formset, change):
        """Guardar el formset de historial de cargos con usuario creador"""
        instances = formset.save(commit=False)
        
        for instance in instances:
            if isinstance(instance, HistorialCargo):
                if not instance.creado_por_id:
                    instance.creado_por = request.user
                instance.save()
        
        formset.save_m2m()
    
    def crear_usuario_automatico(self, request, empleado):
        """Mostrar mensaje si el usuario ya existe, pero no crear usuario aquí (solo lo hace el signal)."""
        from django.contrib import messages
        if empleado.usuario:
            messages.success(
                request,
                f"✅ Usuario ya existe para el empleado:\n"
                f"👤 Usuario: {empleado.usuario.username}\n"
                f" Email: {empleado.usuario.email}\n"
                f"(Comunicar estas credenciales al empleado)"
            )
            return empleado.usuario
        else:
            messages.warning(
                request,
                f"⚠️ No se pudo crear usuario automáticamente."
            )
            return None
    
    # Acciones personalizadas
    def marcar_como_activo(self, request, queryset):
        """Marcar empleados seleccionados como activos"""
        try:
            estado_activo = EstadoEmpleado.objects.get(codigo='999')
            updated = queryset.update(estado=estado_activo)
            self.message_user(
                request, 
                f'{updated} empleado(s) marcado(s) como activo(s).'
            )
        except EstadoEmpleado.DoesNotExist:
            self.message_user(
                request, 
                'Error: No existe el estado ACTIVO en el sistema.',
                level='ERROR'
            )
    marcar_como_activo.short_description = "Marcar como activo"
    
    def marcar_como_inactivo(self, request, queryset):
        """Marcar empleados seleccionados como inactivos"""
        try:
            estado_inactivo = EstadoEmpleado.objects.get(codigo='INACTIVO')
            updated = queryset.update(estado=estado_inactivo)
            self.message_user(
                request, 
                f'{updated} empleado(s) marcado(s) como inactivo(s).'
            )
        except EstadoEmpleado.DoesNotExist:
            self.message_user(
                request, 
                'Error: No existe el estado INACTIVO en el sistema.',
                level='ERROR'
            )
    marcar_como_inactivo.short_description = "Marcar como inactivo"
    
    def export_to_excel(self, request, queryset):
        """Exportar empleados seleccionados a Excel"""
        from django.http import HttpResponse
        import openpyxl
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Empleados'
        
        # Encabezados
        headers = [
            'Documento', 'Nombres', 'Apellidos', 'Email', 'Teléfono',
            'Cargo', 'Área', 'Estado', 'Fecha Ingreso'
        ]
        
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        # Datos
        for row, empleado in enumerate(queryset.select_related('estado').prefetch_related('historialcargo_set__cargo__area'), 2):
            cargo_actual = empleado.historialcargo_set.filter(activo=True).first()
            
            worksheet.cell(row=row, column=1, value=empleado.numero_documento)
            worksheet.cell(row=row, column=2, value=empleado.nombres)
            worksheet.cell(row=row, column=3, value=empleado.apellidos)
            worksheet.cell(row=row, column=4, value=empleado.correo_electronico)
            worksheet.cell(row=row, column=5, value=empleado.telefono_contacto)
            worksheet.cell(row=row, column=6, value=cargo_actual.cargo.nombre if cargo_actual else '')
            worksheet.cell(row=row, column=7, value=cargo_actual.cargo.area.nombre if cargo_actual else '')
            worksheet.cell(row=row, column=8, value=empleado.estado.nombre)
            worksheet.cell(row=row, column=9, value=empleado.fecha_ingreso.strftime('%d/%m/%Y'))
        
        # Respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="empleados_seleccionados.xlsx"'
        
        workbook.save(response)
        return response
    export_to_excel.short_description = "Exportar a Excel"

# Registro del modelo HistorialCargo en el admin de Django
@admin.register(HistorialCargo)
class HistorialCargoAdmin(admin.ModelAdmin):
    list_display = (
        'empleado', 'cargo', 'get_area', 'jefe_directo',
        'fecha_inicio', 'fecha_fin', 'activo', 'get_duracion'
    )
    list_filter = (
        'activo', 'cargo__area', 'fecha_inicio', 'fecha_fin'
    )
    search_fields = (
        'empleado__nombres', 'empleado__apellidos',
        'cargo__nombre', 'cargo__area__nombre',
        'jefe_directo__nombres', 'jefe_directo__apellidos'
    )
    date_hierarchy = 'fecha_inicio'
    ordering = ('-fecha_inicio',)
    
    fieldsets = (
        ('Información del Cargo', {
            'fields': ('empleado', 'cargo', 'jefe_directo', 'fecha_inicio', 'fecha_fin', 'activo')
        }),
        ('Detalles', {
            'fields': ('salario', 'motivo_cambio', 'observaciones'),
            'classes': ('collapse',)
        }),
        ('Información del Sistema', {
            'fields': ('fecha_creacion', 'creado_por'),
            'classes': ('collapse',)
        }),
    )
    
    readonly_fields = ('fecha_creacion', 'creado_por', 'get_duracion')
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related(
            'empleado', 'cargo__area', 'creado_por', 'jefe_directo'
        )
    
    def get_area(self, obj):
        """Obtener área del cargo"""
        return obj.cargo.area.nombre if obj.cargo and obj.cargo.area else ''
    get_area.short_description = 'Área'
    get_area.admin_order_field = 'cargo__area__nombre'
    
    def get_duracion(self, obj):
        """Calcular duración en el cargo"""
        inicio = obj.fecha_inicio
        fin = obj.fecha_fin or date.today()
        
        if inicio:
            delta = fin - inicio
            years = delta.days // 365
            months = (delta.days % 365) // 30
            days = delta.days % 30
            
            partes = []
            if years > 0:
                partes.append(f"{years} año{'s' if years > 1 else ''}")
            if months > 0:
                partes.append(f"{months} mes{'es' if months != 1 else ''}")
            if not partes and days > 0:
                partes.append(f"{days} días")
            
            duracion = ", ".join(partes) if partes else "Menos de un mes"
            
            if obj.activo:
                return f"{duracion} (actual)"
            return duracion
        
        return "No disponible"
    get_duracion.short_description = 'Duración'
    
    def save_model(self, request, obj, form, change):
        if not change:
            obj.creado_por = request.user
        super().save_model(request, obj, form, change)

# ===================== MARKETPLACE - ADMIN =====================

@admin.register(Categoria)
class CategoriaAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'activa', 'fecha_creacion')
    list_filter = ('activa', 'fecha_creacion')
    search_fields = ('nombre', 'descripcion')
    ordering = ('nombre',)


class ProductoInline(admin.TabularInline):
    """Inline para ver productos en subastas/ventas"""
    model = Producto
    extra = 0
    fields = ('titulo', 'tipo', 'precio_inicial', 'estado')
    readonly_fields = ('fecha_creacion', 'creado_por')


@admin.register(Producto)
class ProductoAdmin(admin.ModelAdmin):
    list_display = ('titulo', 'vendedor', 'tipo', 'categoria', 'precio_inicial', 'estado', 'fecha_creacion')
    list_filter = ('tipo', 'estado', 'categoria', 'fecha_creacion')
    search_fields = ('titulo', 'descripcion', 'vendedor__nombre_completo')
    readonly_fields = ('fecha_creacion', 'creado_por', 'fecha_actualizacion', 'id')
    fieldsets = (
        ('Información General', {
            'fields': ('id', 'titulo', 'descripcion', 'categoria', 'vendedor')
        }),
        ('Tipo de Oferta', {
            'fields': ('tipo', 'precio_inicial', 'estado')
        }),
        ('Imagen', {
            'fields': ('imagen',),
            'classes': ('collapse',)
        }),
        ('Visibilidad', {
            'fields': ('visible_para',),
            'classes': ('collapse',),
            'description': 'Dejar vacío para que sea visible para todos'
        }),
        ('Auditoría', {
            'fields': ('fecha_creacion', 'fecha_actualizacion', 'creado_por'),
            'classes': ('collapse',)
        }),
    )
    filter_horizontal = ('visible_para',)
    ordering = ('-fecha_creacion',)


@admin.register(Venta)
class VentaAdmin(admin.ModelAdmin):
    list_display = ('producto', 'vendedor', 'comprador', 'precio_final', 'estado', 'fecha_venta')
    list_filter = ('estado', 'fecha_venta')
    search_fields = ('producto__titulo', 'vendedor__nombre_completo', 'comprador__nombre_completo')
    readonly_fields = ('fecha_creacion', 'creado_por', 'fecha_actualizacion', 'id')
    fieldsets = (
        ('Producto', {
            'fields': ('id', 'producto', 'vendedor', 'comprador')
        }),
        ('Transacción', {
            'fields': ('precio_final', 'estado', 'observaciones')
        }),
        ('Calificaciones', {
            'fields': ('calificacion_vendedor', 'comentario_vendedor', 'calificacion_comprador', 'comentario_comprador'),
            'classes': ('collapse',)
        }),
        ('Fechas', {
            'fields': ('fecha_venta', 'fecha_completada'),
            'classes': ('collapse',)
        }),
        ('Auditoría', {
            'fields': ('fecha_creacion', 'fecha_actualizacion', 'creado_por'),
            'classes': ('collapse',)
        }),
    )
    ordering = ('-fecha_venta',)


class PujaSubastaInline(admin.TabularInline):
    """Inline para ver pujas en una subasta"""
    model = PujaSubasta
    extra = 0
    fields = ('pujador', 'monto', 'es_puja_automatica', 'fecha_creacion')
    readonly_fields = ('fecha_creacion', 'creado_por')
    ordering = ('-fecha_creacion',)


@admin.register(Subasta)
class SubastaAdmin(admin.ModelAdmin):
    list_display = ('producto', 'vendedor', 'precio_actual', 'pujador_actual', 'estado', 'fecha_fin')
    list_filter = ('estado', 'fecha_inicio', 'fecha_fin')
    search_fields = ('producto__titulo', 'vendedor__nombre_completo', 'pujador_actual__nombre_completo')
    readonly_fields = ('fecha_creacion', 'creado_por', 'fecha_actualizacion', 'id')
    fieldsets = (
        ('Producto', {
            'fields': ('id', 'producto', 'vendedor')
        }),
        ('Subasta', {
            'fields': ('precio_inicial', 'precio_actual', 'incremento_minimo', 'estado')
        }),
        ('Pujadores', {
            'fields': ('pujador_actual', 'ganador')
        }),
        ('Fechas', {
            'fields': ('fecha_inicio', 'fecha_fin')
        }),
        ('Auditoría', {
            'fields': ('fecha_creacion', 'fecha_actualizacion', 'creado_por'),
            'classes': ('collapse',)
        }),
    )
    inlines = [PujaSubastaInline]
    ordering = ('-fecha_inicio',)


@admin.register(PujaSubasta)
class PujaSubastaAdmin(admin.ModelAdmin):
    list_display = ('subasta', 'pujador', 'monto', 'es_puja_automatica', 'fecha_creacion')
    list_filter = ('es_puja_automatica', 'fecha_creacion', 'subasta')
    search_fields = ('pujador__nombre_completo', 'subasta__producto__titulo')
    readonly_fields = ('fecha_creacion', 'creado_por', 'fecha_actualizacion', 'id')
    fieldsets = (
        ('Información', {
            'fields': ('id', 'subasta', 'pujador')
        }),
        ('Puja', {
            'fields': ('monto', 'es_puja_automatica', 'monto_maximo')
        }),
        ('Auditoría', {
            'fields': ('fecha_creacion', 'fecha_actualizacion', 'creado_por'),
            'classes': ('collapse',)
        }),
    )
    ordering = ('-fecha_creacion',)


@admin.register(Regalo)
class RegaloAdmin(admin.ModelAdmin):
    list_display = ('producto', 'donante', 'receptor', 'estado', 'fecha_ofrecimiento')
    list_filter = ('estado', 'fecha_ofrecimiento')
    search_fields = ('producto__titulo', 'donante__nombre_completo', 'receptor__nombre_completo')
    readonly_fields = ('fecha_creacion', 'creado_por', 'fecha_actualizacion', 'id')
    fieldsets = (
        ('Información', {
            'fields': ('id', 'producto', 'donante', 'receptor')
        }),
        ('Regalo', {
            'fields': ('estado', 'mensaje')
        }),
        ('Fechas', {
            'fields': ('fecha_ofrecimiento', 'fecha_aceptacion')
        }),
        ('Auditoría', {
            'fields': ('fecha_creacion', 'fecha_actualizacion', 'creado_por'),
            'classes': ('collapse',)
        }),
    )
    ordering = ('-fecha_ofrecimiento',)


# ===================== MESSAGING - ADMIN =====================

class MensajeInline(admin.TabularInline):
    """Inline para ver mensajes en una conversación"""
    model = Mensaje
    extra = 0
    fields = ('remitente', 'contenido', 'leido', 'fecha_creacion')
    readonly_fields = ('remitente', 'contenido', 'leido', 'fecha_creacion', 'creado_por')
    can_delete = False
    ordering = ('fecha_creacion',)

    def has_add_permission(self, request, obj=None):
        """No permitir agregar mensajes desde inline"""
        return False


@admin.register(Conversacion)
class ConversacionAdmin(admin.ModelAdmin):
    list_display = ('id', 'contexto', 'participantes_display', 'fecha_ultima_actividad', 'archivada')
    list_filter = ('contexto', 'archivada', 'fecha_ultima_actividad')
    search_fields = ('participantes__nombre_completo', 'producto_referencia__titulo')
    readonly_fields = ('fecha_creacion', 'creado_por', 'fecha_actualizacion', 'id')
    fieldsets = (
        ('Información', {
            'fields': ('id', 'contexto', 'titulo', 'producto_referencia', 'archivada')
        }),
        ('Participantes', {
            'fields': ('participantes',)
        }),
        ('Auditoría', {
            'fields': ('fecha_creacion', 'fecha_actualizacion', 'creado_por'),
            'classes': ('collapse',)
        }),
    )
    filter_horizontal = ('participantes',)
    inlines = [MensajeInline]
    ordering = ('-fecha_ultima_actividad',)

    def participantes_display(self, obj):
        """Muestra los participantes de forma legible"""
        return ', '.join([p.nombre_completo for p in obj.participantes.all()[:3]])
    participantes_display.short_description = 'Participantes'


@admin.register(Mensaje)
class MensajeAdmin(admin.ModelAdmin):
    list_display = ('remitente', 'conversacion', 'contenido_preview', 'leido', 'fecha_creacion')
    list_filter = ('leido', 'fecha_creacion', 'conversacion')
    search_fields = ('remitente__nombre_completo', 'contenido', 'conversacion__participantes__nombre_completo')
    readonly_fields = ('fecha_creacion', 'creado_por', 'fecha_actualizacion', 'id')
    fieldsets = (
        ('Información', {
            'fields': ('id', 'conversacion', 'remitente')
        }),
        ('Contenido', {
            'fields': ('contenido', 'archivos_adjuntos')
        }),
        ('Lectura', {
            'fields': ('leido', 'fecha_lectura')
        }),
        ('Auditoría', {
            'fields': ('fecha_creacion', 'fecha_actualizacion', 'creado_por'),
            'classes': ('collapse',)
        }),
    )
    ordering = ('-fecha_creacion',)

    def contenido_preview(self, obj):
        """Muestra una vista previa del contenido"""
        preview = obj.contenido[:50] + '...' if len(obj.contenido) > 50 else obj.contenido
        return preview
    contenido_preview.short_description = 'Contenido'


@admin.register(LecturaConversacion)
class LecturaConversacionAdmin(admin.ModelAdmin):
    list_display = ('empleado', 'conversacion', 'fecha_actualizacion')
    list_filter = ('fecha_actualizacion',)
    search_fields = ('empleado__nombre_completo', 'conversacion__participantes__nombre_completo')
    readonly_fields = ('fecha_actualizacion',)
    ordering = ('-fecha_actualizacion',)


# ===================== FEED/PUBLICACIONES =====================

class ComentarioInline(admin.TabularInline):
    """Inline para comentarios en publicaciones"""
    model = Comentario
    extra = 0
    fields = ('autor', 'contenido', 'fecha_creacion')
    readonly_fields = ('autor', 'fecha_creacion')
    can_delete = True


@admin.register(Publicacion)
class PublicacionAdmin(admin.ModelAdmin):
    """Admin para gestionar publicaciones del feed"""
    list_display = (
        'titulo_display', 'autor', 'tipo_publicacion',
        'fecha_creacion', 'tiene_imagen', 'estado_publicacion'
    )
    list_filter = (
        'es_anuncio', 'es_importante', 'fecha_creacion',
        ('fecha_fin', admin.RelatedOnlyFieldListFilter)
    )
    search_fields = ('titulo', 'contenido', 'autor__nombre_completo')
    readonly_fields = ('id', 'fecha_creacion', 'fecha_actualizacion')

    fieldsets = (
        ('Información Básica', {
            'fields': ('id', 'autor', 'titulo', 'contenido', 'imagen')
        }),
        ('Tipo de Publicación', {
            'fields': ('es_anuncio', 'es_importante', 'fecha_fin'),
            'description': 'Marca es_importante para crear un anuncio destacado con fecha de finalización'
        }),
        ('Estilos', {
            'fields': ('estilos',),
            'classes': ('collapse',),
            'description': 'Estilos CSS personalizados (JSON)'
        }),
        ('Fechas', {
            'fields': ('fecha_creacion', 'fecha_actualizacion', 'fecha_eliminacion_automatica'),
            'classes': ('collapse',),
        }),
    )

    inlines = [ComentarioInline]
    ordering = ('-fecha_creacion',)

    def titulo_display(self, obj):
        """Muestra título o primeras 50 caracteres del contenido"""
        titulo = obj.titulo or obj.contenido[:50]
        return titulo[:50] + '...' if len(titulo) > 50 else titulo
    titulo_display.short_description = 'Título/Contenido'

    def tipo_publicacion(self, obj):
        """Muestra el tipo de publicación"""
        if obj.es_anuncio:
            if obj.es_importante:
                return format_html('<span style="color: red; font-weight: bold;">📢 Anuncio Importante</span>')
            return format_html('<span style="color: orange;">📣 Anuncio</span>')
        return format_html('<span style="color: blue;">📝 Publicación</span>')
    tipo_publicacion.short_description = 'Tipo'

    def tiene_imagen(self, obj):
        """Muestra si tiene imagen"""
        return '✅' if obj.imagen else '❌'
    tiene_imagen.short_description = 'Imagen'
    tiene_imagen.admin_order_field = 'imagen'

    def estado_publicacion(self, obj):
        """Muestra el estado de la publicación"""
        from django.utils import timezone
        if obj.es_anuncio and obj.es_importante and obj.fecha_fin:
            if timezone.now() > obj.fecha_fin:
                return format_html('<span style="color: gray;">⏰ Finalizado</span>')
            else:
                dias_restantes = (obj.fecha_fin - timezone.now()).days
                return format_html(f'<span style="color: green;">⏳ {dias_restantes} días</span>')
        return format_html('<span style="color: green;">✓ Activa</span>')
    estado_publicacion.short_description = 'Estado'


@admin.register(Comentario)
class ComentarioAdmin(admin.ModelAdmin):
    """Admin para gestionar comentarios"""
    list_display = ('autor', 'publicacion', 'contenido_preview', 'fecha_creacion')
    list_filter = ('fecha_creacion', 'autor')
    search_fields = ('contenido', 'autor__nombre_completo', 'publicacion__titulo')
    readonly_fields = ('id', 'fecha_creacion', 'fecha_actualizacion')

    fieldsets = (
        ('Información', {
            'fields': ('id', 'publicacion', 'autor', 'contenido')
        }),
        ('Fechas', {
            'fields': ('fecha_creacion', 'fecha_actualizacion'),
            'classes': ('collapse',),
        }),
    )

    ordering = ('-fecha_creacion',)

    def contenido_preview(self, obj):
        """Muestra vista previa del contenido"""
        preview = obj.contenido[:50] + '...' if len(obj.contenido) > 50 else obj.contenido
        return preview
    contenido_preview.short_description = 'Contenido'


# =============================================================================
# POLLA MUNDIALISTA 2026
# =============================================================================

@admin.register(PartidoMundial)
class PartidoMundialAdmin(admin.ModelAdmin):
    """Admin para gestionar partidos del mundial"""
    list_display = ('equipo_local', 'equipo_visitante', 'fase', 'fecha_hora', 'resultado_display', 'finalizado', 'activo')
    list_filter = ('fase', 'finalizado', 'activo', 'fecha_hora')
    search_fields = ('equipo_local', 'equipo_visitante', 'ciudad', 'estadio')
    readonly_fields = ('fecha_creacion', 'api_id')

    fieldsets = (
        ('Equipos', {
            'fields': ('equipo_local', 'equipo_visitante', 'bandera_local', 'bandera_visitante')
        }),
        ('Información del Partido', {
            'fields': ('fecha_hora', 'fase', 'grupo', 'estadio', 'ciudad')
        }),
        ('Resultado', {
            'fields': ('goles_local', 'goles_visitante', 'finalizado')
        }),
        ('Control', {
            'fields': ('activo', 'api_id', 'fecha_creacion')
        }),
    )

    ordering = ('-fecha_hora',)
    list_per_page = 50

    def resultado_display(self, obj):
        """Muestra el resultado del partido si finalizó"""
        if obj.finalizado and obj.goles_local is not None and obj.goles_visitante is not None:
            return f"{obj.goles_local} - {obj.goles_visitante}"
        return "-"
    resultado_display.short_description = 'Resultado'

    actions = ['marcar_como_finalizado', 'activar_predicciones', 'desactivar_predicciones']

    def marcar_como_finalizado(self, request, queryset):
        """Acción para marcar partidos como finalizados"""
        count = queryset.update(finalizado=True)
        self.message_user(request, f'{count} partido(s) marcado(s) como finalizado(s)')
    marcar_como_finalizado.short_description = 'Marcar como finalizado'

    def activar_predicciones(self, request, queryset):
        """Acción para activar predicciones"""
        count = queryset.update(activo=True)
        self.message_user(request, f'Predicciones activadas para {count} partido(s)')
    activar_predicciones.short_description = 'Activar predicciones'

    def desactivar_predicciones(self, request, queryset):
        """Acción para desactivar predicciones"""
        count = queryset.update(activo=False)
        self.message_user(request, f'Predicciones desactivadas para {count} partido(s)')
    desactivar_predicciones.short_description = 'Desactivar predicciones'


@admin.register(PrediccionMundial)
class PrediccionMundialAdmin(admin.ModelAdmin):
    """Admin para gestionar predicciones de empleados"""
    list_display = ('empleado', 'partido_display', 'prediccion_display', 'puntos_ganados', 'fecha_prediccion')
    list_filter = ('fecha_prediccion', 'partido__fase', 'puntos_ganados')
    search_fields = ('empleado__nombre_completo', 'partido__equipo_local', 'partido__equipo_visitante')
    readonly_fields = ('fecha_prediccion', 'fecha_actualizacion', 'puntos_ganados')

    fieldsets = (
        ('Información', {
            'fields': ('empleado', 'partido')
        }),
        ('Predicción', {
            'fields': ('goles_local_prediccion', 'goles_visitante_prediccion')
        }),
        ('Resultado', {
            'fields': ('puntos_ganados', 'fecha_prediccion', 'fecha_actualizacion')
        }),
    )

    ordering = ('-fecha_prediccion',)
    list_per_page = 100

    def partido_display(self, obj):
        """Muestra información del partido"""
        return f"{obj.partido.equipo_local} vs {obj.partido.equipo_visitante}"
    partido_display.short_description = 'Partido'

    def prediccion_display(self, obj):
        """Muestra la predicción del empleado"""
        return f"{obj.goles_local_prediccion} - {obj.goles_visitante_prediccion}"
    prediccion_display.short_description = 'Predicción'

    actions = ['recalcular_puntos']

    def recalcular_puntos(self, request, queryset):
        """Acción para recalcular puntos de predicciones"""
        count = 0
        for prediccion in queryset:
            if prediccion.partido.finalizado:
                prediccion.calcular_puntos()
                prediccion.save()
                count += 1
        self.message_user(request, f'Puntos recalculados para {count} predicción(es)')
    recalcular_puntos.short_description = 'Recalcular puntos'


# Personalización del sitio admin
admin.site.site_header = "RRHH Pro - Administración"
admin.site.site_title = "RRHH Pro Admin"
admin.site.index_title = "Panel de Administración"


# =============================================================================
# Familia
# =============================================================================

class DocumentoFamiliarInline(admin.TabularInline):
    model = DocumentoFamiliar
    extra = 0
    fields = ('tipo', 'descripcion', 'archivo', 'fecha_vencimiento', 'fecha_subida')
    readonly_fields = ('fecha_subida',)


@admin.register(Familiar)
class FamiliarAdmin(admin.ModelAdmin):
    list_display = ('nombre_completo', 'tipo', 'empleado', 'edad', 'eps', 'activo', 'fecha_creacion')
    list_filter = ('tipo', 'activo', 'convive', 'dependiente_economico')
    search_fields = ('nombres', 'apellidos', 'numero_documento', 'empleado__nombres', 'empleado__apellidos', 'empleado__numero_documento')
    autocomplete_fields = ('empleado', 'tipo_documento', 'creado_por')
    inlines = [DocumentoFamiliarInline]
    readonly_fields = ('fecha_creacion', 'fecha_actualizacion')


@admin.register(DocumentoFamiliar)
class DocumentoFamiliarAdmin(admin.ModelAdmin):
    list_display = ('familiar', 'tipo', 'descripcion', 'fecha_vencimiento', 'fecha_subida')
    list_filter = ('tipo',)
    search_fields = ('familiar__nombres', 'familiar__apellidos', 'descripcion')
    autocomplete_fields = ('familiar',)
    readonly_fields = ('fecha_subida',)


