# =============================================================================
# apps/employees/exports.py - FUNCIONES DE EXPORTACIÓN SOLAMENTE
# =============================================================================

from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

# =============================================================================
# REEMPLAZAR en apps/employees/exports.py - VERSIÓN COMPLETA
# =============================================================================

def export_empleados_excel(empleados):
    """Exportar empleados a Excel con TODOS los campos"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        print(f"🔍 DEBUG: Iniciando exportación de {len(empleados)} empleados")
        
        # Crear workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Empleados'
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ENCABEZADOS COMPLETOS
        headers = [
            'Documento', 'Nombres', 'Apellidos', 'Email', 'Teléfono',
            'Cargo', 'Área', 'Sede', 'Jefe Directo', 'Estado', 'Escolaridad', 'Fecha Ingreso'
        ]
        
        # Escribir encabezados
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Escribir datos COMPLETOS
        for row_num, empleado in enumerate(empleados, 2):
            try:
                # Obtener cargo actual
                cargo_actual = empleado.historialcargo_set.filter(activo=True).first()
                
                # Obtener jefe directo
                jefe_directo_nombre = 'Sin asignar'
                if cargo_actual and cargo_actual.jefe_directo:
                    jefe_directo_nombre = cargo_actual.jefe_directo.nombre_completo

                # TODOS LOS DATOS
                data = [
                    empleado.numero_documento,                                    # Columna 1: Documento
                    empleado.nombres,                                            # Columna 2: Nombres
                    empleado.apellidos,                                          # Columna 3: Apellidos
                    empleado.correo_electronico or '',                          # Columna 4: Email
                    empleado.telefono_contacto,                                 # Columna 5: Teléfono
                    cargo_actual.cargo.nombre if cargo_actual else 'Sin cargo', # Columna 6: Cargo
                    cargo_actual.cargo.area.nombre if cargo_actual else 'Sin área', # Columna 7: Área
                    empleado.sede.nombre,                                        # Columna 8: Sede
                    jefe_directo_nombre,                                         # Columna 9: Jefe Directo
                    empleado.estado.nombre,                                      # Columna 10: Estado
                    empleado.escolaridad.nivel if empleado.escolaridad else 'No especificado', # Columna 11: Escolaridad
                    empleado.fecha_ingreso,                                      # Columna 12: Fecha Ingreso
                ]
                
                # Escribir cada dato
                for col, value in enumerate(data, 1):
                    cell = worksheet.cell(row=row_num, column=col, value=value)
                    cell.border = border
                    
                    # Formato especial para fecha (columna 11)
                    if col == 11 and hasattr(value, 'strftime'):
                        cell.number_format = 'DD/MM/YYYY'
                        
            except Exception as e:
                print(f"❌ Error procesando empleado {empleado.id}: {e}")
                continue
        
        # Ajustar ancho de columnas automáticamente
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            for row in worksheet[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            
            # Ancho mínimo y máximo
            adjusted_width = min(max(max_length + 2, 12), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Metadata del archivo
        total_rows = len(empleados) + 2
        worksheet[f'A{total_rows + 1}'] = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"
        worksheet[f'A{total_rows + 2}'] = f"Total empleados: {len(empleados)}"
        
        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f'empleados_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        print(f"✅ DEBUG: Headers configurados - {filename}")
        
        # Guardar workbook
        workbook.save(response)
        print(f"✅ DEBUG: Archivo generado exitosamente con {len(headers)} columnas")
        
        return response
        
    except ImportError as e:
        error_msg = f"Error: openpyxl no está instalado - {e}"
        print(f"❌ {error_msg}")
        return HttpResponse(error_msg, content_type="text/plain", status=500)
        
    except Exception as e:
        error_msg = f"Error general en exportación: {e}"
        print(f"❌ {error_msg}")
        return HttpResponse(error_msg, content_type="text/plain", status=500)


def export_empleados_pdf(empleados):
    """Exportar empleados a PDF con todos los campos principales"""
    try:
        print(f"🔍 DEBUG: Iniciando exportación PDF de {len(empleados)} empleados")
        
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        
        # Crear respuesta HTTP
        response = HttpResponse(content_type='application/pdf')
        filename = f'empleados_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Crear documento en orientación horizontal para más columnas
        doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=0.5*inch)
        elements = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.darkblue,
            alignment=1,  # Centrado
            spaceAfter=20
        )
        
        # Título
        title = Paragraph("Listado de Empleados", title_style)
        elements.append(title)
        
        # Información del reporte
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            alignment=1
        )
        
        fecha_generacion = timezone.now().strftime('%d/%m/%Y %H:%M:%S')
        info = Paragraph(f"Generado el: {fecha_generacion} | Total empleados: {len(empleados)}", info_style)
        elements.append(info)
        elements.append(Spacer(1, 20))
        
        # Preparar datos para la tabla (más columnas en horizontal)
        data = [['Documento', 'Nombre Completo', 'Email', 'Cargo', 'Área', 'Jefe Directo', 'Estado', 'Fecha Ingreso']]

        for empleado in empleados:
            try:
                cargo_actual = empleado.historialcargo_set.filter(activo=True).first()

                # Obtener jefe directo
                jefe_directo_nombre = 'Sin asignar'
                if cargo_actual and cargo_actual.jefe_directo:
                    jefe_directo_nombre = cargo_actual.jefe_directo.nombre_completo

                row = [
                    empleado.numero_documento,
                    f"{empleado.nombres} {empleado.apellidos}",
                    empleado.correo_electronico or '',
                    cargo_actual.cargo.nombre if cargo_actual else 'Sin cargo',
                    cargo_actual.cargo.area.nombre if cargo_actual else 'Sin área',
                    jefe_directo_nombre,
                    empleado.estado.nombre,
                    empleado.fecha_ingreso.strftime('%d/%m/%Y')
                ]
                data.append(row)
                
            except Exception as e:
                print(f"❌ Error procesando empleado {empleado.id} para PDF: {e}")
                continue
        
        # Crear tabla con anchos específicos (8 columnas ahora)
        table = Table(data, colWidths=[0.9*inch, 1.8*inch, 1.5*inch, 1.3*inch, 1.2*inch, 1.5*inch, 0.9*inch, 1*inch])
        
        # Estilo de tabla
        table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Contenido
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Bordes
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            
            # Alternar colores de filas
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        print(f"✅ DEBUG: PDF generado exitosamente")
        return response
        
    except ImportError as e:
        error_msg = f"Error: reportlab no está instalado - {e}"
        print(f"❌ {error_msg}")
        return HttpResponse(error_msg, content_type="text/plain", status=500)
        
    except Exception as e:
        error_msg = f"Error en exportación PDF: {e}"
        print(f"❌ {error_msg}")
        return HttpResponse(error_msg, content_type="text/plain", status=500)


def export_empleados_csv(empleados):
    """Exportar empleados a CSV con todos los campos"""
    import csv
    
    print(f"🔍 DEBUG: Iniciando exportación CSV de {len(empleados)} empleados")
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    filename = f'empleados_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # BOM para UTF-8 (para Excel)
    response.write('\ufeff')
    
    writer = csv.writer(response)
    
    # Encabezados completos
    writer.writerow([
        'Documento', 'Nombres', 'Apellidos', 'Email', 'Teléfono',
        'Cargo', 'Área', 'Sede', 'Jefe Directo', 'Estado', 'Escolaridad', 'Fecha Ingreso'
    ])

    # Datos completos
    for empleado in empleados:
        try:
            cargo_actual = empleado.historialcargo_set.filter(activo=True).first()

            # Obtener jefe directo
            jefe_directo_nombre = 'Sin asignar'
            if cargo_actual and cargo_actual.jefe_directo:
                jefe_directo_nombre = cargo_actual.jefe_directo.nombre_completo

            writer.writerow([
                empleado.numero_documento,
                empleado.nombres,
                empleado.apellidos,
                empleado.correo_electronico or '',
                empleado.telefono_contacto,
                cargo_actual.cargo.nombre if cargo_actual else 'Sin cargo',
                cargo_actual.cargo.area.nombre if cargo_actual else 'Sin área',
                empleado.sede.nombre,
                jefe_directo_nombre,
                empleado.estado.nombre,
                empleado.escolaridad.nivel if empleado.escolaridad else 'No especificado',
                empleado.fecha_ingreso.strftime('%d/%m/%Y')
            ])
        except Exception as e:
            print(f"❌ Error procesando empleado {empleado.id} para CSV: {e}")
            continue
    
    print(f"✅ DEBUG: CSV generado exitosamente")
    return response

def export_empleado_perfil_pdf(empleado):
    """Exportar perfil completo de un empleado a PDF"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.units import inch
        
        print(f"🔍 DEBUG: Generando perfil PDF para {empleado.nombre_completo}")
        
        # Crear respuesta HTTP
        response = HttpResponse(content_type='application/pdf')
        filename = f'perfil_{empleado.nombres}_{empleado.apellidos}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Crear documento
        doc = SimpleDocTemplate(response, pagesize=A4, topMargin=0.7*inch, bottomMargin=0.7*inch)
        story = []
        
        # Estilos
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.darkblue,
            alignment=1,  # Centrado
            spaceAfter=30
        )
        
        subtitle_style = ParagraphStyle(
            'SubTitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.darkblue,
            spaceAfter=15,
            spaceBefore=20
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=10
        )
        
        # TÍTULO PRINCIPAL
        title = Paragraph(f"PERFIL DE EMPLEADO<br/>{empleado.nombre_completo.upper()}", title_style)
        story.append(title)
        
        # Fecha de generación
        fecha_gen = Paragraph(f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}", 
                             ParagraphStyle('DateStyle', parent=styles['Normal'], fontSize=9, 
                                          textColor=colors.grey, alignment=1))
        story.append(fecha_gen)
        story.append(Spacer(1, 20))
        
        # INFORMACIÓN BÁSICA
        story.append(Paragraph("INFORMACIÓN BÁSICA", subtitle_style))
        
        # Obtener cargo actual
        cargo_actual = empleado.historialcargo_set.filter(activo=True).first()
        
        info_basica = [
            ['Campo', 'Información'],
            ['Tipo de Documento', empleado.tipo_documento.nombre],
            ['Número de Documento', empleado.numero_documento],
            ['Nombres Completos', empleado.nombres],
            ['Apellidos Completos', empleado.apellidos],
            ['Email', empleado.correo_electronico or 'No registrado'],
            ['Teléfono', empleado.telefono_contacto],
            ['Fecha de Ingreso', empleado.fecha_ingreso.strftime('%d/%m/%Y')],
            ['Sede', empleado.sede.nombre],
            ['Estado Actual', empleado.estado.nombre],
            ['Cargo Actual', cargo_actual.cargo.nombre if cargo_actual else 'Sin cargo asignado'],
            ['Área', cargo_actual.cargo.area.nombre if cargo_actual else 'Sin área asignada'],
        ]
        
        table_info = Table(info_basica, colWidths=[2*inch, 4*inch])
        table_info.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        story.append(table_info)
        story.append(Spacer(1, 20))
        
        # INFORMACIÓN PERSONAL
        if empleado.fecha_nacimiento or empleado.ciudad_nacimiento or empleado.escolaridad:
            story.append(Paragraph("INFORMACIÓN PERSONAL", subtitle_style))
            
            info_personal = [['Campo', 'Información']]
            
            if empleado.fecha_nacimiento:
                edad = (timezone.now().date() - empleado.fecha_nacimiento).days // 365
                info_personal.append(['Fecha de Nacimiento', f"{empleado.fecha_nacimiento.strftime('%d/%m/%Y')} ({edad} años)"])
            
            if empleado.ciudad_nacimiento:
                info_personal.append(['Ciudad de Nacimiento', empleado.ciudad_nacimiento])
            
            if empleado.escolaridad:
                info_personal.append(['Escolaridad', empleado.escolaridad.nivel])
            
            if empleado.contacto_emergencia_nombre:
                info_personal.append(['Contacto de Emergencia', empleado.contacto_emergencia_nombre])
            
            if empleado.contacto_emergencia_telefono:
                info_personal.append(['Teléfono de Emergencia', empleado.contacto_emergencia_telefono])
            
            table_personal = Table(info_personal, colWidths=[2*inch, 4*inch])
            table_personal.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            story.append(table_personal)
            story.append(Spacer(1, 20))
        
        # HISTORIAL DE CARGOS
        historial_cargos = empleado.historialcargo_set.all().order_by('-fecha_inicio')
        if historial_cargos.exists():
            story.append(Paragraph("HISTORIAL DE CARGOS", subtitle_style))
            
            historial_data = [['Cargo', 'Área', 'Fecha Inicio', 'Fecha Fin', 'Estado', 'Motivo']]
            
            for hist in historial_cargos:
                historial_data.append([
                    hist.cargo.nombre,
                    hist.cargo.area.nombre,
                    hist.fecha_inicio.strftime('%d/%m/%Y'),
                    hist.fecha_fin.strftime('%d/%m/%Y') if hist.fecha_fin else 'Actual',
                    'Activo' if hist.activo else 'Finalizado',
                    hist.motivo_cambio or '-'
                ])
            
            table_historial = Table(historial_data, colWidths=[1.2*inch, 1.2*inch, 1*inch, 1*inch, 0.8*inch, 1.8*inch])
            table_historial.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkorange),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            story.append(table_historial)
        
        # Generar PDF
        doc.build(story)
        
        print(f"✅ DEBUG: Perfil PDF generado exitosamente")
        return response
        
    except ImportError as e:
        error_msg = f"Error: reportlab no disponible - {e}"
        print(f"❌ {error_msg}")
        return HttpResponse(error_msg, content_type="text/plain", status=500)
        
    except Exception as e:
        error_msg = f"Error generando perfil PDF: {e}"
        print(f"❌ {error_msg}")
        return HttpResponse(error_msg, content_type="text/plain", status=500)


def export_empleado_excel(empleado):
    """Exportar datos básicos de un empleado a Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        print(f"🔍 DEBUG: Generando Excel individual para {empleado.nombre_completo}")
        
        # Crear workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = f'{empleado.nombres} {empleado.apellidos}'
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        data_font = Font(size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        worksheet.merge_cells('A1:B1')
        title_cell = worksheet['A1']
        title_cell.value = f"PERFIL DE EMPLEADO: {empleado.nombre_completo.upper()}"
        title_cell.font = Font(bold=True, size=14, color="366092")
        title_cell.alignment = Alignment(horizontal="center")
        
        # Datos del empleado
        cargo_actual = empleado.historialcargo_set.filter(activo=True).first()
        
        datos = [
            ('Campo', 'Información'),
            ('Documento', f"{empleado.tipo_documento.nombre}: {empleado.numero_documento}"),
            ('Nombres', empleado.nombres),
            ('Apellidos', empleado.apellidos),
            ('Email', empleado.correo_electronico or 'No registrado'),
            ('Teléfono', empleado.telefono_contacto),
            ('Cargo Actual', cargo_actual.cargo.nombre if cargo_actual else 'Sin cargo'),
            ('Área', cargo_actual.cargo.area.nombre if cargo_actual else 'Sin área'),
            ('Sede', empleado.sede.nombre),
            ('Estado', empleado.estado.nombre),
            ('Fecha de Ingreso', empleado.fecha_ingreso),
            ('Escolaridad', empleado.escolaridad.nivel if empleado.escolaridad else 'No especificado'),
        ]
        
        if empleado.fecha_nacimiento:
            edad = (timezone.now().date() - empleado.fecha_nacimiento).days // 365
            datos.append(('Fecha de Nacimiento', f"{empleado.fecha_nacimiento} ({edad} años)"))
        
        if empleado.ciudad_nacimiento:
            datos.append(('Ciudad de Nacimiento', empleado.ciudad_nacimiento))
        
        if empleado.contacto_emergencia_nombre:
            datos.append(('Contacto de Emergencia', f"{empleado.contacto_emergencia_nombre} - {empleado.contacto_emergencia_telefono}"))
        
        # Escribir datos
        for row, (campo, valor) in enumerate(datos, 3):
            # Campo
            cell_campo = worksheet.cell(row=row, column=1, value=campo)
            cell_campo.font = header_font if row == 3 else data_font
            cell_campo.fill = header_fill if row == 3 else PatternFill()
            cell_campo.border = border
            
            # Valor
            cell_valor = worksheet.cell(row=row, column=2, value=valor)
            cell_valor.font = header_font if row == 3 else data_font
            cell_valor.fill = header_fill if row == 3 else PatternFill()
            cell_valor.border = border
            
            # Formato fecha
            if isinstance(valor, timezone.datetime.date):
                cell_valor.number_format = 'DD/MM/YYYY'
        
        # Ajustar columnas
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 40
        
        # Metadata
        last_row = len(datos) + 4
        worksheet[f'A{last_row}'] = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f'perfil_{empleado.nombres}_{empleado.apellidos}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        workbook.save(response)
        
        print(f"✅ DEBUG: Excel individual generado exitosamente")
        return response
        
    except ImportError as e:
        error_msg = f"Error: openpyxl no disponible - {e}"
        print(f"❌ {error_msg}")
        return HttpResponse(error_msg, content_type="text/plain", status=500)
        
    except Exception as e:
        error_msg = f"Error generando Excel individual: {e}"
        print(f"❌ {error_msg}")
        return HttpResponse(error_msg, content_type="text/plain", status=500)