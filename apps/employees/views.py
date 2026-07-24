# =============================================================================
# apps/employees/views.py - VISTAS CORREGIDAS DE EMPLEADOS
# =============================================================================

# Importaciones estándar
import json
import logging
from datetime import timedelta, date

# Importaciones de Django
from django import forms
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth import get_user_model
from django.contrib import messages
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, View, TemplateView
from django.urls import reverse_lazy, reverse
from django.db.models import Q, Count, Avg, Prefetch, Sum
from django.core.paginator import Paginator
from django.http import Http404, JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST, require_http_methods
from django.template.loader import render_to_string
from django.utils import timezone
from django.db import transaction, IntegrityError
from django.core.exceptions import ValidationError

# Importaciones locales
from .exports import export_empleados_excel, export_empleados_pdf, export_empleados_csv, export_empleado_perfil_pdf, export_empleado_excel
from .models import (
    Empleado, TipoDocumento, Escolaridad, EstadoEmpleado, HistorialCargo,
    Producto, Venta, Subasta, PujaSubasta, Regalo, Reserva, Conversacion, Mensaje, Categoria,
    Familiar, DocumentoFamiliar,
    SolicitudVacacion,
    AsistenciaDiaria,
)
from .forms import (
    EmpleadoForm, BusquedaEmpleadoForm,
    ProductoForm, VentaForm, SubastaForm, PujaForm, RegaloForm, ConversacionForm, MensajeForm,
    EstadoCivilForm, FamiliarForm, DocumentoFamiliarForm,
    SolicitudVacacionForm,
)
from apps.organizational.models import AreaEmpresa, Cargo, Sede
from apps.training.models import InscripcionCapacitacion
from apps.evaluations.models import AsignacionEvaluacion
from apps.documents.models import DocumentoEmpleado, TipoDocumentoEmpleado
from apps.recognition.models import HistorialPuntos, InsigniaEmpleado


# Configurar logging
logger = logging.getLogger(__name__)
User = get_user_model()


# =============================================================================
# MIXINS REUTILIZABLES
# =============================================================================

class EmpleadoRequiredMixin(LoginRequiredMixin):
    """
    Mixin que requiere que el usuario autenticado tenga un registro de Empleado.
    Si no lo tiene, redirige al dashboard con un mensaje de error.
    Los administradores pueden acceder sin importar si tienen empleado.
    """
    def dispatch(self, request, *args, **kwargs):
        # Los administradores pueden acceder sin restricción
        if request.user.is_staff or request.user.is_superuser:
            logger.info(f"[EmpleadoRequiredMixin] Usuario staff/super accediendo: {request.user.username}")
            return super().dispatch(request, *args, **kwargs)

        try:
            # Intenta acceder al empleado del usuario regular
            empleado = request.user.empleado
            logger.info(f"[EmpleadoRequiredMixin] Usuario con empleado accediendo: {request.user.username} -> {empleado.nombre_completo}")
            return super().dispatch(request, *args, **kwargs)
        except AttributeError as e:
            # No tiene el atributo 'empleado'
            logger.error(f"[EmpleadoRequiredMixin] AttributeError para usuario {request.user.username}: {str(e)}")
            messages.error(
                request,
                'Tu cuenta de usuario no está asociada a un registro de empleado. '
                'Contacta con Recursos Humanos para completar tu perfil.'
            )
            return redirect('core:dashboard')
        except Empleado.DoesNotExist as e:
            # El empleado no existe en la BD
            logger.error(f"[EmpleadoRequiredMixin] Empleado.DoesNotExist para usuario {request.user.username}: {str(e)}")
            messages.error(
                request,
                'Tu cuenta de usuario no está asociada a un registro de empleado. '
                'Contacta con Recursos Humanos para completar tu perfil.'
            )
            return redirect('core:dashboard')


class EmpleadoListView(LoginRequiredMixin, ListView):
    """Vista para listar empleados con filtros y búsqueda"""
    model = Empleado
    template_name = 'employees/empleado_list.html'
    context_object_name = 'empleados'
    paginate_by = 20
    
    def get_queryset(self):
        """Obtener empleados con filtros aplicados"""
        queryset = Empleado.objects.select_related(
            'estado', 'sede', 'tipo_documento', 'usuario', 'escolaridad'
        ).prefetch_related(
            Prefetch(
                'historialcargo_set',
                queryset=HistorialCargo.objects.filter(activo=True).select_related('cargo__area'),
                to_attr='cargo_actual_list'
            )
        )
        
        # Obtener parámetros de filtros
        search = self.request.GET.get('search', '').strip()
        estado = self.request.GET.get('estado')
        area = self.request.GET.get('area')
        cargo = self.request.GET.get('cargo')
        sede = self.request.GET.get('sede')
        
        # Aplicar filtro de búsqueda
        if search:
            queryset = queryset.filter(
                Q(nombres__istartswith=search) |
                Q(apellidos__istartswith=search)
            )
        
        # Aplicar filtros específicos
        if estado:
            queryset = queryset.filter(estado_id=estado)
        
        if sede:
            queryset = queryset.filter(sede_id=sede)
        
        if area:
            queryset = queryset.filter(
                historialcargo__cargo__area_id=area,
                historialcargo__activo=True
            )
        
        if cargo:
            queryset = queryset.filter(
                historialcargo__cargo_id=cargo,
                historialcargo__activo=True
            )
        
        return queryset.distinct().order_by('nombres', 'apellidos')
    
    def get_context_data(self, **kwargs):
        """Agregar contexto adicional"""
        context = super().get_context_data(**kwargs)
        
        # Formulario de búsqueda
        form_data = self.request.GET.copy()
        context['form'] = BusquedaEmpleadoForm(form_data)
        
        # Estadísticas generales
        context.update(self.get_estadisticas())
        
        return context
    
    def get_estadisticas(self):
        """Calcular estadísticas para el dashboard"""
        # Totales generales
        total_empleados = Empleado.objects.count()
        
        # Por estado
        try:
            estado_activo = EstadoEmpleado.objects.get(codigo='999')
            empleados_activos = Empleado.objects.filter(estado=estado_activo).count()
        except EstadoEmpleado.DoesNotExist:
            empleados_activos = 0
        
        try:
            estado_prueba = EstadoEmpleado.objects.get(codigo='p-prue')
            empleados_prueba = Empleado.objects.filter(estado=estado_prueba).count()
        except EstadoEmpleado.DoesNotExist:
            empleados_prueba = 0
        
        # Nuevos empleados este mes
        inicio_mes = timezone.now().replace(day=1).date()
        nuevos_mes = Empleado.objects.filter(fecha_ingreso__gte=inicio_mes).count()
        
        return {
            'total_empleados': total_empleados,
            'empleados_activos': empleados_activos,
            'empleados_prueba': empleados_prueba,
            'nuevos_mes': nuevos_mes,
        }


class EmpleadoDetailView(LoginRequiredMixin, DetailView):
    """Vista de detalle completo del empleado"""
    model = Empleado
    template_name = 'employees/empleado_detail.html'
    context_object_name = 'empleado'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        empleado = self.get_object()
        
        # Historial de cargos
        context['historial_cargos'] = empleado.historialcargo_set.select_related(
            'cargo__area'
        ).order_by('-fecha_inicio')
        
        # Capacitaciones con todas las relaciones necesarias
        try:
            context['capacitaciones'] = InscripcionCapacitacion.objects.filter(
                empleado=empleado
            ).select_related(
                'capacitacion',
                'capacitacion__tipo',
                'inscrito_por',
                'aprobado_por'
            ).prefetch_related(
                'capacitacion__modulocapacitacion_set'
            ).order_by('-fecha_inscripcion')
            
            # Calcular estadísticas de capacitaciones
            capacitaciones = context['capacitaciones']
            context.update({
                'capacitaciones_obligatorias': capacitaciones.filter(
                    capacitacion__tipo__codigo='OBLIGATORIA'
                ).count(),
                'capacitaciones_en_progreso': capacitaciones.filter(
                    estado='en_progreso'
                ).count(),
                'capacitaciones_completadas': capacitaciones.filter(
                    estado='completado'
                ).count()
            })
            
        except Exception as e:
            context['capacitaciones'] = []
            context.update({
                'capacitaciones_obligatorias': 0,
                'capacitaciones_en_progreso': 0,
                'capacitaciones_completadas': 0
            })
        
        # Evaluaciones
        try:
            context['evaluaciones'] = AsignacionEvaluacion.objects.filter(
                empleado_evaluado=empleado
            ).select_related('evaluacion', 'evaluador').order_by('-fecha_asignacion')[:5]

            # Obtener planes de mejora asociados a cada evaluación
            try:
                from apps.evaluations.models import PlanMejoraPredefinido
                planes_dict = {}
                for evaluacion in context['evaluaciones']:
                    plan = PlanMejoraPredefinido.objects.filter(
                        asignacion_evaluacion=evaluacion
                    ).first()
                    if plan:
                        planes_dict[str(evaluacion.id)] = plan
                context['planes_mejora'] = planes_dict
            except Exception as e:
                context['planes_mejora'] = {}
        except:
            context['evaluaciones'] = []
            context['planes_mejora'] = {}
        
        # Estadísticas reales de documentos
        documentos = empleado.documentoempleado_set.all()
        context['documentos_completos'] = documentos.filter(estado_aprobacion='aprobado').count()
        context['documentos_pendientes'] = documentos.filter(estado_aprobacion='pendiente').count()
        
        # Obtener cargo actual
        cargo_actual = empleado.historialcargo_set.filter(activo=True).first()
        context['cargo_actual'] = cargo_actual
        
        # Obtener capacitaciones obligatorias del cargo actual
        if cargo_actual:
            try:
                context['capacitaciones_cargo'] = InscripcionCapacitacion.objects.filter(
                    cargo=cargo_actual.cargo,
                    capacitacion__tipo__codigo='OBLIGATORIA'
                ).exclude(
                    empleado=empleado
                ).select_related('capacitacion')
            except:
                context['capacitaciones_cargo'] = []
        
        # Datos de reconocimientos (solo para administradores)
        if self.request.user.is_staff:
            try:
                from django.db.models import Sum, Count, Q
                from django.db.models.functions import Coalesce
                
                # Calcular puntos totales
                puntos_totales = HistorialPuntos.objects.filter(
                    empleado=empleado,
                    validado=True
                ).aggregate(total=Sum('puntos'))['total'] or 0
                
                # Puntos del mes actual
                inicio_mes = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                puntos_mes = HistorialPuntos.objects.filter(
                    empleado=empleado,
                    validado=True,
                    fecha_obtencion__gte=inicio_mes
                ).aggregate(total=Sum('puntos'))['total'] or 0
                
                # Posición en ranking
                from apps.employees.models import Empleado
                empleados_ranking = Empleado.objects.exclude(estado__codigo='inactivo').annotate(
                    puntos_totales=Coalesce(
                        Sum('historialpuntos__puntos', filter=Q(historialpuntos__validado=True)), 
                        0
                    )
                ).order_by('-puntos_totales')
                
                ranking_posicion = None
                for i, emp in enumerate(empleados_ranking, 1):
                    if emp.id == empleado.id:
                        ranking_posicion = i
                        break
                
                # Número de actividades
                actividades_count = HistorialPuntos.objects.filter(
                    empleado=empleado,
                    validado=True
                ).count()
                
                # Historial reciente de puntos (últimos 30 días)
                hace_30_dias = timezone.now() - timedelta(days=30)
                historial_puntos = HistorialPuntos.objects.filter(
                    empleado=empleado,
                    fecha_obtencion__gte=hace_30_dias
                ).select_related('tipo_actividad', 'validado_por').order_by('-fecha_obtencion')[:10]
                
                # Datos de insignias
                insignias_empleado = InsigniaEmpleado.objects.filter(
                    empleado=empleado
                ).select_related('tipo_insignia', 'otorgado_por').order_by('-fecha_otorgamiento')
                
                insignias_count = insignias_empleado.count()
                
                # Insignias del mes
                insignias_mes = insignias_empleado.filter(
                    fecha_otorgamiento__gte=inicio_mes
                ).count()
                
                # Insignias especiales (manualmente otorgadas)
                insignias_especiales = insignias_empleado.filter(
                    otorgado_automaticamente=False
                ).count()
                
                # Agregar al contexto
                context.update({
                    'puntos_totales': puntos_totales,
                    'puntos_mes': puntos_mes,
                    'ranking_posicion': ranking_posicion,
                    'actividades_count': actividades_count,
                    'historial_puntos': historial_puntos,
                    'insignias_empleado': insignias_empleado,
                    'insignias_count': insignias_count,
                    'insignias_mes': insignias_mes,
                    'insignias_especiales': insignias_especiales,
                })
                
                # Agregar datos al empleado para uso en template
                empleado.puntos_totales = puntos_totales
                empleado.puntos_mes = puntos_mes
                empleado.ranking_posicion = ranking_posicion
                empleado.actividades_count = actividades_count
                empleado.insignias_count = insignias_count
                empleado.insignias_mes = insignias_mes
                empleado.insignias_especiales = insignias_especiales
                
            except ImportError:
                # Módulo de reconocimientos no disponible
                pass
            except Exception as e:
                # Error al obtener datos de reconocimientos
                pass
        
        # Agregar datos necesarios para el modal de cambio de cargo
        # Excluir el cargo actual del empleado de la lista
        cargo_actual_id = None
        if cargo_actual and cargo_actual.cargo:
            cargo_actual_id = cargo_actual.cargo.id

        if cargo_actual_id:
            context['cargos'] = Cargo.objects.filter(activo=True).exclude(id=cargo_actual_id).order_by('nombre')
        else:
            context['cargos'] = Cargo.objects.filter(activo=True).order_by('nombre')

        context['sedes'] = Sede.objects.filter(activa=True).order_by('nombre')

        # === EMPLEADOS A CARGO ===
        # Usar el mismo método que EmpleadoPerfilView
        context.update(self.get_empleados_a_cargo(empleado))

        return context

    def get_empleados_a_cargo(self, empleado):
        """Obtener empleados a cargo según jerarquía organizacional con lógica híbrida"""
        try:
            # Obtener cargo actual del empleado
            cargo_actual = empleado.historialcargo_set.filter(activo=True).first()

            if not cargo_actual or not cargo_actual.cargo:
                return {
                    'empleados_a_cargo': {
                        'total': 0,
                        'activos': 0,
                        'en_prueba': 0,
                        'con_alertas': 0,
                        'lista_empleados': [],
                        'es_jefe': False,
                        'tipo_jefe': None
                    }
                }

            # Determinar tipo de jefe según nombre del cargo
            cargo_nombre = cargo_actual.cargo.nombre.lower()
            tipo_jefe = None

            if 'gerente' in cargo_nombre:
                tipo_jefe = 'gerente'
            elif 'director' in cargo_nombre:
                tipo_jefe = 'director'
            elif 'coordinador' in cargo_nombre:
                tipo_jefe = 'coordinador'
            elif 'supervisor' in cargo_nombre or 'jefe' in cargo_nombre:
                tipo_jefe = 'supervisor'

            if not tipo_jefe:
                # No es un cargo de jefatura
                return {
                    'empleados_a_cargo': {
                        'total': 0,
                        'activos': 0,
                        'en_prueba': 0,
                        'con_alertas': 0,
                        'lista_empleados': [],
                        'es_jefe': False,
                        'tipo_jefe': None
                    }
                }

            # Buscar empleados a cargo usando SOLO jefe_directo
            # Solo mostrar empleados que:
            # 1. Tienen jefe_directo asignado = empleado actual
            # 2. Están ACTIVOS o en PERÍODO DE PRUEBA
            empleados_subordinados = Empleado.objects.filter(
                historialcargo__activo=True,
                historialcargo__jefe_directo=empleado,
                estado__codigo__in=['999', 'p-prue']  # Activos y en período de prueba
            ).distinct()

            # Calcular estadísticas
            total = empleados_subordinados.count()
            activos = empleados_subordinados.filter(estado__codigo='999').count()
            en_prueba = empleados_subordinados.filter(estado__codigo='p-prue').count()

            # Preparar lista de empleados con detalles
            lista_empleados = []
            con_alertas = 0

            for emp in empleados_subordinados.select_related('estado').prefetch_related('historialcargo_set__cargo'):
                # Verificar si tiene alertas
                tiene_alertas = False
                alertas = []

                # Verificar documentos vencidos/pendientes
                try:
                    from apps.documents.models import DocumentoEmpleado
                    docs_vencidos = DocumentoEmpleado.objects.filter(
                        empleado=emp,
                        fecha_vencimiento__lt=date.today()
                    ).count()
                    if docs_vencidos > 0:
                        tiene_alertas = True
                        alertas.append(f"{docs_vencidos} doc. vencidos")
                except ImportError:
                    pass

                # Verificar evaluaciones pendientes
                try:
                    from apps.evaluations.models import AsignacionEvaluacion
                    eval_pendientes = AsignacionEvaluacion.objects.filter(
                        empleado_evaluado=emp,
                        estado__in=['asignada', 'en_progreso'],
                        fecha_vencimiento__lt=date.today()
                    ).count()
                    if eval_pendientes > 0:
                        tiene_alertas = True
                        alertas.append(f"{eval_pendientes} eval. vencidas")
                except ImportError:
                    pass

                if tiene_alertas:
                    con_alertas += 1

                # Obtener cargo actual del subordinado
                cargo_subordinado = emp.historialcargo_set.filter(activo=True).first()

                lista_empleados.append({
                    'empleado': emp,
                    'cargo_actual': cargo_subordinado,
                    'tiene_alertas': tiene_alertas,
                    'alertas': alertas,
                    'dias_empresa': (date.today() - emp.fecha_ingreso).days
                })

            return {
                'empleados_a_cargo': {
                    'total': total,
                    'activos': activos,
                    'en_prueba': en_prueba,
                    'con_alertas': con_alertas,
                    'lista_empleados': lista_empleados,
                    'es_jefe': True,
                    'tipo_jefe': tipo_jefe
                }
            }

        except Exception as e:
            logger.error(f"Error al obtener empleados a cargo: {e}")
            return {
                'empleados_a_cargo': {
                    'total': 0,
                    'activos': 0,
                    'en_prueba': 0,
                    'con_alertas': 0,
                    'lista_empleados': [],
                    'es_jefe': False,
                    'tipo_jefe': None
                }
            }


class EmpleadoCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Vista para crear nuevo empleado"""
    model = Empleado
    form_class = EmpleadoForm
    template_name = 'employees/empleado_form.html'
    permission_required = 'employees.add_empleado'
    success_url = reverse_lazy('employees:empleado_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Crear Nuevo Empleado'
        return context
    
    def form_valid(self, form):
        """Procesar formulario válido y crear usuario automáticamente"""
        try:
            with transaction.atomic():
                # Asignar usuario creador
                form.instance.creado_por = self.request.user
                
                # Determinar estado inicial basado en fecha de ingreso.
                # Los aprendices (rol AP001) nunca están en periodo de prueba: su
                # vínculo es contrato de aprendizaje, no laboral.
                cargo_form = form.cleaned_data.get('cargo')
                es_aprendiz = bool(
                    cargo_form
                    and getattr(cargo_form, 'rol_automatico', None)
                    and cargo_form.rol_automatico.codigo == 'AP001'
                )
                dias_desde_ingreso = (date.today() - form.instance.fecha_ingreso).days

                if dias_desde_ingreso <= 60 and not es_aprendiz:  # Período de prueba
                    try:
                        estado_prueba = EstadoEmpleado.objects.get(codigo='p-prue')
                        form.instance.estado = estado_prueba
                    except EstadoEmpleado.DoesNotExist:
                        logger.warning("Estado código 'p-prue' (Periodo de prueba) no encontrado")
                        # Usar el primer estado disponible
                        form.instance.estado = EstadoEmpleado.objects.first()
                else:
                    try:
                        estado_activo = EstadoEmpleado.objects.get(codigo='999')
                        form.instance.estado = estado_activo
                    except EstadoEmpleado.DoesNotExist:
                        logger.warning("Estado código '999' (Activo) no encontrado")
                        form.instance.estado = EstadoEmpleado.objects.first()
                
                # Guardar empleado primero
                empleado = form.save()

                # Crear usuario automáticamente solo si el cargo del empleado permite
                # acceso al sistema. Para cargos sin acceso (ej: aprendiz lectiva),
                # el usuario se creará al rotar a un cargo con acceso vía la señal
                # de HistorialCargo.
                cargo = cargo_form
                permite_crear_usuario = (
                    cargo is None or getattr(cargo, 'crea_usuario_sistema', True)
                )
                if empleado.numero_documento and permite_crear_usuario:
                    usuario_creado = self.crear_usuario_automatico(empleado)
                    if usuario_creado:
                        empleado.usuario = usuario_creado
                        empleado.save()
                elif cargo and not permite_crear_usuario:
                    messages.info(
                        self.request,
                        f"ℹ️ El cargo '{cargo.nombre}' no crea usuario en el sistema. "
                        f"El usuario se generará cuando el empleado pase a un cargo con acceso."
                    )

                # Crear historial de cargo si se especificó cargo
                if cargo:
                    # Obtener el jefe directo seleccionado (si existe)
                    jefe_directo = form.cleaned_data.get('jefe_directo')

                    HistorialCargo.objects.create(
                        empleado=empleado,
                        cargo=cargo,
                        fecha_inicio=empleado.fecha_ingreso,
                        activo=True,
                        creado_por=self.request.user,
                        jefe_directo=jefe_directo
                    )
                
                messages.success(
                    self.request, 
                    f'✅ Empleado {empleado.nombre_completo} creado exitosamente.'
                )
                
                return super().form_valid(form)
                
        except Exception as e:
            logger.error(f"Error creando empleado: {str(e)}")
            messages.error(
                self.request,
                f'❌ Error al crear el empleado: {str(e)}'
            )
            return self.form_invalid(form)
    
    def crear_usuario_automatico(self, empleado):
        """Crear usuario automáticamente con nombre y documento"""
        try:
            # Generar username: primer_nombre.primer_apellido
            primer_nombre = empleado.nombres.split()[0].lower()
            primer_apellido = empleado.apellidos.split()[0].lower()
            username_base = f"{primer_nombre}.{primer_apellido}"
            
            # Asegurar username único
            username = username_base
            counter = 1
            while User.objects.filter(username=username).exists():
                username = f"{username_base}{counter}"
                counter += 1
            
            # Generar password: Primer nombre + documento
            password = f"{primer_nombre.capitalize()}{empleado.numero_documento}"
            
            # Crear usuario
            user = User.objects.create_user(
                username=username,
                email=empleado.correo_electronico,
                first_name=empleado.nombres,
                last_name=empleado.apellidos,
                password=password,
                is_active=True
            )
            
            # Mostrar mensaje con credenciales
            messages.success(
                self.request,
                f"✅ Usuario creado automáticamente:\n"
                f"👤 Usuario: {username}\n"
                f"🔑 Contraseña: {password}\n"
                f"📧 Email: {empleado.correo_electronico}\n"
                f"(Comunicar estas credenciales al empleado)"
            )
            
            logger.info(f"Usuario creado para empleado {empleado.numero_documento}: {username}")
            return user
            
        except Exception as e:
            logger.error(f"Error creando usuario para empleado {empleado.numero_documento}: {str(e)}")
            messages.warning(
                self.request,
                f"⚠️ No se pudo crear usuario automáticamente: {str(e)}"
            )
            return None
    
    def form_invalid(self, form):
        """Manejar formulario inválido"""
        logger.warning(f"Formulario inválido: {form.errors}")
        
        # Agregar errores específicos a los mensajes
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(self.request, f"{field}: {error}")
        
        return super().form_invalid(form)


class EmpleadoUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Vista para editar empleado"""
    model = Empleado
    form_class = EmpleadoForm
    template_name = 'employees/empleado_form.html'
    permission_required = 'employees.change_empleado'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = f'Editar Empleado - {self.object.nombre_completo}'
        return context
    
    def get_initial(self):
        """Preparar datos iniciales del formulario"""
        initial = super().get_initial()
        # Formatear las fechas en el formato correcto para el input type="date"
        if self.object.fecha_ingreso:
            initial['fecha_ingreso'] = self.object.fecha_ingreso.strftime('%Y-%m-%d')
        if self.object.fecha_nacimiento:
            initial['fecha_nacimiento'] = self.object.fecha_nacimiento.strftime('%Y-%m-%d')
        
        # Obtener el cargo actual si existe
        cargo_actual = self.object.historialcargo_set.filter(activo=True).first()
        if cargo_actual:
            initial['cargo'] = cargo_actual.cargo
        
        return initial
    
    def get_success_url(self):
        return reverse_lazy('employees:empleado_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        """Procesar cambios en el formulario"""
        try:
            with transaction.atomic():
                # Verificar si cambió el cargo o jefe directo
                cargo_nuevo = form.cleaned_data.get('cargo')
                jefe_directo_nuevo = form.cleaned_data.get('jefe_directo')
                cargo_actual = None
                jefe_directo_actual = None

                # Obtener cargo actual
                historial_actual = self.object.historialcargo_set.filter(activo=True).first()
                if historial_actual:
                    cargo_actual = historial_actual.cargo
                    jefe_directo_actual = historial_actual.jefe_directo

                # Si cambió el cargo, crear nuevo historial
                if cargo_nuevo and cargo_nuevo != cargo_actual:
                    # Desactivar cargo actual
                    if historial_actual:
                        historial_actual.activo = False
                        historial_actual.fecha_fin = date.today()
                        historial_actual.save()

                    # Crear nuevo historial
                    HistorialCargo.objects.create(
                        empleado=self.object,
                        cargo=cargo_nuevo,
                        fecha_inicio=date.today(),
                        activo=True,
                        motivo_cambio="Actualización de cargo",
                        creado_por=self.request.user,
                        jefe_directo=jefe_directo_nuevo
                    )

                    # Si el empleado no tenía usuario porque venía de un cargo sin acceso
                    # y el nuevo cargo sí crea usuario, generarlo aquí para mostrar las
                    # credenciales al RRHH en pantalla. La señal de HistorialCargo es
                    # idempotente y no duplicará el usuario.
                    if (
                        not self.object.usuario_id
                        and getattr(cargo_nuevo, 'crea_usuario_sistema', True)
                        and cargo_actual is not None
                        and not getattr(cargo_actual, 'crea_usuario_sistema', True)
                    ):
                        from apps.employees.signals import _crear_usuario_para_empleado
                        resultado = _crear_usuario_para_empleado(self.object, cargo=cargo_nuevo)
                        if resultado:
                            user_creado, password_generado = resultado
                            self.object.refresh_from_db()
                            messages.success(
                                self.request,
                                f"✅ Usuario creado automáticamente al promover desde "
                                f"'{cargo_actual.nombre}' a '{cargo_nuevo.nombre}':\n"
                                f"👤 Usuario: {user_creado.username}\n"
                                f"🔑 Contraseña: {password_generado}\n"
                                f"📧 Email: {self.object.correo_electronico or '(sin email)'}\n"
                                f"(Comunicar estas credenciales al empleado)"
                            )
                # Si solo cambió el jefe directo (mismo cargo)
                elif historial_actual and jefe_directo_nuevo != jefe_directo_actual:
                    historial_actual.jefe_directo = jefe_directo_nuevo
                    historial_actual.save()

                response = super().form_valid(form)

                messages.success(
                    self.request,
                    f'Empleado {self.object.nombre_completo} actualizado exitosamente.'
                )

                return response
                
        except Exception as e:
            logger.error(f"Error actualizando empleado: {str(e)}")
            messages.error(
                self.request,
                f'Error al actualizar el empleado: {str(e)}'
            )
            return self.form_invalid(form)


@login_required
@require_http_methods(["POST"])
def empleado_inactivar(request, pk):
    """
    Vista para inactivar un empleado.
    Cambia el estado del empleado a INACTIVO y desactiva sus cargos.
    """
    try:
        empleado = get_object_or_404(Empleado, pk=pk)

        # Verificar permisos
        if not request.user.has_perm('employees.change_empleado'):
            return JsonResponse({
                'success': False,
                'message': 'No tienes permisos para inactivar empleados'
            }, status=403)

        # Verificar que no esté ya inactivo
        if empleado.estado.codigo.lower() == 'inactivo':
            return JsonResponse({
                'success': False,
                'message': 'El empleado ya está inactivo'
            }, status=400)

        with transaction.atomic():
            # Buscar el estado INACTIVO
            estado_inactivo = EstadoEmpleado.objects.filter(
                codigo__iexact='INACTIVO'
            ).first()

            if not estado_inactivo:
                # Si no existe, intentar crear el estado INACTIVO
                estado_inactivo = EstadoEmpleado.objects.create(
                    codigo='INACTIVO',
                    nombre='Inactivo'
                )

            # Cambiar estado del empleado
            empleado.estado = estado_inactivo
            empleado.save()

            # Desactivar todos los cargos activos del empleado
            HistorialCargo.objects.filter(
                empleado=empleado,
                activo=True
            ).update(
                activo=False,
                fecha_fin=timezone.now().date()
            )

            # Desactivar usuario (no puede acceder al sistema)
            if empleado.usuario:
                empleado.usuario.is_active = False
                empleado.usuario.save()

            logger.info(
                f"Empleado inactivado: {empleado.nombre_completo} "
                f"(ID: {empleado.pk}) por {request.user.username}"
            )

        return JsonResponse({
            'success': True,
            'message': f'Empleado {empleado.nombre_completo} inactivado exitosamente'
        })

    except Exception as e:
        logger.error(f"Error al inactivar empleado {pk}: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Error al inactivar empleado: {str(e)}'
        }, status=500)


@login_required
def empleado_search_api(request):
    """API para búsqueda de empleados (autocompletado)"""
    query = request.GET.get('q', '').strip()
    limit = int(request.GET.get('limit', 10))
    
    if len(query) < 2:
        return JsonResponse({'results': []})
    
    empleados = Empleado.objects.filter(
        Q(nombres__icontains=query) | 
        Q(apellidos__icontains=query) |
        Q(numero_documento__icontains=query)
    ).select_related('usuario', 'estado')[:limit]
    
    results = []
    for emp in empleados:
        results.append({
            'id': str(emp.pk),
            'text': emp.nombre_completo,
            'email': emp.correo_electronico,
            'documento': emp.numero_documento,
            'estado': emp.estado.nombre
        })
    
    return JsonResponse({'results': results})



def empleado_export(request):
    """Exportar lista de empleados - VERSIÓN MEJORADA"""
    formato = request.GET.get('export', 'excel')
    
    # Obtener empleados con los mismos filtros de la lista
    view = EmpleadoListView()
    view.request = request
    empleados = list(view.get_queryset())
    
    try:
        if formato == 'excel':
            return export_empleados_excel(empleados)
        elif formato == 'pdf':
            return export_empleados_pdf(empleados)
        elif formato == 'csv':
            return export_empleados_csv(empleados)
        else:
            messages.error(request, 'Formato de exportación no válido')
            return redirect('employees:empleado_list')
            
    except Exception as e:
        logger.error(f"Error en exportación {formato}: {e}")
        messages.error(request, f'Error al exportar: {str(e)}')
        return redirect('employees:empleado_list')

@login_required
def empleado_export_individual(request, pk):
    """Exportar perfil individual de empleado"""
    empleado = get_object_or_404(Empleado, pk=pk)
    formato = request.GET.get('format', 'pdf')
    
    try:
        if formato == 'pdf':
            return export_empleado_perfil_pdf(empleado)
        elif formato == 'excel':
            return export_empleado_excel(empleado)
        else:
            messages.error(request, 'Formato no válido')
            return redirect('employees:empleado_detail', pk=pk)
            
    except Exception as e:
        logger.error(f"Error en exportación individual {formato}: {e}")
        messages.error(request, f'Error al generar {formato.upper()}: {str(e)}')
        return redirect('employees:empleado_detail', pk=pk)


@login_required
def empleado_print_view(request, pk):
    """Vista optimizada para impresión del empleado"""
    empleado = get_object_or_404(Empleado, pk=pk)
    
    # Obtener datos adicionales
    historial_cargos = empleado.historialcargo_set.select_related(
        'cargo__area'
    ).order_by('-fecha_inicio')
    
    context = {
        'empleado': empleado,
        'historial_cargos': historial_cargos,
        'fecha_generacion': timezone.now(),
    }
    
    return render(request, 'employees/empleado_print.html', context)


@login_required
def empleado_historial_export(request, pk):
    """Exportar historial completo de cambios del empleado"""
    empleado = get_object_or_404(Empleado, pk=pk)
    formato = request.GET.get('format', 'pdf')
    
    try:
        historial_cargos = empleado.historialcargo_set.all().order_by('-fecha_inicio')
        
        if formato == 'excel':
            return export_historial_excel(empleado, historial_cargos)
        else:  # PDF por defecto
            return export_historial_pdf(empleado, historial_cargos)
            
    except Exception as e:
        logger.error(f"Error exportando historial: {e}")
        messages.error(request, f'Error al exportar historial: {str(e)}')
        return redirect('employees:empleado_detail', pk=pk)


# Funciones auxiliares para historial
def export_historial_excel(empleado, historial_cargos):
    """Exportar historial de cargos a Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Historial de Cargos'
        
        # Título
        worksheet.merge_cells('A1:F1')
        title_cell = worksheet['A1']
        title_cell.value = f"HISTORIAL DE CARGOS - {empleado.nombre_completo.upper()}"
        title_cell.font = Font(bold=True, size=14)
        
        # Encabezados
        headers = ['Cargo', 'Área', 'Fecha Inicio', 'Fecha Fin', 'Estado', 'Motivo del Cambio']
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Datos
        for row, hist in enumerate(historial_cargos, 4):
            worksheet.cell(row=row, column=1, value=hist.cargo.nombre)
            worksheet.cell(row=row, column=2, value=hist.cargo.area.nombre)
            worksheet.cell(row=row, column=3, value=hist.fecha_inicio)
            worksheet.cell(row=row, column=4, value=hist.fecha_fin or 'Actual')
            worksheet.cell(row=row, column=5, value='Activo' if hist.activo else 'Finalizado')
            worksheet.cell(row=row, column=6, value=hist.motivo_cambio or '-')
        
        # Ajustar columnas
        for col in range(1, 7):
            worksheet.column_dimensions[chr(64 + col)].width = 20
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'historial_{empleado.nombres}_{empleado.apellidos}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        workbook.save(response)
        return response
        
    except Exception as e:
        return HttpResponse(f"Error: {e}", content_type="text/plain", status=500)


def export_historial_pdf(empleado, historial_cargos):
    """Exportar historial de cargos a PDF"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        
        response = HttpResponse(content_type='application/pdf')
        filename = f'historial_{empleado.nombres}_{empleado.apellidos}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        doc = SimpleDocTemplate(response, pagesize=A4)
        elements = []
        
        styles = getSampleStyleSheet()
        
        # Título
        title = Paragraph(f"HISTORIAL DE CARGOS<br/>{empleado.nombre_completo}", styles['Title'])
        elements.append(title)
        
        # Datos
        data = [['Cargo', 'Área', 'Fecha Inicio', 'Fecha Fin', 'Estado', 'Motivo']]
        
        for hist in historial_cargos:
            data.append([
                hist.cargo.nombre,
                hist.cargo.area.nombre,
                hist.fecha_inicio.strftime('%d/%m/%Y'),
                hist.fecha_fin.strftime('%d/%m/%Y') if hist.fecha_fin else 'Actual',
                'Activo' if hist.activo else 'Finalizado',
                hist.motivo_cambio or '-'
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        return response
        
    except Exception as e:
        return HttpResponse(f"Error: {e}", content_type="text/plain", status=500)
    

class EmpleadoPerfilView(LoginRequiredMixin, DetailView):
    """Vista del perfil personal del empleado - Dashboard personal"""
    model = Empleado
    template_name = 'employees/empleado_perfil.html'
    context_object_name = 'empleado'
    
    def get_object(self):
        """Obtener el empleado asociado al usuario logueado o por PK"""
        # Si se pasa pk en la URL, usar ese empleado
        if 'pk' in self.kwargs:
            return get_object_or_404(Empleado, pk=self.kwargs['pk'])
        
        # Si no hay pk, buscar empleado del usuario logueado
        try:
            empleado = Empleado.objects.get(usuario=self.request.user)
            return empleado
        except Empleado.DoesNotExist:
            # Si no hay empleado asociado, mostrar error 404
            raise Http404("No se encontró un perfil de empleado asociado a tu usuario.")
    
    def dispatch(self, request, *args, **kwargs):
        """Verificar permisos antes de mostrar la vista"""
        try:
            empleado = self.get_object()
            
            # Verificar si el usuario puede ver este perfil
            if empleado.usuario != request.user and not request.user.is_staff:
                messages.error(
                    request, 
                    'No tienes permisos para ver este perfil.'
                )
                return redirect('core:dashboard')
            
            return super().dispatch(request, *args, **kwargs)
            
        except Http404:
            messages.error(
                request, 
                'No se encontró un perfil de empleado asociado a tu usuario. Contacta al administrador.'
            )
            return redirect('core:dashboard')
        except Exception as e:
            logger.error(f"Error en dispatch de EmpleadoPerfilView: {e}")
            messages.error(request, 'Error al acceder al perfil.')
            return redirect('core:dashboard')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        empleado = self.get_object()
        
        # === INFORMACIÓN BÁSICA ===
        context.update(self.get_informacion_basica(empleado))
        
        # === ESTADO DE DOCUMENTOS ===
        context.update(self.get_estado_documentos(empleado))
        
        # === CAPACITACIONES ===
        context.update(self.get_capacitaciones(empleado))
        
        # === EVALUACIONES ===
        context.update(self.get_evaluaciones(empleado))

        # === ENCUESTAS ===
        context.update(self.get_encuestas(empleado))

        # === EMPLEADOS A CARGO ===
        context.update(self.get_empleados_a_cargo(empleado))
        
        # === SISTEMA DE PUNTOS ===
        context.update(self.get_sistema_puntos(empleado))
        
        # === NOTIFICACIONES RECIENTES ===
        context.update(self.get_notificaciones(empleado))
        
        # === ACTIVIDAD RECIENTE ===
        context.update(self.get_actividad_reciente(empleado))

        # === Tiles del módulo de vacaciones ===
        # Equipo: SOLO si el empleado tiene subordinados directos.
        # Admin: SOLO si el usuario es staff. Son independientes; un usuario
        # puede ver los dos tiles si cumple ambas condiciones.
        context['puede_ver_vacaciones_equipo'] = HistorialCargo.objects.filter(
            jefe_directo=empleado, activo=True
        ).exists()
        context['puede_ver_vacaciones_admin'] = self.request.user.is_staff
        # Asistencia: mismo criterio que vacaciones_equipo — tiene subordinados directos.
        context['puede_ver_asistencia_equipo'] = context['puede_ver_vacaciones_equipo']

        return context

    def get_informacion_basica(self, empleado):
        """Obtener información básica del empleado"""
        # Cargo actual
        cargo_actual = empleado.historialcargo_set.filter(activo=True).first()
        
        # Días en la empresa
        dias_empresa = (date.today() - empleado.fecha_ingreso).days
        
        # Formatear tiempo en empresa de forma amigable
        tiempo_empresa_texto = self.formatear_tiempo_empresa(dias_empresa)
        
        # Tiempo en cargo actual
        tiempo_cargo = None
        if cargo_actual:
            tiempo_cargo = (date.today() - cargo_actual.fecha_inicio).days
        
        return {
            'cargo_actual': cargo_actual,
            'dias_empresa': dias_empresa,
            'tiempo_empresa_texto': tiempo_empresa_texto,
            'tiempo_cargo': tiempo_cargo,
            'anos_empresa': dias_empresa // 365,
        }
    
    def formatear_tiempo_empresa(self, dias_totales):
        """
        Formatear el tiempo en la empresa de forma amigable:
        - Menos de 30 días: "X días"
        - Menos de 365 días: "X meses y Y días"
        - 365 días o más: "X años, Y meses y Z días"
        """
        if dias_totales < 30:
            return f"{dias_totales} día{'s' if dias_totales != 1 else ''}"
        
        elif dias_totales < 365:
            meses = dias_totales // 30
            dias_restantes = dias_totales % 30
            
            texto = f"{meses} mes{'es' if meses != 1 else ''}"
            if dias_restantes > 0:
                texto += f" y {dias_restantes} día{'s' if dias_restantes != 1 else ''}"
            
            return texto
        
        else:
            anos = dias_totales // 365
            dias_restantes = dias_totales % 365
            meses = dias_restantes // 30
            dias = dias_restantes % 30
            
            texto = f"{anos} año{'s' if anos != 1 else ''}"
            
            if meses > 0:
                texto += f", {meses} mes{'es' if meses != 1 else ''}"
            
            if dias > 0:
                texto += f" y {dias} día{'s' if dias != 1 else ''}"
            
            return texto
    
    def get_estado_documentos(self, empleado):
        """Obtener estado de documentos del empleado"""
        try:
            # Importar aquí para evitar errores si no existe el módulo
            from apps.documents.models import DocumentoEmpleado, TipoDocumentoEmpleado
            
            # Documentos del empleado
            documentos = DocumentoEmpleado.objects.filter(empleado=empleado)
            
            # Tipos de documentos obligatorios
            tipos_obligatorios = TipoDocumentoEmpleado.objects.filter(
                obligatorio=True, 
                activo=True
            )
            
            # Documentos aprobados
            docs_aprobados = documentos.filter(estado_aprobacion='aprobado')
            
            # Documentos pendientes
            docs_pendientes = documentos.filter(estado_aprobacion='pendiente')
            
            # Documentos próximos a vencer (30 días)
            fecha_limite = date.today() + timedelta(days=30)
            docs_por_vencer = docs_aprobados.filter(
                fecha_vencimiento__isnull=False,
                fecha_vencimiento__lte=fecha_limite,
                fecha_vencimiento__gte=date.today()
            )
            
            # Documentos vencidos
            docs_vencidos = docs_aprobados.filter(
                fecha_vencimiento__isnull=False,
                fecha_vencimiento__lt=date.today()
            )
            
            # Documentos faltantes (obligatorios no subidos)
            tipos_subidos = documentos.values_list('tipo_documento_id', flat=True)
            tipos_faltantes = tipos_obligatorios.exclude(id__in=tipos_subidos)
            
            # Calcular progreso de documentación
            total_obligatorios = tipos_obligatorios.count()
            aprobados_obligatorios = docs_aprobados.filter(
                tipo_documento__in=tipos_obligatorios
            ).count()
            
            progreso_documentos = 0
            if total_obligatorios > 0:
                progreso_documentos = round((aprobados_obligatorios / total_obligatorios) * 100)
            
            return {
                'documentos': {
                    'total': documentos.count(),
                    'aprobados': docs_aprobados.count(),
                    'pendientes': docs_pendientes.count(),
                    'por_vencer': docs_por_vencer.count(),
                    'vencidos': docs_vencidos.count(),
                    'faltantes': tipos_faltantes.count(),
                    'progreso': progreso_documentos,
                    'lista_por_vencer': docs_por_vencer[:5],  # Primeros 5
                    'lista_vencidos': docs_vencidos[:5],
                    'lista_faltantes': tipos_faltantes[:5],
                    'documentacion_completa': tipos_faltantes.count() == 0 and docs_pendientes.count() == 0
                }
            }
        except ImportError:
            logger.info("Módulo de documentos no disponible")
            return {
                'documentos': {
                    'total': 0, 'aprobados': 0, 'pendientes': 0,
                    'por_vencer': 0, 'vencidos': 0, 'faltantes': 0,
                    'progreso': 0, 'documentacion_completa': True,
                    'lista_por_vencer': [], 'lista_vencidos': [], 'lista_faltantes': []
                }
            }
        except Exception as e:
            logger.error(f"Error obteniendo estado de documentos: {e}")
            return {
                'documentos': {
                    'total': 0, 'aprobados': 0, 'pendientes': 0,
                    'por_vencer': 0, 'vencidos': 0, 'faltantes': 0,
                    'progreso': 0, 'documentacion_completa': False,
                    'lista_por_vencer': [], 'lista_vencidos': [], 'lista_faltantes': []
                }
            }
    
    def get_capacitaciones(self, empleado):
        """Obtener estado de capacitaciones del empleado"""
        try:
            from apps.training.models import InscripcionCapacitacion, Capacitacion
            
            # Capacitaciones asignadas
            inscripciones = InscripcionCapacitacion.objects.filter(
                empleado=empleado
            ).select_related('capacitacion')
            
            # Estados
            completadas = inscripciones.filter(estado='completado')
            en_progreso = inscripciones.filter(estado='en_progreso')
            pendientes = inscripciones.filter(
                estado='no_iniciado'
            )
            
            # Próximas a vencer (si tienen fecha límite)
            fecha_limite = date.today() + timedelta(days=7)
            proximas_vencer = en_progreso.filter(
                fecha_limite__isnull=False,
                fecha_limite__lte=fecha_limite
            )
            
            # Progreso general
            total_capacitaciones = inscripciones.count()
            capacitaciones_completadas = completadas.count()
            
            progreso_capacitaciones = 0
            if total_capacitaciones > 0:
                progreso_capacitaciones = round(
                    (capacitaciones_completadas / total_capacitaciones) * 100
                )
            
            return {
                'capacitaciones': {
                    'total': total_capacitaciones,
                    'completadas': capacitaciones_completadas,
                    'en_progreso': en_progreso.count(),
                    'pendientes': pendientes.count(),
                    'proximas_vencer': proximas_vencer.count(),
                    'progreso': progreso_capacitaciones,
                    'lista_pendientes': pendientes[:5],
                    'lista_en_progreso': en_progreso[:5],
                    'lista_proximas_vencer': proximas_vencer[:3],
                }
            }
        except ImportError:
            logger.info("Módulo de capacitaciones no disponible")
            return {
                'capacitaciones': {
                    'total': 0, 'completadas': 0, 'en_progreso': 0,
                    'pendientes': 0, 'progreso': 100,
                    'lista_pendientes': [], 'lista_en_progreso': [], 'lista_proximas_vencer': []
                }
            }
        except Exception as e:
            logger.error(f"Error obteniendo capacitaciones: {e}")
            return {
                'capacitaciones': {
                    'total': 0, 'completadas': 0, 'en_progreso': 0,
                    'pendientes': 0, 'progreso': 0,
                    'lista_pendientes': [], 'lista_en_progreso': [], 'lista_proximas_vencer': []
                }
            }
    
    def get_evaluaciones(self, empleado):
        """Obtener estado de evaluaciones del empleado"""
        try:
            from apps.evaluations.models import AsignacionEvaluacion
            
            # Evaluaciones asignadas
            evaluaciones = AsignacionEvaluacion.objects.filter(
                empleado_evaluado=empleado
            ).select_related('evaluacion', 'evaluador')
            
            # Pendientes (estado pendiente o en progreso, y no vencidas)
            pendientes = evaluaciones.filter(
                estado__in=['pendiente', 'en_progreso']
            )
            
            # Vencidas (estado vencida o vencidas por fecha)
            vencidas = evaluaciones.filter(
                estado='vencida'
            ) | evaluaciones.filter(
                fecha_completada__isnull=True,
                fecha_vencimiento__lt=date.today()
            )
            
            # Completadas este año (estado completada)
            completadas_año = evaluaciones.filter(
                estado='completada',
                fecha_completada__year=date.today().year
            )
            completadas_año = evaluaciones.filter(
                estado='completada',
                fecha_completada__year=date.today().year
            )
            
            # Próximas (pendientes o en progreso, siguientes 15 días)
            fecha_limite = date.today() + timedelta(days=15)
            proximas = pendientes.filter(fecha_vencimiento__lte=fecha_limite)
            
            return {
                'evaluaciones': {
                    'pendientes': pendientes.count(),
                    'vencidas': vencidas.count(),
                    'completadas_año': completadas_año.count(),
                    'proximas': proximas.count(),
                    'lista_pendientes': pendientes[:5],
                    'lista_proximas': proximas[:3],
                    'ultima_evaluacion': evaluaciones.filter(fecha_completada__isnull=False).order_by('-fecha_completada').first()
                }
            }
        except ImportError:
            logger.info("Módulo de evaluaciones no disponible")
            return {
                'evaluaciones': {
                    'pendientes': 0, 'vencidas': 0, 'completadas_año': 0,
                    'proximas': 0, 'lista_pendientes': [], 'lista_proximas': [],
                    'ultima_evaluacion': None
                }
            }
        except Exception as e:
            logger.error(f"Error obteniendo evaluaciones: {e}")
            return {
                'evaluaciones': {
                    'pendientes': 0, 'vencidas': 0, 'completadas_año': 0,
                    'proximas': 0, 'lista_pendientes': [], 'lista_proximas': [],
                    'ultima_evaluacion': None
                }
            }

    def get_encuestas(self, empleado):
        """Obtener estado de encuestas del empleado"""
        try:
            from apps.surveys.models import ParticipacionEncuesta, Encuesta

            # Participaciones del empleado
            participaciones = ParticipacionEncuesta.objects.filter(
                empleado=empleado
            ).select_related('encuesta', 'encuesta__tipo_encuesta')

            # Encuestas pendientes (no completadas y activas)
            pendientes = participaciones.filter(
                completada=False,
                encuesta__activa=True,
                encuesta__fecha_inicio__lte=date.today(),
                encuesta__fecha_fin__gte=date.today()
            )

            # Encuestas completadas
            completadas = participaciones.filter(completada=True)

            # Encuestas completadas este año
            completadas_año = completadas.filter(
                fecha_completada__year=date.today().year
            )

            # Encuestas próximas a vencer (siguientes 7 días)
            fecha_limite = date.today() + timedelta(days=7)
            proximas_vencer = pendientes.filter(
                encuesta__fecha_fin__lte=fecha_limite
            )

            return {
                'encuestas': {
                    'pendientes': pendientes.count(),
                    'completadas': completadas.count(),
                    'completadas_año': completadas_año.count(),
                    'proximas_vencer': proximas_vencer.count(),
                    'lista_pendientes': pendientes[:5],  # Primeras 5 pendientes
                    'lista_proximas_vencer': proximas_vencer[:3],
                    'ultima_completada': completadas.order_by('-fecha_completada').first()
                }
            }
        except ImportError:
            logger.info("Módulo de encuestas no disponible")
            return {
                'encuestas': {
                    'pendientes': 0, 'completadas': 0, 'completadas_año': 0,
                    'proximas_vencer': 0, 'lista_pendientes': [],
                    'lista_proximas_vencer': [], 'ultima_completada': None
                }
            }
        except Exception as e:
            logger.error(f"Error obteniendo encuestas: {e}")
            return {
                'encuestas': {
                    'pendientes': 0, 'completadas': 0, 'completadas_año': 0,
                    'proximas_vencer': 0, 'lista_pendientes': [],
                    'lista_proximas_vencer': [], 'ultima_completada': None
                }
            }

    def get_empleados_a_cargo(self, empleado):
        """Obtener empleados a cargo según jerarquía organizacional"""
        try:
            # Obtener cargo actual del empleado
            cargo_actual = empleado.historialcargo_set.filter(activo=True).first()
            
            if not cargo_actual or not cargo_actual.cargo:
                return {
                    'empleados_a_cargo': {
                        'total': 0,
                        'activos': 0,
                        'en_prueba': 0,
                        'con_alertas': 0,
                        'lista_empleados': [],
                        'es_jefe': False,
                        'tipo_jefe': None
                    }
                }
            
            # Determinar tipo de jefe según nombre del cargo
            cargo_nombre = cargo_actual.cargo.nombre.lower()
            tipo_jefe = None
            
            if 'gerente' in cargo_nombre:
                tipo_jefe = 'gerente'
            elif 'director' in cargo_nombre:
                tipo_jefe = 'director'  
            elif 'coordinador' in cargo_nombre:
                tipo_jefe = 'coordinador'
            elif 'supervisor' in cargo_nombre or 'jefe' in cargo_nombre:
                tipo_jefe = 'supervisor'
            
            if not tipo_jefe:
                # No es un cargo de jefatura
                return {
                    'empleados_a_cargo': {
                        'total': 0,
                        'activos': 0,
                        'en_prueba': 0,
                        'con_alertas': 0,
                        'lista_empleados': [],
                        'es_jefe': False,
                        'tipo_jefe': None
                    }
                }
            
            # Buscar empleados a cargo usando SOLO jefe_directo
            # Solo mostrar empleados que:
            # 1. Tienen jefe_directo asignado = empleado actual
            # 2. Están ACTIVOS o en PERÍODO DE PRUEBA
            empleados_subordinados = Empleado.objects.filter(
                historialcargo__activo=True,
                historialcargo__jefe_directo=empleado,
                estado__codigo__in=['999', 'p-prue']  # Activos y en período de prueba
            ).distinct()

            # Calcular estadísticas
            total_subordinados = empleados_subordinados.count()
            activos = empleados_subordinados.filter(estado__codigo='999').count()
            en_prueba = empleados_subordinados.filter(estado__codigo='p-prue').count()
            
            # Empleados con alertas (documentos vencidos, evaluaciones pendientes, etc.)
            con_alertas = 0
            lista_empleados = []

            for emp in empleados_subordinados.select_related('estado').prefetch_related('historialcargo_set__cargo'):
                # Verificar si tiene alertas
                tiene_alertas = False
                alertas = []
                
                # Verificar documentos vencidos/pendientes
                try:
                    from apps.documents.models import DocumentoEmpleado
                    docs_vencidos = DocumentoEmpleado.objects.filter(
                        empleado=emp,
                        fecha_vencimiento__lt=date.today()
                    ).count()
                    if docs_vencidos > 0:
                        tiene_alertas = True
                        alertas.append(f"{docs_vencidos} doc. vencidos")
                except ImportError:
                    pass
                
                # Verificar evaluaciones pendientes
                try:
                    from apps.evaluations.models import AsignacionEvaluacion
                    eval_pendientes = AsignacionEvaluacion.objects.filter(
                        empleado_evaluado=emp,
                        estado__in=['asignada', 'en_progreso'],
                        fecha_vencimiento__lt=date.today()
                    ).count()
                    if eval_pendientes > 0:
                        tiene_alertas = True
                        alertas.append(f"{eval_pendientes} eval. vencidas")
                except ImportError:
                    pass
                
                if tiene_alertas:
                    con_alertas += 1
                
                # Obtener cargo actual del subordinado
                cargo_subordinado = emp.historialcargo_set.filter(activo=True).first()
                
                lista_empleados.append({
                    'empleado': emp,
                    'cargo_actual': cargo_subordinado,
                    'tiene_alertas': tiene_alertas,
                    'alertas': alertas,
                    'dias_empresa': (date.today() - emp.fecha_ingreso).days
                })
            
            return {
                'empleados_a_cargo': {
                    'total': total_subordinados,
                    'activos': activos,
                    'en_prueba': en_prueba,
                    'con_alertas': con_alertas,
                    'lista_empleados': lista_empleados,
                    'es_jefe': True,
                    'tipo_jefe': tipo_jefe
                }
            }
            
        except Exception as e:
            logger.error(f"Error obteniendo empleados a cargo: {e}")
            return {
                'empleados_a_cargo': {
                    'total': 0,
                    'activos': 0,
                    'en_prueba': 0,
                    'con_alertas': 0,
                    'lista_empleados': [],
                    'es_jefe': False,
                    'tipo_jefe': None
                }
            }

    def get_sistema_puntos(self, empleado):
        """Obtener información del sistema de puntos y reconocimientos"""
        try:
            from apps.recognition.models import HistorialPuntos, InsigniaEmpleado
            from django.db import models
            
            # Puntos totales
            puntos_totales = HistorialPuntos.objects.filter(
                empleado=empleado,
                validado=True
            ).aggregate(
                total=models.Sum('puntos')
            )['total'] or 0
            
            # Puntos este mes
            inicio_mes = date.today().replace(day=1)
            puntos_mes = HistorialPuntos.objects.filter(
                empleado=empleado,
                validado=True,
                fecha_obtencion__gte=inicio_mes
            ).aggregate(
                total=models.Sum('puntos')
            )['total'] or 0
            
            # Insignias obtenidas
            insignias = InsigniaEmpleado.objects.filter(
                empleado=empleado
            ).select_related('tipo_insignia')
            
            # Posición en ranking (aproximada)
            empleados_mas_puntos = HistorialPuntos.objects.filter(
                validado=True
            ).values('empleado').annotate(
                total_puntos=models.Sum('puntos')
            ).filter(
                total_puntos__gt=puntos_totales
            ).count()
            
            posicion_ranking = empleados_mas_puntos + 1 if puntos_totales > 0 else 0
            
            return {
                'puntos': {
                    'total': puntos_totales,
                    'este_mes': puntos_mes,
                    'insignias': insignias.count(),
                    'lista_insignias': insignias[:10],
                    'ranking_posicion': posicion_ranking,
                    'puntos_disponibles': puntos_totales  # Para canje
                }
            }
        except ImportError:
            logger.info("Módulo de reconocimientos no disponible")
            return {
                'puntos': {
                    'total': 0, 'este_mes': 0, 'insignias': 0,
                    'ranking_posicion': 0, 'lista_insignias': [],
                    'puntos_disponibles': 0
                }
            }
        except Exception as e:
            logger.error(f"Error obteniendo sistema de puntos: {e}")
            return {
                'puntos': {
                    'total': 0, 'este_mes': 0, 'insignias': 0,
                    'ranking_posicion': 0, 'lista_insignias': [],
                    'puntos_disponibles': 0
                }
            }
    
    def get_notificaciones(self, empleado):
        """Obtener notificaciones recientes del empleado"""
        try:
            from apps.notifications.models import NotificacionEmpleado
            
            # Notificaciones recientes (últimos 7 días)
            fecha_limite = timezone.now() - timedelta(days=7)
            notificaciones = NotificacionEmpleado.objects.filter(
                empleado=empleado,
                fecha_creacion__gte=fecha_limite
            ).order_by('-fecha_creacion')[:10]
            
            # No leídas
            no_leidas = NotificacionEmpleado.objects.filter(
                empleado=empleado,
                leida=False
            ).count()
            
            return {
                'notificaciones': {
                    'recientes': notificaciones,
                    'no_leidas': no_leidas,
                    'total_recientes': notificaciones.count()
                }
            }
        except ImportError:
            logger.info("Módulo de notificaciones no disponible")
            return {
                'notificaciones': {
                    'recientes': [], 'no_leidas': 0, 'total_recientes': 0
                }
            }
        except Exception as e:
            logger.error(f"Error obteniendo notificaciones: {e}")
            return {
                'notificaciones': {
                    'recientes': [], 'no_leidas': 0, 'total_recientes': 0
                }
            }
    
    def get_actividad_reciente(self, empleado):
        """Obtener actividad reciente del empleado"""
        actividades = []
        
        try:
            # Documentos subidos recientemente
            try:
                from apps.documents.models import DocumentoEmpleado
                
                docs_recientes = DocumentoEmpleado.objects.filter(
                    empleado=empleado,
                    fecha_carga__gte=timezone.now() - timedelta(days=15)
                ).order_by('-fecha_carga')[:3]
                
                for doc in docs_recientes:
                    actividades.append({
                        'tipo': 'documento',
                        'icono': 'fas fa-file-upload',
                        'titulo': f'Documento subido: {doc.tipo_documento.nombre}',
                        'fecha': doc.fecha_carga,
                        'estado': doc.estado_aprobacion
                    })
            except ImportError:
                pass
            
            # Capacitaciones completadas recientemente
            try:
                from apps.training.models import InscripcionCapacitacion
                
                caps_completadas = InscripcionCapacitacion.objects.filter(
                    empleado=empleado,
                    estado='completado',
                    fecha_finalizacion__gte=timezone.now() - timedelta(days=15)
                ).order_by('-fecha_finalizacion')[:3]
                
                for cap in caps_completadas:
                    actividades.append({
                        'tipo': 'capacitacion',
                        'icono': 'fas fa-graduation-cap',
                        'titulo': f'Capacitación completada: {cap.capacitacion.nombre}',
                        'fecha': cap.fecha_finalizacion,
                        'estado': 'completado'
                    })
            except ImportError:
                pass
            
            # Evaluaciones completadas pendientes de aceptación
            try:
                from apps.evaluations.models import AsignacionEvaluacion
                
                evaluaciones_pendientes_aceptacion = AsignacionEvaluacion.objects.filter(
                    empleado_evaluado=empleado,
                    estado='completada',
                    fecha_completada__isnull=False,
                    fecha_completada__gte=timezone.now() - timedelta(days=30),  # Últimos 30 días
                    puntaje_total__gte=14  # Solo evaluaciones aprobadas (>= 14 puntos)
                ).exclude(
                    observaciones__icontains='[ACEPTADO_EMPLEADO:'  # Excluir las ya aceptadas
                ).order_by('-fecha_completada')
                
                for evaluacion in evaluaciones_pendientes_aceptacion:
                    # Verificar si tiene resultados para aceptar
                    if hasattr(evaluacion, 'resultadoevaluacion'):
                        resultado = evaluacion.resultadoevaluacion
                        
                        # Verificar si existe plan de mejora predefinido
                        try:
                            from apps.evaluations.models import PlanMejoraPredefinido
                            plan_mejora = PlanMejoraPredefinido.objects.get(asignacion_evaluacion=evaluacion)
                            # Si existe plan predefinido, dirigir al plan
                            url_destino = f'/evaluaciones/ver-plan-mejora/{plan_mejora.id}/'
                            accion_texto = 'Ver Plan de Mejora'
                            titulo_texto = f'📋 Mi Plan de Mejora - {evaluacion.evaluacion.nombre}'
                            descripcion_texto = f'Plan personalizado en estado: {plan_mejora.get_estado_display()}'
                        except PlanMejoraPredefinido.DoesNotExist:
                            # Si no hay plan predefinido, dirigir a resultados
                            url_destino = f'/evaluaciones/ver-resultados/{evaluacion.id}/'
                            accion_texto = 'Ver Resultados'
                            titulo_texto = f'📋 Plan de mejora disponible - {evaluacion.evaluacion.nombre}'
                            descripcion_texto = 'Revisa y acepta tu plan de desarrollo profesional'
                        
                        actividades.append({
                            'tipo': 'evaluacion_aceptar',
                            'icono': 'fas fa-clipboard-check',
                            'color': 'warning',
                            'titulo': titulo_texto,
                            'descripcion': descripcion_texto,
                            'fecha': evaluacion.fecha_completada,
                            'estado': 'pendiente_aceptacion',
                            'url': url_destino,
                            'accion': accion_texto,
                            'evaluacion_id': evaluacion.id,
                            'puntaje': resultado.puntaje_final if resultado else None
                        })
            except ImportError:
                pass
            
            # Ordenar por fecha
            actividades.sort(key=lambda x: x['fecha'], reverse=True)
            
            return {
                'actividades_recientes': actividades[:10]
            }
        except Exception as e:
            logger.error(f"Error obteniendo actividad reciente: {e}")
            return {
                'actividades_recientes': []
            }


class EmpleadoDetailSupervisorView(LoginRequiredMixin, DetailView):
    """Vista para que supervisores vean detalles de sus empleados subordinados"""
    model = Empleado
    template_name = 'employees/empleado_detail_supervisor.html'
    context_object_name = 'empleado'
    
    def dispatch(self, request, *args, **kwargs):
        """Verificar permisos antes de procesar la vista"""
        empleado_a_ver = get_object_or_404(Empleado, pk=kwargs['pk'])
        
        # Si es superusuario, permitir acceso
        if request.user.is_superuser:
            return super().dispatch(request, *args, **kwargs)
        
        # Verificar si el usuario actual es supervisor del empleado
        try:
            usuario_actual = Empleado.objects.get(usuario=request.user)
            cargo_actual = usuario_actual.cargo_actual
            
            if not cargo_actual:
                raise Http404("No tienes permisos para ver este perfil.")
            
            # Verificar si el empleado está bajo mi supervisión
            empleado_cargo = empleado_a_ver.cargo_actual
            if not empleado_cargo:
                raise Http404("Empleado sin cargo asignado.")
            
            # Verificar jerarquía: debe ser mi subordinado directo o indirecto
            es_subordinado = False
            
            # 1. Subordinado directo (mi cargo es su cargo_jefe)
            if empleado_cargo.cargo.cargo_jefe == cargo_actual.cargo:
                es_subordinado = True
            
            # 2. Misma área y nivel jerárquico mayor (subordinado indirecto)
            elif (empleado_cargo.cargo.area == cargo_actual.cargo.area and 
                  empleado_cargo.cargo.nivel_jerarquico > cargo_actual.cargo.nivel_jerarquico):
                es_subordinado = True
            
            if not es_subordinado:
                raise Http404("No tienes permisos para ver este perfil.")
                
        except Empleado.DoesNotExist:
            raise Http404("Usuario sin perfil de empleado.")
        
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        empleado = self.object
        
        # Información básica del empleado
        context['cargo_actual'] = empleado.cargo_actual
        context['tiempo_empresa'] = (date.today() - empleado.fecha_ingreso).days
        context['tiempo_empresa_texto'] = self._calcular_tiempo_empresa_texto(empleado.fecha_ingreso)
        
        # Documentos
        context['documentos'] = self._get_documentos_info(empleado)
        
        # Capacitaciones
        context['capacitaciones'] = self._get_capacitaciones_info(empleado)
        
        # Evaluaciones
        context['evaluaciones'] = self._get_evaluaciones_info(empleado)
        
        # Sistema de puntos
        context['puntos'] = self._get_sistema_puntos(empleado)
        
        # Historial de cargos
        context['historial_cargos'] = HistorialCargo.objects.filter(
            empleado=empleado
        ).select_related('cargo', 'cargo__area').order_by('-fecha_inicio')
        
        # Actividad reciente
        context['actividades_recientes'] = self._get_actividad_reciente(empleado)[:10]
        
        # Información del supervisor
        context['es_vista_supervisor'] = True
        context['supervisor_actual'] = self.request.user.empleado if hasattr(self.request.user, 'empleado') else None
        
        return context
    
    def _calcular_tiempo_empresa_texto(self, fecha_ingreso):
        """Calcular texto legible del tiempo en empresa"""
        try:
            tiempo_total = date.today() - fecha_ingreso
            años = tiempo_total.days // 365
            meses = (tiempo_total.days % 365) // 30
            
            if años > 0:
                return f"{años} año{'s' if años > 1 else ''}"
            elif meses > 0:
                return f"{meses} mes{'es' if meses > 1 else ''}"
            else:
                return f"{tiempo_total.days} días"
        except:
            return "No disponible"
    
    def _get_documentos_info(self, empleado):
        """Obtener información de documentos del empleado"""
        try:
            tipos_requeridos = TipoDocumentoEmpleado.objects.filter(activo=True)
            documentos_empleado = DocumentoEmpleado.objects.filter(empleado=empleado)
            
            aprobados = documentos_empleado.filter(estado='aprobado').count()
            pendientes = documentos_empleado.filter(estado__in=['pendiente', 'revision']).count()
            total = tipos_requeridos.count()
            faltantes = total - documentos_empleado.count()
            
            return {
                'aprobados': aprobados,
                'pendientes': pendientes,
                'total': total,
                'faltantes': faltantes,
                'progreso': int((aprobados / total * 100)) if total > 0 else 0,
            }
        except:
            return {'aprobados': 0, 'pendientes': 0, 'total': 0, 'faltantes': 0, 'progreso': 0}
    
    def _get_capacitaciones_info(self, empleado):
        """Obtener información de capacitaciones del empleado"""
        try:
            inscripciones = InscripcionCapacitacion.objects.filter(empleado=empleado)
            completadas = inscripciones.filter(estado='completada').count()
            en_progreso = inscripciones.filter(estado='en_progreso').count()
            total = inscripciones.count()
            
            return {
                'completadas': completadas,
                'en_progreso': en_progreso,
                'total': total,
                'progreso': int((completadas / total * 100)) if total > 0 else 0,
            }
        except:
            return {'completadas': 0, 'en_progreso': 0, 'total': 0, 'progreso': 0}
    
    def _get_evaluaciones_info(self, empleado):
        """Obtener información de evaluaciones del empleado"""
        try:
            evaluaciones = AsignacionEvaluacion.objects.filter(empleado_evaluado=empleado)
            
            # Pendientes: estado pendiente o en_progreso
            pendientes = evaluaciones.filter(estado__in=['pendiente', 'en_progreso']).count()
            
            # Completadas: estado completada  
            completadas = evaluaciones.filter(estado='completada').count()
            
            # Vencidas: estado vencida o vencidas por fecha
            vencidas = evaluaciones.filter(estado='vencida').count()
            vencidas += evaluaciones.filter(
                fecha_completada__isnull=True,
                fecha_vencimiento__lt=date.today()
            ).exclude(estado='vencida').count()
            
            return {
                'pendientes': pendientes,
                'completadas': completadas,
                'vencidas': vencidas,
                'total': evaluaciones.count(),
            }
        except:
            return {'pendientes': 0, 'completadas': 0, 'vencidas': 0, 'total': 0}
    
    def _get_sistema_puntos(self, empleado):
        """Obtener información del sistema de puntos"""
        try:
            from apps.recognition.models import HistorialPuntos, InsigniaEmpleado
            
            puntos_totales = HistorialPuntos.objects.filter(empleado=empleado).aggregate(
                total=Sum('puntos')
            )['total'] or 0
            
            insignias = InsigniaEmpleado.objects.filter(empleado=empleado).count()
            
            return {
                'total': puntos_totales,
                'insignias': insignias,
            }
        except:
            return {'total': 0, 'insignias': 0}
    
    def _get_actividad_reciente(self, empleado):
        """Obtener actividad reciente del empleado"""
        actividades = []
        
        try:
            # Documentos recientes
            docs_recientes = DocumentoEmpleado.objects.filter(
                empleado=empleado
            ).order_by('-fecha_subida')[:5]
            
            for doc in docs_recientes:
                actividades.append({
                    'tipo': 'documento',
                    'titulo': f'Documento subido: {doc.tipo_documento.nombre}',
                    'fecha': doc.fecha_subida,
                    'estado': doc.estado,
                    'icono': 'fas fa-file-upload'
                })
            
            # Capacitaciones recientes
            caps_recientes = InscripcionCapacitacion.objects.filter(
                empleado=empleado
            ).order_by('-fecha_inscripcion')[:5]
            
            for cap in caps_recientes:
                actividades.append({
                    'tipo': 'capacitacion',
                    'titulo': f'Capacitación: {cap.capacitacion.titulo}',
                    'fecha': cap.fecha_inscripcion,
                    'estado': cap.estado,
                    'icono': 'fas fa-graduation-cap'
                })
            
            # Ordenar por fecha
            actividades.sort(key=lambda x: x['fecha'], reverse=True)
            
        except Exception as e:
            logger.error(f"Error obteniendo actividad reciente: {e}")
        
        return actividades[:10]


@login_required
def empleado_perfil_redirect(request):
    """Vista para redirigir al perfil del empleado logueado"""
    try:
        empleado = get_object_or_404(Empleado, usuario=request.user)
        return redirect('employees:empleado_perfil_detail', pk=empleado.pk)
    except Empleado.DoesNotExist:
        messages.error(
            request, 
            'No se encontró un perfil de empleado asociado a tu usuario. Contacta al administrador.'
        )
        return redirect('core:dashboard')
    except Exception as e:
        logger.error(f"Error en empleado_perfil_redirect: {e}")
        messages.error(request, 'Error al acceder al perfil del empleado.')
        return redirect('core:dashboard')


@login_required
@require_POST
def cambiar_cargo_empleado(request, pk):
    """Vista para cambiar el cargo de un empleado"""
    try:
        empleado = get_object_or_404(Empleado, pk=pk)
        
        # Verificar permisos (solo staff puede cambiar cargos)
        if not request.user.is_staff:
            return JsonResponse({'success': False, 'message': 'No tienes permisos para realizar esta acción'})
        
        # Obtener datos del formulario
        nuevo_cargo_id = request.POST.get('nuevo_cargo')
        nueva_sede_id = request.POST.get('nueva_sede')
        fecha_inicio = request.POST.get('fecha_inicio')
        salario = request.POST.get('salario')
        motivo = request.POST.get('motivo', '')
        
        # Validaciones
        if not all([nuevo_cargo_id, nueva_sede_id, fecha_inicio]):
            return JsonResponse({'success': False, 'message': 'Todos los campos obligatorios deben ser completados'})
        
        try:
            nuevo_cargo = Cargo.objects.get(pk=nuevo_cargo_id, activo=True)
            nueva_sede = Sede.objects.get(pk=nueva_sede_id, activa=True)
            
            # Salario es opcional
            salario_valor = None
            if salario and salario.strip():
                salario_valor = float(salario)
                if salario_valor <= 0:
                    return JsonResponse({'success': False, 'message': 'El salario debe ser mayor a 0'})
            else:
                # Si no se proporciona salario, mantener el actual
                cargo_actual = empleado.historialcargo_set.filter(activo=True).first()
                if cargo_actual and cargo_actual.salario:
                    salario_valor = cargo_actual.salario
                
        except (Cargo.DoesNotExist, Sede.DoesNotExist):
            return JsonResponse({'success': False, 'message': 'Cargo o sede no válidos'})
        except ValueError:
            return JsonResponse({'success': False, 'message': 'Salario no válido'})
        
        # Verificar que no sea el mismo cargo actual
        cargo_actual_check = empleado.historialcargo_set.filter(activo=True).first()
        if cargo_actual_check and cargo_actual_check.cargo.id == int(nuevo_cargo_id):
            return JsonResponse({'success': False, 'message': 'El empleado ya tiene asignado este cargo'})
        
        # Realizar el cambio de cargo en una transacción
        with transaction.atomic():
            # Desactivar cargo actual
            if cargo_actual_check:
                cargo_actual_check.activo = False
                cargo_actual_check.fecha_fin = fecha_inicio
                cargo_actual_check.save()
            
            # Crear nuevo historial de cargo
            nuevo_historial = HistorialCargo.objects.create(
                empleado=empleado,
                cargo=nuevo_cargo,
                fecha_inicio=fecha_inicio,
                salario=salario_valor,
                motivo_cambio=motivo,
                activo=True,
                creado_por=request.user
            )
            
            # Actualizar datos del empleado
            empleado.cargo = nuevo_cargo
            empleado.sede = nueva_sede
            if salario_valor:
                empleado.salario = salario_valor
            empleado.modificado_por = request.user
            empleado.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Cargo cambiado exitosamente a {nuevo_cargo.nombre}'
        })
        
    except Exception as e:
        logger.error(f"Error al cambiar cargo del empleado {pk}: {e}")
        return JsonResponse({'success': False, 'message': 'Error interno del servidor'})


@login_required
def empleados_periodo_prueba_reporte(request):
    """Vista para mostrar empleados en periodo de prueba y sus fechas de activación automática"""
    
    try:
        # Obtener estado de periodo de prueba
        estado_prueba = EstadoEmpleado.objects.get(codigo='p-prue')
    except EstadoEmpleado.DoesNotExist:
        messages.error(request, 'No se encontró el estado de periodo de prueba')
        return redirect('employees:empleado_list')
    
    # Obtener días de antelación de los parámetros GET o usar 7 por defecto
    dias_antelacion = int(request.GET.get('dias_antelacion', 7))
    if dias_antelacion < 1 or dias_antelacion > 30:
        dias_antelacion = 7
    
    # Obtener empleados en periodo de prueba
    empleados_prueba = Empleado.objects.filter(
        estado=estado_prueba
    ).select_related(
        'sede', 'tipo_documento', 'estado'
    ).prefetch_related(
        'historialcargo_set__cargo__area'
    ).order_by('fecha_ingreso')
    
        # Calcular información de activación para cada empleado
    empleados_info = []
    hoy = timezone.now().date()
    
    for empleado in empleados_prueba:
        dias_transcurridos = (hoy - empleado.fecha_ingreso).days
        dias_restantes = 60 - dias_transcurridos
        fecha_activacion = empleado.fecha_ingreso + timedelta(days=60)
        
        # Determinar estado de activación
        if dias_transcurridos >= 60:
            estado_activacion = 'Listo para activar'
        elif dias_restantes <= dias_antelacion:
            estado_activacion = f'Próximo a activar'
        else:
            estado_activacion = f'En período de prueba'
        
        # Obtener información de evaluación si existe
        evaluacion_info = None
        try:
            from apps.evaluations.models import AsignacionEvaluacion
            evaluacion = AsignacionEvaluacion.objects.filter(
                empleado_evaluado=empleado,
                estado='completada'
            ).first()
            
            if evaluacion:
                evaluacion_info = {
                    'existe': True,
                    'puntaje': evaluacion.puntaje_total,
                    'aprobado': evaluacion.puntaje_total >= 14 if evaluacion.puntaje_total else False,
                    'estado_aprobacion': evaluacion.estado_aprobacion,
                    'fecha_completada': evaluacion.fecha_completada,
                    'requiere_desactivacion': (
                        evaluacion.estado_aprobacion == 'aprobada' and 
                        evaluacion.puntaje_total and 
                        evaluacion.puntaje_total <= 13
                    )
                }
            else:
                evaluacion_info = {'existe': False}
        except:
            evaluacion_info = {'existe': False}
        
        empleados_info.append({
            'empleado': empleado,
            'dias_transcurridos': dias_transcurridos,
            'dias_restantes': max(0, dias_restantes),
            'fecha_activacion': fecha_activacion,
            'estado_activacion': estado_activacion,
            'cargo_actual': empleado.cargo_actual,
            'evaluacion': evaluacion_info,
        })
    
    # Estadísticas generales
    total_empleados = len(empleados_info)
    listos_activar = len([e for e in empleados_info if e['dias_restantes'] == 0])
    proximos_activar = len([e for e in empleados_info if 0 < e['dias_restantes'] <= dias_antelacion])
    
    context = {
        'empleados_info': empleados_info,
        'total_empleados': total_empleados,
        'listos_activar': listos_activar,
        'proximos_activar': proximos_activar,
        'fecha_actual': hoy,
        'dias_antelacion': dias_antelacion,
    }
    
    return render(request, 'employees/periodo_prueba_reporte.html', context)


@login_required
@require_POST
def activar_empleado_individual(request, pk):
    """Activa individualmente un empleado que cumplió el período de prueba"""
    try:
        # Obtener empleado
        empleado = get_object_or_404(Empleado, pk=pk)
        
        # Verificar que esté en período de prueba
        estado_prueba = EstadoEmpleado.objects.get(codigo='p-prue')
        if empleado.estado != estado_prueba:
            return JsonResponse({
                'success': False,
                'message': 'El empleado no está en período de prueba'
            })
        
        # Verificar que haya cumplido los 60 días
        dias_transcurridos = (timezone.now().date() - empleado.fecha_ingreso).days
        if dias_transcurridos < 60:
            return JsonResponse({
                'success': False,
                'message': f'El empleado aún no cumple los 60 días. Días transcurridos: {dias_transcurridos}'
            })
        
        # Obtener estado activo
        estado_activo = EstadoEmpleado.objects.get(codigo='999')
        
        # Cambiar estado
        empleado.estado = estado_activo
        empleado.save()
        
        # Log de activación manual
        logger.info(
            f'ACTIVACIÓN MANUAL: Empleado {empleado.numero_documento} '
            f'({empleado.nombre_completo}) activado manualmente por {request.user.username} '
            f'después de {dias_transcurridos} días en período de prueba'
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Empleado {empleado.nombre_completo} activado exitosamente',
            'empleado_nombre': empleado.nombre_completo,
            'dias_transcurridos': dias_transcurridos
        })
        
    except EstadoEmpleado.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Error: Estados de empleado no configurados correctamente'
        })
    except Exception as e:
        logger.error(f"Error al activar empleado {pk}: {e}")
        return JsonResponse({
            'success': False,
            'message': 'Error interno del servidor'
        })


@require_POST
@login_required
def desactivar_empleado_reprobado(request, pk):
    """Desactiva un empleado que reprobó la evaluación de período de prueba"""
    try:
        # Obtener empleado
        empleado = get_object_or_404(Empleado, pk=pk)
        
        # Verificar que esté en período de prueba
        estado_prueba = EstadoEmpleado.objects.get(codigo='p-prue')
        if empleado.estado != estado_prueba:
            return JsonResponse({
                'success': False,
                'message': 'El empleado no está en período de prueba'
            })
        
        # Verificar que tenga una evaluación reprobada
        try:
            from apps.evaluations.models import AsignacionEvaluacion
            evaluacion = AsignacionEvaluacion.objects.filter(
                empleado_evaluado=empleado,
                estado='completada',
                estado_aprobacion='aprobada'
            ).first()
            
            if not evaluacion or not evaluacion.puntaje_total or evaluacion.puntaje_total > 13:
                return JsonResponse({
                    'success': False,
                    'message': 'El empleado no tiene una evaluación reprobada que justifique la desactivación'
                })
        except Exception as e:
            logger.error(f"Error al verificar evaluación: {e}")
            return JsonResponse({
                'success': False,
                'message': 'Error al verificar la evaluación del empleado'
            })
        
        # Obtener estado inactivo
        estado_inactivo = EstadoEmpleado.objects.get(codigo='000')
        
        # Desactivar empleado
        empleado.estado = estado_inactivo
        empleado.observaciones = f'Desactivado por reprobación en evaluación de período de prueba. Puntaje: {evaluacion.puntaje_total}/21 puntos. Fecha: {timezone.now().date()}'
        empleado.save()
        
        # Log de desactivación
        logger.info(
            f'DESACTIVACIÓN POR REPROBACIÓN: Empleado {empleado.numero_documento} '
            f'({empleado.nombre_completo}) desactivado por {request.user.username} '
            f'debido a evaluación reprobada ({evaluacion.puntaje_total}/21 puntos)'
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Empleado {empleado.nombre_completo} desactivado por reprobación',
            'empleado_nombre': empleado.nombre_completo,
            'puntaje': evaluacion.puntaje_total,
            'motivo': 'Evaluación reprobada'
        })
        
    except EstadoEmpleado.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Error: Estados de empleado no configurados correctamente'
        })
    except Exception as e:
        logger.error(f"Error al desactivar empleado reprobado {pk}: {e}")
        return JsonResponse({
            'success': False,
            'message': 'Error interno del servidor'
        })


# Agregar import para agregaciones
from django.db import models


@login_required
def obtener_jefes_potenciales(request, cargo_id):
    """
    API para obtener los posibles jefes directos según el cargo seleccionado.
    Retorna lista de empleados que ocupan el cargo_jefe del cargo dado.
    """
    try:
        cargo = get_object_or_404(Cargo, pk=cargo_id, activo=True)

        # Si el cargo no tiene cargo_jefe definido, no hay jefes potenciales
        if not cargo.cargo_jefe:
            return JsonResponse({
                'success': True,
                'tiene_cargo_jefe': False,
                'cargo_jefe_nombre': None,
                'jefes': [],
                'mensaje': 'Este cargo no tiene un cargo jefe definido'
            })

        # Buscar empleados activos que ocupen el cargo_jefe
        jefes = Empleado.objects.filter(
            historialcargo__cargo=cargo.cargo_jefe,
            historialcargo__activo=True,
            estado__codigo__in=['999', 'p-prue']
        ).distinct().select_related('estado').prefetch_related(
            Prefetch(
                'historialcargo_set',
                queryset=HistorialCargo.objects.filter(activo=True).select_related('cargo'),
                to_attr='cargo_actual_list'
            )
        )

        jefes_lista = []
        for jefe in jefes:
            cargo_actual = jefe.cargo_actual_list[0] if jefe.cargo_actual_list else None
            jefes_lista.append({
                'id': str(jefe.pk),
                'nombre_completo': jefe.nombre_completo,
                'documento': jefe.numero_documento,
                'cargo': cargo_actual.cargo.nombre if cargo_actual else 'Sin cargo',
                'area': cargo_actual.cargo.area.nombre if cargo_actual and cargo_actual.cargo.area else ''
            })

        return JsonResponse({
            'success': True,
            'tiene_cargo_jefe': True,
            'cargo_jefe_nombre': cargo.cargo_jefe.nombre,
            'jefes': jefes_lista,
            'cantidad': len(jefes_lista),
            'mensaje': f'Se encontraron {len(jefes_lista)} posibles jefes'
        })

    except Exception as e:
        logger.error(f"Error obteniendo jefes potenciales para cargo {cargo_id}: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e),
            'mensaje': 'Error al obtener los jefes potenciales'
        }, status=500)


# =============================================================================
# MARKETPLACE - PRODUCTOS
# =============================================================================

class ProductoListView(EmpleadoRequiredMixin, ListView):
    """Listar productos del marketplace con filtros"""
    model = Producto
    paginate_by = 12
    context_object_name = 'productos'
    template_name = 'employees/marketplace/producto_list.html'

    def get_queryset(self):
        """Filtrar productos activos y visibles"""
        queryset = Producto.objects.filter(estado='activo').select_related('vendedor', 'categoria')

        # Filtro por tipo
        tipo = self.request.GET.get('tipo')
        if tipo in ['venta', 'regalo', 'subasta']:
            queryset = queryset.filter(tipo=tipo)

        # Filtro por categoría
        categoria_id = self.request.GET.get('categoria')
        if categoria_id:
            queryset = queryset.filter(categoria_id=categoria_id)

        # Búsqueda por título o descripción
        search = self.request.GET.get('search')
        if search:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(titulo__icontains=search) | Q(descripcion__icontains=search)
            )

        return queryset.order_by('-fecha_creacion')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categorias'] = Categoria.objects.filter(activa=True)
        context['tipos'] = Producto.TIPO_CHOICES
        context['search'] = self.request.GET.get('search', '')
        context['tipo_filter'] = self.request.GET.get('tipo', '')
        context['categoria_filter'] = self.request.GET.get('categoria', '')
        return context


class ProductoDetailView(EmpleadoRequiredMixin, DetailView):
    """Ver detalles de un producto"""
    model = Producto
    context_object_name = 'producto'
    template_name = 'employees/marketplace/producto_detail.html'

    def get_queryset(self):
        return Producto.objects.select_related('vendedor', 'categoria')

    def get(self, request, *args, **kwargs):
        """Redirigir a subasta_detail si el producto es tipo subasta"""
        self.object = self.get_object()

        # Si el producto es tipo subasta, redirigir a la vista de subasta
        if self.object.tipo == 'subasta':
            logger.info(f"[PRODUCTO_DETAIL] Producto tipo subasta detectado: {self.object.titulo}")
            subasta = self.object.subastas.filter(estado='activa').first()

            if subasta:
                logger.info(f"[PRODUCTO_DETAIL] Redirigiendo a subasta activa: {subasta.pk}")
                return redirect('employees:subasta_detail', pk=subasta.pk)
            else:
                # No hay subasta activa, verificar si hay alguna subasta
                todas_subastas = self.object.subastas.all()
                logger.warning(f"[PRODUCTO_DETAIL] No hay subasta activa. Total subastas: {todas_subastas.count()}")
                if todas_subastas.exists():
                    logger.warning(f"[PRODUCTO_DETAIL] Estados de subastas: {[s.estado for s in todas_subastas]}")

                messages.warning(request, 'Este producto está configurado como subasta pero no tiene una subasta activa.')

        # Si no es subasta o no tiene subasta activa, mostrar vista normal
        context = self.get_context_data(object=self.object)
        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        producto = self.object

        # Solo agregamos información si el usuario tiene un perfil de empleado
        try:
            usuario = self.request.user.empleado

            # Agregar información según tipo
            if producto.tipo == 'venta':
                context['puede_comprar'] = usuario.id != producto.vendedor.id
                # Si es el vendedor, mostrar lista de reservas
                if usuario.id == producto.vendedor.id:
                    context['reservas_activas'] = producto.reservas.filter(estado='activa').select_related('comprador').order_by('-fecha_creacion')
            elif producto.tipo == 'subasta':
                context['subasta'] = producto.subastas.filter(estado='activa').first()
            elif producto.tipo == 'regalo':
                context['puede_recibir'] = usuario.id != producto.vendedor.id
                # Si es el vendedor/donante, mostrar solicitudes de regalo
                if usuario.id == producto.vendedor.id:
                    context['solicitudes_regalo'] = producto.regalos.all().order_by('-fecha_ofrecimiento')

            # Agregar información de reservaciones (solo para el vendedor)
            if usuario.id == producto.vendedor.id:
                context['reservaciones_activas'] = producto.get_cantidad_reservada()
                context['es_vendedor'] = True
        except Empleado.DoesNotExist:
            # Los superusuarios/admin pueden ver y comprar si es un producto tipo venta
            context['puede_comprar'] = (self.request.user.is_superuser or self.request.user.is_staff) and producto.tipo == 'venta'
            context['puede_recibir'] = (self.request.user.is_superuser or self.request.user.is_staff) and producto.tipo == 'regalo'
            context['es_vendedor'] = False

        return context


class CrearProductoView(EmpleadoRequiredMixin, CreateView):
    """Crear nuevo producto"""
    model = Producto
    form_class = ProductoForm
    template_name = 'employees/marketplace/producto_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['es_creacion'] = True
        return context

    def get_form(self, form_class=None):
        """Personalizar el formulario según el tipo de usuario"""
        form = super().get_form(form_class)

        if not (self.request.user.is_superuser or self.request.user.is_staff):
            # Para empleados normales: esconder el campo vendedor y usar el empleado actual
            form.fields['vendedor'].widget = forms.HiddenInput()
            try:
                form.initial['vendedor'] = self.request.user.empleado
                form.fields['vendedor'].required = False
            except Empleado.DoesNotExist:
                # Si no tiene empleado, mostrar error
                messages.error(self.request, 'Debes tener un perfil de empleado para crear productos.')

        return form

    def form_valid(self, form):
        """Validar y procesar el formulario"""
        # El vendedor ya viene del formulario
        form.instance.creado_por = self.request.user
        response = super().form_valid(form)

        # Si el producto es tipo subasta, crear automáticamente la subasta
        if self.object.tipo == 'subasta':
            from datetime import timedelta
            from django.utils import timezone

            # Usar precio_inicial del producto o un valor por defecto
            precio = self.object.precio_inicial or 10000

            # Crear subasta con valores por defecto
            subasta = Subasta.objects.create(
                producto=self.object,
                vendedor=self.object.vendedor,
                precio_inicial=precio,
                precio_actual=precio,
                incremento_minimo=1000,  # $1,000 de incremento mínimo por defecto
                fecha_inicio=timezone.now(),
                fecha_fin=timezone.now() + timedelta(days=7),  # 7 días por defecto
                estado='activa',
                creado_por=self.request.user
            )
            logger.info(f"[CREAR_PRODUCTO] Subasta creada automáticamente: {subasta.pk} para producto {self.object.titulo}")
            messages.success(self.request, f'Producto creado como subasta. Finaliza en 7 días.')

        return response

    def get_success_url(self):
        return reverse('employees:producto_detail', kwargs={'pk': self.object.pk})


class EditarProductoView(EmpleadoRequiredMixin, UpdateView):
    """Editar un producto existente"""
    model = Producto
    form_class = ProductoForm
    template_name = 'employees/marketplace/producto_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['es_creacion'] = False
        return context

    def get_queryset(self):
        """Solo el vendedor o admin pueden editar"""
        user = self.request.user
        if user.is_superuser or user.is_staff:
            return Producto.objects.all()
        try:
            return Producto.objects.filter(vendedor=user.empleado)
        except Empleado.DoesNotExist:
            return Producto.objects.none()

    def get_form(self, form_class=None):
        """Personalizar el formulario según el tipo de usuario"""
        form = super().get_form(form_class)

        if not (self.request.user.is_superuser or self.request.user.is_staff):
            # Para empleados normales: esconder el campo vendedor
            form.fields['vendedor'].widget = forms.HiddenInput()
            form.fields['vendedor'].required = False

        return form

    def form_valid(self, form):
        """Validar y procesar el formulario"""
        response = super().form_valid(form)

        # Si el producto es tipo subasta y no tiene subasta activa, crearla
        if self.object.tipo == 'subasta':
            if not self.object.subastas.filter(estado='activa').exists():
                from datetime import timedelta
                from django.utils import timezone

                # Usar precio_inicial del producto o un valor por defecto
                precio = self.object.precio_inicial or 10000

                # Crear subasta con valores por defecto
                subasta = Subasta.objects.create(
                    producto=self.object,
                    vendedor=self.object.vendedor,
                    precio_inicial=precio,
                    precio_actual=precio,
                    incremento_minimo=1000,  # $1,000 de incremento mínimo por defecto
                    fecha_inicio=timezone.now(),
                    fecha_fin=timezone.now() + timedelta(days=7),  # 7 días por defecto
                    estado='activa',
                    creado_por=self.request.user
                )
                logger.info(f"[EDITAR_PRODUCTO] Subasta creada para producto existente: {subasta.pk}")
                messages.success(self.request, f'Subasta activada. Finaliza en 7 días.')

        return response

    def get_success_url(self):
        return reverse('employees:producto_detail', kwargs={'pk': self.object.pk})


class EliminarProductoView(EmpleadoRequiredMixin, DeleteView):
    """Eliminar un producto"""
    model = Producto
    template_name = 'employees/marketplace/producto_confirm_delete.html'
    context_object_name = 'producto'

    def get_queryset(self):
        """Solo el vendedor o admin pueden eliminar"""
        user = self.request.user
        queryset = Producto.objects.all()

        # Admin puede eliminar cualquier producto
        if user.is_superuser or user.is_staff:
            return queryset

        # Empleados normales solo pueden eliminar sus propios productos
        try:
            return queryset.filter(vendedor=user.empleado)
        except Empleado.DoesNotExist:
            return queryset.none()

    def delete(self, request, *args, **kwargs):
        """Mostrar mensaje de confirmación"""
        messages.success(request, 'Producto eliminado correctamente.')
        return super().delete(request, *args, **kwargs)

    def get_success_url(self):
        return reverse('employees:producto_list')


class MisProductosView(EmpleadoRequiredMixin, ListView):
    """Mis productos creados"""
    model = Producto
    paginate_by = 12
    context_object_name = 'productos'
    template_name = 'employees/marketplace/mis_productos.html'

    def get_queryset(self):
        return Producto.objects.filter(
            vendedor=self.request.user.empleado
        ).select_related('categoria').order_by('-fecha_creacion')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        empleado = self.request.user.empleado
        context['total_productos'] = self.get_queryset().count()
        context['productos_activos'] = self.get_queryset().filter(estado='activo').count()
        context['productos_vendidos'] = self.get_queryset().filter(estado='vendido').count()

        # Agregar información de reservas para cada producto
        from apps.employees.models import Reserva
        productos_con_reservas = {}
        for producto in context.get('productos', []):
            reservas_activas = producto.reservas.filter(estado='activa').select_related('comprador')
            productos_con_reservas[producto.pk] = reservas_activas
        context['productos_con_reservas'] = productos_con_reservas

        return context


class MisComprasView(EmpleadoRequiredMixin, ListView):
    """Mi historial de compras"""
    model = Venta
    paginate_by = 12
    context_object_name = 'compras'
    template_name = 'employees/marketplace/mis_compras.html'

    def get_queryset(self):
        return Venta.objects.filter(
            comprador=self.request.user.empleado
        ).select_related('producto', 'vendedor', 'reserva_origen').order_by('-fecha_venta')

    def get_context_data(self, **kwargs):
        from apps.employees.models import Reserva
        context = super().get_context_data(**kwargs)
        queryset = self.get_queryset()
        empleado = self.request.user.empleado

        # Estadísticas de compras completadas (considerando cantidades)
        context['total_compras'] = queryset.count()
        total_gastado = 0
        for venta in queryset:
            cantidad = venta.reserva_origen.cantidad_solicitada if hasattr(venta, 'reserva_origen') and venta.reserva_origen else 1
            total_gastado += venta.precio_final * cantidad
        context['total_gastado'] = total_gastado
        context['compras_completadas'] = queryset.filter(estado='completada').count()

        # Reservas activas del comprador (productos separados pendientes)
        reservas_activas = Reserva.objects.filter(
            comprador=empleado,
            estado='activa'
        ).select_related('producto', 'producto__vendedor', 'producto__categoria').order_by('-fecha_creacion')

        context['reservas_activas'] = reservas_activas
        context['total_reservas_activas'] = reservas_activas.count()
        context['total_unidades_reservadas'] = sum(r.cantidad_solicitada for r in reservas_activas)

        return context


class ComprarProductoView(EmpleadoRequiredMixin, CreateView):
    """Flujo de compra de un producto"""
    model = Venta
    form_class = VentaForm
    template_name = 'employees/marketplace/comprar_producto.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['vendedor'] = self.get_producto().vendedor
        kwargs['comprador'] = self.request.user.empleado
        return kwargs

    def get_producto(self):
        """Obtener producto a comprar"""
        if not hasattr(self, '_producto'):
            self._producto = get_object_or_404(Producto, pk=self.kwargs['producto_pk'])
        return self._producto

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['producto'] = self.get_producto()
        return context

    def form_valid(self, form):
        """Procesar compra"""
        form.instance.producto = self.get_producto()
        form.instance.vendedor = self.get_producto().vendedor
        form.instance.comprador = self.request.user.empleado
        form.instance.creado_por = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('employees:mis_compras')


# =============================================================================
# MARKETPLACE - SUBASTAS
# =============================================================================

class SubastaListView(EmpleadoRequiredMixin, ListView):
    """Listar subastas activas"""
    model = Subasta
    paginate_by = 12
    context_object_name = 'subastas'
    template_name = 'employees/marketplace/subasta_list.html'

    def get_queryset(self):
        return Subasta.objects.filter(
            estado='activa'
        ).select_related('producto', 'vendedor').order_by('-fecha_inicio')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_subastas'] = self.get_queryset().count()
        return context


class PujarView(EmpleadoRequiredMixin, CreateView):
    """Realizar puja en subasta"""
    model = PujaSubasta
    form_class = PujaForm
    template_name = 'employees/marketplace/puja_form.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['subasta'] = self.get_subasta()
        kwargs['pujador'] = self.request.user.empleado
        return kwargs

    def get_subasta(self):
        """Obtener subasta"""
        if not hasattr(self, '_subasta'):
            self._subasta = get_object_or_404(Subasta, pk=self.kwargs['subasta_pk'])
        return self._subasta

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        subasta = self.get_subasta()
        context['subasta'] = subasta
        context['producto'] = subasta.producto
        context['precio_minimo'] = subasta.precio_actual + subasta.incremento_minimo
        return context

    def form_valid(self, form):
        """Procesar puja"""
        subasta = self.get_subasta()
        form.instance.subasta = subasta
        form.instance.pujador = self.request.user.empleado
        form.instance.creado_por = self.request.user

        # Guardar puja
        response = super().form_valid(form)

        # Actualizar precio actual de la subasta
        subasta.precio_actual = form.instance.monto
        subasta.pujador_actual = self.request.user.empleado
        subasta.save()

        return response

    def get_success_url(self):
        return reverse('employees:subasta_detail', kwargs={'pk': self.kwargs['subasta_pk']})


class SubastaDetailView(EmpleadoRequiredMixin, DetailView):
    """Ver detalles de subasta"""
    model = Subasta
    context_object_name = 'subasta'
    template_name = 'employees/marketplace/subasta_detail.html'

    def get_queryset(self):
        return Subasta.objects.select_related('producto', 'vendedor')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        subasta = self.object
        context['pujas'] = subasta.pujas.select_related('pujador').order_by('-fecha_creacion')
        context['precio_minimo'] = subasta.precio_actual + subasta.incremento_minimo

        # Verificar si el usuario puede pujar
        puede_pujar = False
        try:
            empleado_actual = self.request.user.empleado
            # No es el vendedor y está activo
            puede_pujar = (
                empleado_actual.id != subasta.vendedor.id and
                empleado_actual.estado.codigo == '999'
            )
        except (AttributeError, Empleado.DoesNotExist):
            # Usuario sin empleado asociado no puede pujar
            puede_pujar = False

        context['puede_pujar'] = puede_pujar

        # Agregar formulario de puja si puede pujar
        if puede_pujar:
            from .forms import PujaForm
            context['form'] = PujaForm()

        return context


# =============================================================================
# MARKETPLACE - REGALOS
# =============================================================================

class RegalarProductoView(EmpleadoRequiredMixin, CreateView):
    """Regalar un producto"""
    model = Regalo
    form_class = RegaloForm
    template_name = 'employees/marketplace/regalar_producto.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['donante'] = self.request.user.empleado
        return kwargs

    def get_producto(self):
        """Obtener producto a regalar"""
        if not hasattr(self, '_producto'):
            self._producto = get_object_or_404(Producto, pk=self.kwargs['producto_pk'])
        return self._producto

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['producto'] = self.get_producto()
        return context

    def form_valid(self, form):
        """Procesar regalo"""
        form.instance.producto = self.get_producto()
        form.instance.donante = self.request.user.empleado
        form.instance.creado_por = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('employees:mis_productos')


class ReceibirRegaloAjaxView(EmpleadoRequiredMixin, View):
    """Vista AJAX para que un usuario solicite un regalo disponible en el marketplace"""

    def post(self, request, *args, **kwargs):
        """POST: Usuario solicita recibir un regalo"""
        try:
            producto_pk = kwargs.get('producto_pk')
            producto = get_object_or_404(Producto, pk=producto_pk, tipo='regalo', estado='activo')

            receptor = request.user.empleado

            # Validar que no es el donante (vendedor)
            if producto.vendedor == receptor:
                return JsonResponse({
                    'success': False,
                    'error': 'No puedes recibir tu propio regalo'
                }, status=400)

            # Validar que hay regalos disponibles
            if not producto.tiene_disponible():
                return JsonResponse({
                    'success': False,
                    'error': 'No hay regalos disponibles'
                }, status=400)

            # Validar que el usuario no ya solicitó este regalo
            regalo_existente = Regalo.objects.filter(
                producto=producto,
                receptor=receptor,
                estado__in=['pendiente', 'aceptado']
            ).exists()

            if regalo_existente:
                return JsonResponse({
                    'success': False,
                    'error': 'Ya has solicitado este regalo'
                }, status=400)

            # Crear regalo con estado "pendiente" (usuario solicitó, esperando confirmación del donante)
            regalo = Regalo.objects.create(
                producto=producto,
                donante=producto.vendedor,
                receptor=receptor,
                estado='pendiente',
                creado_por=request.user
            )

            logger.info(f"[REGALO] {receptor.nombre_completo} solicitó el regalo '{producto.titulo}' de {producto.vendedor.nombre_completo}")

            return JsonResponse({
                'success': True,
                'message': f'¡Has solicitado el regalo "{producto.titulo}"! El donante será notificado.',
                'regalo_id': str(regalo.id)
            })

        except Exception as e:
            logger.error(f"[REGALO] Error al solicitar regalo: {e}", exc_info=True)
            return JsonResponse({
                'success': False,
                'error': 'Error al solicitar el regalo'
            }, status=500)


class MisRegalosView(EmpleadoRequiredMixin, ListView):
    """Vista para que donantes vean y gestionen solicitudes de regalo"""
    model = Regalo
    paginate_by = 12
    context_object_name = 'regalos'
    template_name = 'employees/marketplace/mis_regalos.html'

    def get_queryset(self):
        """Regalos donde el usuario es donante"""
        empleado = self.request.user.empleado
        return Regalo.objects.filter(
            donante=empleado
        ).select_related('producto', 'receptor').order_by('-fecha_ofrecimiento')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = self.get_queryset()

        # Estadísticas
        context['total_regalos'] = queryset.count()
        context['regalos_pendientes'] = queryset.filter(estado='pendiente').count()
        context['regalos_entregados'] = queryset.filter(estado='aceptado').count()
        context['regalos_cancelados'] = queryset.filter(estado='cancelado').count()

        # Agrupar por producto para mostrar quién ha solicitado qué
        regalos_por_producto = {}
        for regalo in queryset.filter(estado__in=['pendiente', 'aceptado']):
            if regalo.producto.pk not in regalos_por_producto:
                regalos_por_producto[regalo.producto.pk] = {
                    'producto': regalo.producto,
                    'solicitantes': []
                }
            regalos_por_producto[regalo.producto.pk]['solicitantes'].append(regalo)

        context['regalos_por_producto'] = regalos_por_producto

        return context


class ConfirmarEntregaRegaloAjaxView(EmpleadoRequiredMixin, View):
    """Vista AJAX para que donante confirme la entrega del regalo"""

    def post(self, request, **kwargs):
        """POST: Donante confirma entrega del regalo"""
        try:
            regalo_id = kwargs.get('regalo_id')
            regalo = get_object_or_404(Regalo, id=regalo_id)

            # Validar que el usuario es el donante
            if regalo.donante != request.user.empleado:
                return JsonResponse({
                    'success': False,
                    'error': 'No tienes permiso para confirmar este regalo'
                }, status=403)

            # Cambiar estado a 'aceptado' y marcar como confirmado
            regalo.estado = 'aceptado'
            regalo.confirmado_por_donante = True
            regalo.fecha_aceptacion = timezone.now()
            regalo.save()

            logger.info(f"[REGALO] {request.user.empleado.nombre_completo} confirmó entrega del regalo a {regalo.receptor.nombre_completo}")

            return JsonResponse({
                'success': True,
                'message': f'Entrega confirmada para {regalo.receptor.nombre_completo}'
            })

        except Exception as e:
            logger.error(f"[REGALO] Error al confirmar entrega: {e}", exc_info=True)
            return JsonResponse({
                'success': False,
                'error': 'Error al confirmar la entrega'
            }, status=500)


class RevertirSolicitudRegaloAjaxView(EmpleadoRequiredMixin, View):
    """Vista AJAX para que donante revierta solicitud de regalo"""

    def post(self, request, **kwargs):
        """POST: Donante revierte solicitud de regalo"""
        try:
            regalo_id = kwargs.get('regalo_id')
            regalo = get_object_or_404(Regalo, id=regalo_id)

            # Validar que el usuario es el donante
            if regalo.donante != request.user.empleado:
                return JsonResponse({
                    'success': False,
                    'error': 'No tienes permiso para revertir este regalo'
                }, status=403)

            # Cambiar estado a 'cancelado'
            regalo.estado = 'cancelado'
            regalo.save()

            logger.info(f"[REGALO] {request.user.empleado.nombre_completo} revirtió solicitud de {regalo.receptor.nombre_completo}")

            return JsonResponse({
                'success': True,
                'message': f'Solicitud de {regalo.receptor.nombre_completo} revertida'
            })

        except Exception as e:
            logger.error(f"[REGALO] Error al revertir solicitud: {e}", exc_info=True)
            return JsonResponse({
                'success': False,
                'error': 'Error al revertir la solicitud'
            }, status=500)


# =============================================================================
# MESSAGING - CONVERSACIONES
# =============================================================================

class InboxView(EmpleadoRequiredMixin, ListView):
    """Inbox de conversaciones"""
    model = Conversacion
    paginate_by = 20
    context_object_name = 'conversaciones'
    template_name = 'employees/messaging/inbox.html'

    def get_queryset(self):
        """Conversaciones del usuario actual"""
        # Validar que el usuario tiene empleado (por si el mixin no alcanza)
        try:
            empleado = self.request.user.empleado
        except (AttributeError, Empleado.DoesNotExist):
            if not (self.request.user.is_staff or self.request.user.is_superuser):
                messages.error(
                    self.request,
                    'Tu cuenta de usuario no está asociada a un registro de empleado. '
                    'Contacta con Recursos Humanos para completar tu perfil.'
                )
                return Conversacion.objects.none()
            # Los admins ven todas las conversaciones
            return Conversacion.objects.exclude(
                archivada=True
            ).select_related('creado_por').prefetch_related('participantes').order_by('-fecha_ultima_actividad')

        return Conversacion.objects.filter(
            participantes=empleado
        ).exclude(
            archivada=True
        ).select_related('creado_por').prefetch_related('participantes').order_by('-fecha_ultima_actividad')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_conversaciones'] = self.get_queryset().count()
        return context


class ConversacionDetailView(EmpleadoRequiredMixin, DetailView):
    """Ver conversación detallada"""
    model = Conversacion
    context_object_name = 'conversacion'
    template_name = 'employees/messaging/conversacion_detail.html'

    def get_queryset(self):
        """Solo conversaciones donde es participante"""
        try:
            empleado = self.request.user.empleado
        except (AttributeError, Empleado.DoesNotExist):
            if not (self.request.user.is_staff or self.request.user.is_superuser):
                messages.error(
                    self.request,
                    'Tu cuenta de usuario no está asociada a un registro de empleado. '
                    'Contacta con Recursos Humanos para completar tu perfil.'
                )
                return Conversacion.objects.none()
            # Los admins ven todas las conversaciones
            return Conversacion.objects.prefetch_related('participantes', 'mensajes')

        return Conversacion.objects.filter(
            participantes=empleado
        ).prefetch_related('participantes', 'mensajes')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        conversacion = self.object

        # Marcar mensajes como leídos para el usuario actual
        try:
            empleado = self.request.user.empleado
            # Marcar como leído todos los mensajes que no son del usuario actual
            conversacion.mensajes.filter(leido=False).exclude(remitente=empleado).update(leido=True)
        except (AttributeError, Empleado.DoesNotExist):
            # Los admins también pueden marcar como leídos
            if self.request.user.is_staff or self.request.user.is_superuser:
                conversacion.mensajes.filter(leido=False).update(leido=True)

        context['mensajes'] = conversacion.mensajes.select_related('remitente').order_by('fecha_creacion')
        context['form'] = MensajeForm()
        return context


class IniciarConversacionView(EmpleadoRequiredMixin, CreateView):
    """Iniciar nueva conversación"""
    model = Conversacion
    form_class = ConversacionForm
    template_name = 'employees/messaging/iniciar_conversacion.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        try:
            kwargs['usuario_actual'] = self.request.user.empleado
        except (AttributeError, Empleado.DoesNotExist):
            # Los admins sin empleado no pueden iniciar conversaciones
            if not (self.request.user.is_staff or self.request.user.is_superuser):
                kwargs['usuario_actual'] = None
        return kwargs

    def form_valid(self, form):
        """Crear conversación con participantes"""
        form.instance.creado_por = self.request.user
        response = super().form_valid(form)

        # Agregar participantes
        try:
            empleado_actual = self.request.user.empleado
            participante = form.cleaned_data['participante_id']
            self.object.participantes.add(empleado_actual, participante)
        except (AttributeError, Empleado.DoesNotExist):
            # Los admins sin empleado no pueden agregar participantes
            pass

        return response

    def get_success_url(self):
        return reverse('employees:conversacion_detail', kwargs={'pk': self.object.pk})


class EnviarMensajeView(EmpleadoRequiredMixin, CreateView):
    """Enviar mensaje en conversación"""
    model = Mensaje
    form_class = MensajeForm
    template_name = 'employees/messaging/enviar_mensaje.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        try:
            kwargs['remitente'] = self.request.user.empleado
        except (AttributeError, Empleado.DoesNotExist):
            # Los admins sin empleado no pueden enviar mensajes
            kwargs['remitente'] = None
        return kwargs

    def get_conversacion(self):
        """Obtener conversación"""
        if not hasattr(self, '_conversacion'):
            try:
                empleado = self.request.user.empleado
                self._conversacion = get_object_or_404(
                    Conversacion,
                    pk=self.kwargs['conversacion_pk'],
                    participantes=empleado
                )
            except (AttributeError, Empleado.DoesNotExist):
                # Los admins ven todas las conversaciones
                if self.request.user.is_staff or self.request.user.is_superuser:
                    self._conversacion = get_object_or_404(
                        Conversacion,
                        pk=self.kwargs['conversacion_pk']
                    )
                else:
                    raise Http404("No tienes acceso a esta conversación")
        return self._conversacion

    def form_valid(self, form):
        """Guardar mensaje"""
        form.instance.conversacion = self.get_conversacion()
        try:
            form.instance.remitente = self.request.user.empleado
        except (AttributeError, Empleado.DoesNotExist):
            # Solo empleados o usuarios con empleado pueden enviar mensajes
            if not (self.request.user.is_staff or self.request.user.is_superuser):
                messages.error(
                    self.request,
                    'No tienes permiso para enviar mensajes. Debes tener un perfil de empleado.'
                )
            # No continuar si no se pudo asignar el remitente
            return super().form_valid(form)

        form.instance.creado_por = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('employees:conversacion_detail', kwargs={'pk': self.kwargs['conversacion_pk']})


@login_required
def marcar_mensajes_leidos(request, conversacion_pk):
    """Vista AJAX para marcar mensajes de una conversación como leídos"""
    if request.method == 'POST':
        try:
            empleado = request.user.empleado
            conversacion = get_object_or_404(
                Conversacion,
                pk=conversacion_pk,
                participantes=empleado
            )
            # Marcar mensajes como leídos
            conversacion.mensajes.filter(leido=False).exclude(remitente=empleado).update(leido=True)
            return JsonResponse({'success': True, 'mensaje': 'Mensajes marcados como leídos'})
        except (AttributeError, Empleado.DoesNotExist):
            # Los admins también pueden marcar como leídos
            if request.user.is_staff or request.user.is_superuser:
                conversacion = get_object_or_404(Conversacion, pk=conversacion_pk)
                conversacion.mensajes.filter(leido=False).update(leido=True)
                return JsonResponse({'success': True, 'mensaje': 'Mensajes marcados como leídos'})
            return JsonResponse({'success': False, 'error': 'No tienes acceso'}, status=403)
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)


@login_required
def contactar_empleado(request, empleado_id):
    """
    Iniciar o abrir conversación existente con un empleado específico.
    Si ya existe una conversación entre los dos, redirige a ella.
    Si no existe, crea una nueva conversación.
    """
    try:
        empleado_actual = request.user.empleado
        empleado_destino = get_object_or_404(Empleado, id=empleado_id)

        # No puede iniciar conversación consigo mismo
        if empleado_actual == empleado_destino:
            messages.warning(request, 'No puedes iniciar una conversación contigo mismo.')
            return redirect('employees:inbox')

        # Buscar conversación existente entre estos dos empleados
        # Una conversación entre dos personas específicas
        conversaciones_comunes = Conversacion.objects.filter(
            participantes=empleado_actual
        ).filter(
            participantes=empleado_destino
        ).annotate(
            num_participantes=models.Count('participantes')
        ).filter(
            num_participantes=2  # Solo conversaciones 1 a 1
        ).order_by('-fecha_actualizacion')

        if conversaciones_comunes.exists():
            # Ya existe una conversación, redirigir a ella
            conversacion = conversaciones_comunes.first()
            return redirect('employees:conversacion_detail', pk=conversacion.pk)

        # No existe conversación, crear una nueva
        conversacion = Conversacion.objects.create(
            titulo=f"Chat con {empleado_destino.nombre_completo}",
            contexto='general',
            creado_por=request.user
        )
        conversacion.participantes.add(empleado_actual, empleado_destino)

        messages.success(request, f'Conversación iniciada con {empleado_destino.nombre_completo}')
        return redirect('employees:conversacion_detail', pk=conversacion.pk)

    except Empleado.DoesNotExist:
        messages.error(request, 'No tienes un perfil de empleado asociado.')
        return redirect('employees:inbox')


@login_required
@require_POST
def enviar_mensaje_producto(request, pk):
    """Enviar mensaje al vendedor desde la página del producto"""
    try:
        # Obtener el producto y validar que existe
        producto = get_object_or_404(Producto, pk=pk)

        # Validar que el usuario sea empleado (no puede ser superusuario sin perfil)
        try:
            remitente = request.user.empleado
        except Empleado.DoesNotExist:
            messages.error(request, 'Debes tener un perfil de empleado para enviar mensajes.')
            return redirect('employees:producto_detail', pk=pk)

        # No puede enviarse mensaje a sí mismo
        if remitente.id == producto.vendedor.id:
            messages.error(request, 'No puedes enviarte mensajes a ti mismo.')
            return redirect('employees:producto_detail', pk=pk)

        # Obtener el contenido del mensaje
        contenido = request.POST.get('contenido', '').strip()

        if not contenido:
            messages.error(request, 'El mensaje no puede estar vacío.')
            return redirect('employees:producto_detail', pk=pk)

        # Buscar conversación existente entre los dos empleados para este producto
        conversaciones = Conversacion.objects.filter(
            participantes=remitente,
            producto_referencia=producto
        ).filter(
            participantes=producto.vendedor
        )

        if conversaciones.exists():
            conversacion = conversaciones.first()
        else:
            # Crear nueva conversación
            conversacion = Conversacion.objects.create(
                titulo=f"Pregunta sobre {producto.titulo}",
                contexto='venta',
                producto_referencia=producto,
                creado_por=request.user
            )
            # Agregar participantes
            conversacion.participantes.add(remitente, producto.vendedor)

        # Crear el mensaje
        mensaje = Mensaje.objects.create(
            conversacion=conversacion,
            remitente=remitente,
            contenido=contenido,
            creado_por=request.user
        )

        messages.success(request, 'Mensaje enviado al vendedor exitosamente.')
        return redirect('employees:producto_detail', pk=pk)

    except Exception as e:
        logger.error(f'Error enviando mensaje del producto: {str(e)}')
        messages.error(request, 'Hubo un error al enviar el mensaje. Intenta de nuevo.')
        return redirect('employees:producto_detail', pk=pk)


@login_required
def separar_producto_view(request, pk):
    """Vista que muestra el formulario de advertencia para separar un producto"""
    from apps.employees.models import Producto, Reserva

    producto = get_object_or_404(Producto, pk=pk)

    # Verificar que el usuario sea empleado
    try:
        comprador = request.user.empleado
    except Empleado.DoesNotExist:
        messages.error(request, 'Debes tener un perfil de empleado para separar productos.')
        return redirect('employees:producto_detail', pk=pk)

    # Verificar que no sea el vendedor
    if comprador.id == producto.vendedor.id:
        messages.error(request, 'No puedes separar tu propio producto.')
        return redirect('employees:producto_detail', pk=pk)

    # Verificar disponibilidad mínima de 1 unidad
    if not producto.tiene_disponible(cantidad_requerida=1):
        messages.error(request, 'Este producto no tiene cantidad disponible para separar.')
        return redirect('employees:producto_detail', pk=pk)

    # Verificar si ya tiene una reserva activa
    reserva_existente = Reserva.objects.filter(
        producto=producto,
        comprador=comprador,
        estado='activa'
    ).first()

    # Calcular disponible para mostrar en el formulario
    tiene_limite = producto.cantidad_disponible is not None
    if tiene_limite:
        cantidad_disponible_actual = producto.cantidad_disponible - producto.get_cantidad_reservada()
    else:
        cantidad_disponible_actual = None

    # GET: mostrar formulario de advertencia
    if request.method == 'GET':
        context = {
            'producto': producto,
            'cantidad_reservada': producto.get_cantidad_reservada(),
            'disponible': producto.get_disponible_texto(),
            'tiene_limite': tiene_limite,
            'cantidad_disponible_actual': cantidad_disponible_actual,
            'reserva_existente': reserva_existente,
        }
        return render(request, 'employees/marketplace/separar_producto.html', context)

    # POST: crear la reserva
    elif request.method == 'POST':
        confirmado = request.POST.get('confirmado') == 'true'
        acepto_terminos = request.POST.get('acepto_terminos') == 'true'

        # Leer y validar cantidad solicitada
        try:
            cantidad_solicitada = int(request.POST.get('cantidad_solicitada', 1))
            if cantidad_solicitada < 1:
                cantidad_solicitada = 1
        except (ValueError, TypeError):
            cantidad_solicitada = 1

        def _render_form(msg_type=None, msg=None):
            if msg_type == 'warning':
                messages.warning(request, msg)
            elif msg_type == 'error':
                messages.error(request, msg)
            context = {
                'producto': producto,
                'cantidad_reservada': producto.get_cantidad_reservada(),
                'disponible': producto.get_disponible_texto(),
                'tiene_limite': tiene_limite,
                'cantidad_disponible_actual': cantidad_disponible_actual,
                'cantidad_solicitada': cantidad_solicitada,
                'reserva_existente': reserva_existente,
            }
            return render(request, 'employees/marketplace/separar_producto.html', context)

        if not confirmado:
            return _render_form('warning', 'Debes confirmar la separación del producto.')

        if not acepto_terminos:
            return _render_form('error', 'Debes aceptar los términos y condiciones para continuar.')

        # Si existe una reserva, validar disponibilidad solo para las unidades ADICIONALES
        if reserva_existente:
            # Validar disponibilidad para la nueva cantidad adicional
            # (sin contar las unidades ya reservadas por este usuario)
            cantidad_ya_reservada_usuario = reserva_existente.cantidad_solicitada
            if tiene_limite:
                # Total reservado por OTROS usuarios
                otras_reservas = producto.get_cantidad_reservada() - cantidad_ya_reservada_usuario
                disponible_para_agregar = producto.cantidad_disponible - otras_reservas

                if cantidad_solicitada > disponible_para_agregar:
                    return _render_form('error', f'No puedes agregar {cantidad_solicitada} unidades. Solo hay {disponible_para_agregar} disponibles.')
        else:
            # Nueva reserva: validar disponibilidad total
            if not producto.tiene_disponible(cantidad_requerida=cantidad_solicitada):
                unidades_disp = cantidad_disponible_actual if tiene_limite else '∞'
                return _render_form('error', f'No hay suficiente disponibilidad. Solo quedan {unidades_disp} unidades.')

        try:
            if reserva_existente:
                # Actualizar la reserva existente incrementando la cantidad
                nueva_cantidad = reserva_existente.cantidad_solicitada + cantidad_solicitada
                reserva_existente.cantidad_solicitada = nueva_cantidad
                reserva_existente.save()

                unidades_txt = f'{cantidad_solicitada} unidad{"es" if cantidad_solicitada > 1 else ""}'
                messages.success(
                    request,
                    f'Has agregado {unidades_txt} a tu separación. Total ahora: {nueva_cantidad} unidad{"es" if nueva_cantidad > 1 else ""} de "{producto.titulo}".'
                )
            else:
                # Crear nueva reserva con la cantidad solicitada
                reserva = Reserva.objects.create(
                    producto=producto,
                    comprador=comprador,
                    creado_por=request.user,
                    estado='activa',
                    cantidad_solicitada=cantidad_solicitada,
                )

                unidades_txt = f'{cantidad_solicitada} unidad{"es" if cantidad_solicitada > 1 else ""}'
                messages.success(
                    request,
                    f'Separación exitosa: {unidades_txt} de "{producto.titulo}". '
                    f'Tienes 7 días para confirmar la compra o la reserva se cancelará.'
                )

            return redirect('employees:producto_detail', pk=pk)

        except IntegrityError:
            messages.error(request, 'Ya tienes una separación activa de este producto.')
            return redirect('employees:producto_detail', pk=pk)
        except Exception as e:
            logger.error(f'Error procesando reserva: {str(e)}')
            messages.error(request, 'Hubo un error al procesar tu solicitud. Intenta de nuevo.')
            return redirect('employees:producto_detail', pk=pk)


@login_required
@require_POST
def confirmar_entrega_reserva(request, reserva_id):
    """Vista para que el vendedor confirme la entrega y finalice la venta"""
    from apps.employees.models import Reserva, Venta

    # Obtener la reserva
    reserva = get_object_or_404(Reserva, pk=reserva_id)

    # Verificar que el usuario sea empleado
    try:
        vendedor = request.user.empleado
    except Empleado.DoesNotExist:
        messages.error(request, 'Debes tener un perfil de empleado.')
        return redirect('employees:mis_productos')

    # Verificar que el usuario sea el vendedor del producto
    if reserva.producto.vendedor.id != vendedor.id:
        messages.error(request, 'No tienes permiso para confirmar esta reserva.')
        return redirect('employees:mis_productos')

    # Verificar que la reserva esté activa
    if reserva.estado != 'activa':
        messages.warning(request, f'Esta reserva ya está en estado: {reserva.get_estado_display()}')
        return redirect('employees:mis_productos')

    # Verificar que no tenga ya una venta asociada
    if reserva.venta_asociada:
        messages.warning(request, 'Esta reserva ya tiene una venta asociada.')
        return redirect('employees:mis_productos')

    try:
        with transaction.atomic():
            # Crear la venta completada
            venta = Venta.objects.create(
                producto=reserva.producto,
                vendedor=reserva.producto.vendedor,
                comprador=reserva.comprador,
                precio_final=reserva.producto.precio_inicial or 0,
                estado='completada',
                observaciones=f'Venta confirmada desde reserva {reserva.id}',
                fecha_completada=timezone.now(),
                creado_por=request.user
            )

            # Actualizar la reserva
            reserva.estado = 'confirmada'
            reserva.venta_asociada = venta
            reserva.save()

            # Actualizar cantidad del producto si aplica (descontar unidades solicitadas)
            if reserva.producto.cantidad_disponible is not None:
                unidades = reserva.cantidad_solicitada or 1
                nuevo_stock = reserva.producto.cantidad_disponible - unidades
                reserva.producto.cantidad_disponible = max(nuevo_stock, 0)
                reserva.producto.save()

            # Log de éxito
            logger.info(
                f'[CONFIRMAR_ENTREGA] Reserva {reserva.id} confirmada. '
                f'Venta {venta.id} creada. Producto: {reserva.producto.titulo}'
            )

            messages.success(
                request,
                f'¡Venta confirmada exitosamente! Se ha completado la transacción con {reserva.comprador.nombre_completo}.'
            )

    except Exception as e:
        logger.error(f'Error confirmando entrega de reserva {reserva_id}: {str(e)}')
        messages.error(request, 'Hubo un error al confirmar la entrega. Intenta de nuevo.')

    return redirect('employees:mis_productos')


@login_required
@require_http_methods(["POST"])
def calificar_venta(request, venta_id):
    """Vista AJAX para que el comprador califique una venta completada"""
    from apps.employees.models import Venta

    # Obtener la venta
    venta = get_object_or_404(Venta, pk=venta_id)

    # Verificar que el usuario sea empleado
    try:
        comprador = request.user.empleado
    except Empleado.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Debes tener un perfil de empleado.'}, status=403)

    # Verificar que el usuario sea el comprador
    if venta.comprador.id != comprador.id:
        return JsonResponse({'success': False, 'message': 'No tienes permiso para calificar esta venta.'}, status=403)

    # Verificar que la venta esté completada
    if venta.estado != 'completada':
        return JsonResponse({'success': False, 'message': 'Solo puedes calificar ventas completadas.'}, status=400)

    # Verificar que no haya sido calificada ya
    if venta.calificacion_comprador:
        return JsonResponse({'success': False, 'message': 'Ya has calificado esta venta.'}, status=400)

    # Obtener datos del formulario
    try:
        calificacion = int(request.POST.get('calificacion'))
        comentario = request.POST.get('comentario', '').strip()

        # Validar calificación
        if calificacion < 1 or calificacion > 5:
            return JsonResponse({'success': False, 'message': 'La calificación debe estar entre 1 y 5.'}, status=400)

        # Guardar calificación
        venta.calificacion_comprador = calificacion
        venta.comentario_comprador = comentario
        venta.save()

        logger.info(f'[CALIFICAR_VENTA] Venta {venta.id} calificada con {calificacion} estrellas por {comprador.nombre_completo}')

        return JsonResponse({
            'success': True,
            'message': '¡Gracias por tu calificación!',
            'calificacion': calificacion,
            'comentario': comentario
        })

    except (ValueError, TypeError) as e:
        logger.error(f'Error procesando calificación de venta {venta_id}: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Datos inválidos.'}, status=400)
    except Exception as e:
        logger.error(f'Error guardando calificación de venta {venta_id}: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Error al guardar la calificación.'}, status=500)


# =============================================================================
# MARKETPLACE - TÉRMINOS Y CONDICIONES
# =============================================================================

class TerminosCondicionesView(TemplateView):
    """Vista para mostrar términos y condiciones del marketplace"""
    template_name = 'employees/marketplace/terminos_condiciones.html'


# =============================================================================
# HISTORIAL FAMILIAR — Autogestión del empleado
# =============================================================================

def _empleado_actual(request):
    """Devuelve el Empleado vinculado al usuario logueado, o 404."""
    return get_object_or_404(Empleado, usuario=request.user)


@login_required
def familia_panel(request):
    """Panel principal de Historial Familiar (pestaña en el perfil)."""
    empleado = _empleado_actual(request)
    familiares = empleado.familiares.all().prefetch_related('documentos').order_by('tipo', 'fecha_nacimiento')

    grupos = {'pareja': [], 'hijo': [], 'padre': [], 'hermano': [], 'otro': []}
    for f in familiares:
        grupos.setdefault(f.tipo, []).append(f)

    return render(request, 'employees/familia/panel.html', {
        'empleado': empleado,
        'estado_civil_form': EstadoCivilForm(instance=empleado),
        'familiar_form': FamiliarForm(),
        'grupos': grupos,
        'tipos_documento': TipoDocumento.objects.all().order_by('codigo'),
        'es_padre': empleado.es_padre,
    })


@login_required
@require_POST
def familia_estado_civil_actualizar(request):
    empleado = _empleado_actual(request)
    form = EstadoCivilForm(request.POST, instance=empleado)
    if form.is_valid():
        form.save()
        messages.success(request, 'Estado civil actualizado.')
    else:
        messages.error(request, 'No se pudo actualizar el estado civil.')
    return redirect('employees:familia_panel')


@login_required
@require_POST
def familiar_crear(request):
    empleado = _empleado_actual(request)
    form = FamiliarForm(request.POST)
    if form.is_valid():
        try:
            with transaction.atomic():
                familiar = form.save(commit=False)
                familiar.empleado = empleado
                familiar.creado_por = request.user
                familiar.full_clean()
                familiar.save()
            messages.success(request, f'Familiar "{familiar.nombre_completo}" agregado.')
        except ValidationError as e:
            messages.error(request, '; '.join(e.messages))
    else:
        messages.error(request, 'Revisa los datos: ' + '; '.join(
            f"{k}: {v[0]}" for k, v in form.errors.items()
        ))
    return redirect('employees:familia_panel')


@login_required
@require_POST
def familiar_editar(request, familiar_id):
    empleado = _empleado_actual(request)
    familiar = get_object_or_404(Familiar, id=familiar_id, empleado=empleado)
    form = FamiliarForm(request.POST, instance=familiar)
    if form.is_valid():
        try:
            with transaction.atomic():
                familiar = form.save(commit=False)
                familiar.full_clean()
                familiar.save()
            messages.success(request, 'Datos actualizados.')
        except ValidationError as e:
            messages.error(request, '; '.join(e.messages))
    else:
        messages.error(request, 'Revisa los datos.')
    return redirect('employees:familia_panel')


@login_required
@require_POST
def familiar_eliminar(request, familiar_id):
    empleado = _empleado_actual(request)
    familiar = get_object_or_404(Familiar, id=familiar_id, empleado=empleado)
    nombre = familiar.nombre_completo
    familiar.delete()
    messages.success(request, f'Familiar "{nombre}" eliminado.')
    return redirect('employees:familia_panel')


@login_required
@require_POST
def documento_familiar_subir(request, familiar_id):
    empleado = _empleado_actual(request)
    familiar = get_object_or_404(Familiar, id=familiar_id, empleado=empleado)
    form = DocumentoFamiliarForm(request.POST, request.FILES)
    if form.is_valid():
        doc = form.save(commit=False)
        doc.familiar = familiar
        doc.save()
        messages.success(request, f'Documento subido para {familiar.nombre_completo}.')
    else:
        messages.error(request, 'No se pudo subir el documento. Verifica el tipo y el archivo.')
    return redirect('employees:familia_panel')


@login_required
@require_POST
def documento_familiar_eliminar(request, documento_id):
    empleado = _empleado_actual(request)
    doc = get_object_or_404(DocumentoFamiliar, id=documento_id, familiar__empleado=empleado)
    doc.archivo.delete(save=False)
    doc.delete()
    messages.success(request, 'Documento eliminado.')
    return redirect('employees:familia_panel')


# =============================================================================
# HISTORIAL FAMILIAR — Vista cruzada (staff/RRHH)
# =============================================================================

def _familiares_filtrar(request):
    """Construye el queryset filtrado por GET params. Reutilizado por list y export."""
    qs = Familiar.objects.select_related(
        'empleado', 'empleado__sede', 'empleado__estado', 'tipo_documento'
    ).prefetch_related('documentos').all()

    q = request.GET.get('q', '').strip()
    tipo = request.GET.get('tipo', '').strip()
    sede_id = request.GET.get('sede', '').strip()
    estado_civil = request.GET.get('estado_civil', '').strip()
    convive = request.GET.get('convive', '').strip()
    dependiente = request.GET.get('dependiente', '').strip()
    eps = request.GET.get('eps', '').strip()
    edad_min = request.GET.get('edad_min', '').strip()
    edad_max = request.GET.get('edad_max', '').strip()
    activo = request.GET.get('activo', '').strip()
    padres_madres = request.GET.get('padres_madres', '').strip()  # '' | 'padres' | 'madres' | 'todos'
    sexo_empleado = request.GET.get('sexo_empleado', '').strip()

    if q:
        qs = qs.filter(
            Q(nombres__icontains=q) | Q(apellidos__icontains=q) |
            Q(numero_documento__icontains=q) |
            Q(empleado__nombres__icontains=q) | Q(empleado__apellidos__icontains=q) |
            Q(empleado__numero_documento__icontains=q)
        )
    if tipo:
        qs = qs.filter(tipo=tipo)
    if sede_id:
        qs = qs.filter(empleado__sede_id=sede_id)
    if estado_civil:
        qs = qs.filter(empleado__estado_civil=estado_civil)
    if convive == 'si':
        qs = qs.filter(convive=True)
    elif convive == 'no':
        qs = qs.filter(convive=False)
    if dependiente == 'si':
        qs = qs.filter(dependiente_economico=True)
    elif dependiente == 'no':
        qs = qs.filter(dependiente_economico=False)
    if eps:
        qs = qs.filter(eps__icontains=eps)
    if activo == 'si':
        qs = qs.filter(activo=True)
    elif activo == 'no':
        qs = qs.filter(activo=False)
    if sexo_empleado in ('M', 'F'):
        qs = qs.filter(empleado__sexo_biologico=sexo_empleado)
    if padres_madres in ('padres', 'madres', 'todos'):
        # Restringe a empleados con al menos un hijo registrado. La dedupe
        # universal de _dedupe_por_empleado() se encarga de mostrar una fila
        # por empleado.
        qs = qs.filter(tipo='hijo')
        if padres_madres == 'padres':
            qs = qs.filter(empleado__sexo_biologico='M')
        elif padres_madres == 'madres':
            qs = qs.filter(empleado__sexo_biologico='F')

    # Filtro por edad — requiere fecha_nacimiento y se calcula contra hoy
    hoy = date.today()
    if edad_min.isdigit():
        fecha_tope = date(hoy.year - int(edad_min), hoy.month, hoy.day) if hoy.month != 2 or hoy.day != 29 else date(hoy.year - int(edad_min), 2, 28)
        qs = qs.filter(fecha_nacimiento__lte=fecha_tope)
    if edad_max.isdigit():
        # nació después de hoy - (edad_max + 1) años
        edad_max_int = int(edad_max)
        try:
            fecha_inicio = date(hoy.year - edad_max_int - 1, hoy.month, hoy.day) + timedelta(days=1)
        except ValueError:
            fecha_inicio = date(hoy.year - edad_max_int - 1, 2, 28) + timedelta(days=1)
        qs = qs.filter(fecha_nacimiento__gte=fecha_inicio)

    return qs.order_by('empleado__apellidos', 'empleado__nombres', 'tipo', 'fecha_nacimiento')


def _dedupe_por_empleado(qs):
    """Reduce el queryset a una fila por empleado.

    Cada empleado queda representado por su primer familiar coincidente con
    los filtros (orden interno: tipo, fecha_creacion). Pensado para listados
    de segmentación donde duplicar al empleado por cada hijo es ruido.
    Requiere PostgreSQL por el .distinct('field').
    """
    ids = list(
        qs.order_by('empleado_id', 'tipo', 'fecha_creacion')
          .distinct('empleado_id')
          .values_list('id', flat=True)
    )
    return Familiar.objects.select_related(
        'empleado', 'empleado__sede', 'empleado__estado', 'tipo_documento'
    ).prefetch_related('documentos').filter(id__in=ids).order_by(
        'empleado__apellidos', 'empleado__nombres'
    )


from django.contrib.admin.views.decorators import staff_member_required


@staff_member_required
def panel_admin(request):
    """Punto de entrada de RRHH/Admin. Accesible a is_staff (incluye superuser)
    sin requerir Empleado vinculado al usuario.
    """
    resumen = {
        'empleados_activos': Empleado.objects.filter(estado__codigo='999').count(),
        'empleados_en_prueba': Empleado.objects.filter(estado__codigo='p-prue').count(),
        'familiares': Familiar.objects.count(),
        'empleados_con_hijos': Empleado.objects.filter(familiares__tipo='hijo').distinct().count(),
        'vacaciones_pendientes': SolicitudVacacion.objects.filter(estado_local='enviada_pendiente_rrhh').count(),
        'vacaciones_total_mes': SolicitudVacacion.objects.filter(
            fecha_creacion__gte=timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        ).count(),
    }
    return render(request, 'employees/admin/panel.html', {'resumen': resumen})


@staff_member_required
def familiares_admin_lista(request):
    qs = _familiares_filtrar(request)

    # Resumen: cuentas sobre el set filtrado (sin dedup) para tener los
    # totales reales por tipo. 'empleados' es el conteo único de empleados
    # con al menos un familiar coincidente — coincide con las filas que
    # mostrará la tabla deduplicada.
    resumen = {
        'empleados': qs.values('empleado_id').distinct().count(),
        'familiares': qs.count(),
        'parejas': qs.filter(tipo='pareja').count(),
        'hijos': qs.filter(tipo='hijo').count(),
        'padres': qs.filter(tipo='padre').count(),
        'dependientes': qs.filter(dependiente_economico=True).count(),
        'empleados_padres': Empleado.objects.filter(familiares__tipo='hijo').distinct().count(),
    }

    # Tabla: dedupe solo si NO hay un tipo de familiar explícito.
    # Si el usuario eligió "Tipo = Hijo/a" (o Pareja, Padre, etc.) desde el
    # dropdown, quiere ver cada familiar individualmente — ej. listado de
    # hijos para kit escolar, donde el padre puede repetirse. Sin tipo, la
    # tabla muestra una fila por empleado (segmentación).
    tipo_explicito = bool(request.GET.get('tipo', '').strip())
    qs_tabla = qs if tipo_explicito else _dedupe_por_empleado(qs)
    paginator = Paginator(qs_tabla, 50)
    page_obj = paginator.get_page(request.GET.get('page'))

    from apps.organizational.models import Sede as _Sede
    return render(request, 'employees/familia/admin_lista.html', {
        'page_obj': page_obj,
        'resumen': resumen,
        'sedes': _Sede.objects.filter(activa=True).order_by('nombre'),
        'estado_civil_choices': Empleado.ESTADO_CIVIL_CHOICES,
        'sexo_choices': Empleado.SEXO_BIOLOGICO_CHOICES,
        'tipo_choices': Familiar.TIPO_CHOICES,
        'filtros': request.GET,
    })


@staff_member_required
def familiares_admin_export_excel(request):
    # Consistente con la tabla: si se eligió un tipo explícito (ej. Hijo/a),
    # exporta cada familiar individualmente. Sin tipo, una fila por empleado.
    qs_filtrado = _familiares_filtrar(request)
    tipo_explicito = bool(request.GET.get('tipo', '').strip())
    qs = qs_filtrado if tipo_explicito else _dedupe_por_empleado(qs_filtrado)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl no disponible.', status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Familiares'

    headers = [
        'Empleado', 'Doc. Empleado', 'Sede', 'Sexo empleado', 'Estado Civil', 'Tipo Familiar',
        'Nombres', 'Apellidos', 'Doc. Familiar', 'Núm. Doc.',
        'Fecha Nacimiento', 'Edad', 'EPS', 'Convive', 'Dependiente',
        'Parentesco', 'Activo', '# Documentos',
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        c.alignment = Alignment(horizontal='center')

    for r, f in enumerate(qs, 2):
        ws.cell(row=r, column=1, value=f.empleado.nombre_completo)
        ws.cell(row=r, column=2, value=f.empleado.numero_documento)
        ws.cell(row=r, column=3, value=f.empleado.sede.nombre if f.empleado.sede_id else '')
        ws.cell(row=r, column=4, value=f.empleado.get_sexo_biologico_display() if f.empleado.sexo_biologico else '')
        ws.cell(row=r, column=5, value=f.empleado.get_estado_civil_display() if f.empleado.estado_civil else '')
        ws.cell(row=r, column=6, value=f.get_tipo_display())
        ws.cell(row=r, column=7, value=f.nombres)
        ws.cell(row=r, column=8, value=f.apellidos)
        ws.cell(row=r, column=9, value=f.tipo_documento.codigo if f.tipo_documento_id else '')
        ws.cell(row=r, column=10, value=f.numero_documento)
        ws.cell(row=r, column=11, value=f.fecha_nacimiento.strftime('%Y-%m-%d') if f.fecha_nacimiento else '')
        ws.cell(row=r, column=12, value=f.edad if f.edad is not None else '')
        ws.cell(row=r, column=13, value=f.eps)
        ws.cell(row=r, column=14, value='Sí' if f.convive else 'No')
        ws.cell(row=r, column=15, value='Sí' if f.dependiente_economico else 'No')
        ws.cell(row=r, column=16, value=f.parentesco)
        ws.cell(row=r, column=17, value='Sí' if f.activo else 'No')
        ws.cell(row=r, column=18, value=f.documentos.count())

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    stamp = timezone.now().strftime('%Y%m%d_%H%M')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=familiares_{stamp}.xlsx'
    wb.save(response)
    return response


# =============================================================================
# VACACIONES — Jefes solicitan/aprueban y se envía a Odoo
# =============================================================================

def _puede_solicitar_vacacion_para(usuario, empleado):
    """Reglas de quién puede solicitar la vacación de un empleado:
    - is_staff (RRHH/admin): para cualquiera.
    - Jefe directo (vía HistorialCargo.jefe_directo activo): para sus subordinados.
    """
    if usuario.is_staff:
        return True
    try:
        solicitante = Empleado.objects.get(usuario=usuario)
    except Empleado.DoesNotExist:
        return False
    return HistorialCargo.objects.filter(
        empleado=empleado, activo=True, jefe_directo=solicitante
    ).exists()


def _equipo_del_jefe(usuario):
    """Subordinados directos del usuario (activos o en periodo de prueba).

    Para staff/RRHH NO devuelve "todos los activos": eso vive en la vista admin
    de vacaciones. Aquí siempre son los subordinados directos del jefe; si el
    usuario es staff y además jefe, ve solo a su propio equipo.
    """
    try:
        solicitante = Empleado.objects.get(usuario=usuario)
    except Empleado.DoesNotExist:
        return Empleado.objects.none()
    return Empleado.objects.filter(
        historialcargo__activo=True,
        historialcargo__jefe_directo=solicitante,
        estado__codigo__in=['999', 'p-prue'],
    ).distinct().order_by('apellidos', 'nombres')


@login_required
def mis_vacaciones(request):
    """Historial de solicitudes de vacaciones del empleado autenticado.

    El saldo de días es al ÚLTIMO CORTE MENSUAL (fin de mes anterior) según Odoo.
    No cambia durante el mes. Por eso el cache es mensual: SIGHU solo consulta a
    Odoo si el `fecha_corte` que tiene guardado NO es del mes anterior al actual
    (es decir, si todavía no hemos traído el corte de este mes).
    """
    from datetime import date
    from apps.integraciones.odoo.services import obtener_saldo_vacaciones_odoo

    try:
        empleado = Empleado.objects.get(usuario=request.user)
    except Empleado.DoesNotExist:
        messages.info(request, 'Tu usuario no está vinculado a un empleado.')
        return redirect('core:dashboard')

    # ¿Ya tenemos el saldo del corte del mes anterior al actual?
    # Si sí (fecha_corte cae dentro del mes anterior) → no consultamos.
    hoy = date.today()
    if hoy.month == 1:
        year_mes_anterior, mes_anterior = hoy.year - 1, 12
    else:
        year_mes_anterior, mes_anterior = hoy.year, hoy.month - 1

    tiene_corte_actual = (
        empleado.saldo_vacaciones_fecha_corte is not None
        and empleado.saldo_vacaciones_fecha_corte.year == year_mes_anterior
        and empleado.saldo_vacaciones_fecha_corte.month == mes_anterior
    )

    consulta_saldo_falla = False
    if not tiene_corte_actual:
        resultado = obtener_saldo_vacaciones_odoo(empleado)
        if resultado.get('ok'):
            empleado.refresh_from_db(fields=[
                'saldo_vacaciones_dias',
                'saldo_vacaciones_fecha_corte',
                'saldo_vacaciones_actualizado',
            ])
        else:
            consulta_saldo_falla = True

    solicitudes = (
        SolicitudVacacion.objects
        .filter(empleado=empleado)
        .select_related('jefe_solicitante')
        .order_by('-fecha_creacion')
    )

    return render(request, 'employees/vacaciones/mis_vacaciones.html', {
        'empleado': empleado,
        'solicitudes': solicitudes,
        'saldo_dias': empleado.saldo_vacaciones_dias,
        'saldo_fecha_corte': empleado.saldo_vacaciones_fecha_corte,
        'saldo_actualizado': empleado.saldo_vacaciones_actualizado,
        'consulta_saldo_falla': consulta_saldo_falla,
    })


# ============================================================================
# Asistencia diaria — tablero del jefe
# ============================================================================
# Jornada estándar Construinmuniza: L-V, 7:00 AM a 4:24 PM con 1h de descanso
# (20+40 min). Los sábados/domingos/festivos NO se registran.
from datetime import date as _date_const  # aliased para constantes globales
# Fecha desde la cual el módulo de asistencia está oficialmente activo.
# Antes de esta fecha NO se consideran "días sin registrar" — no era falta
# del jefe, simplemente el sistema aún no operaba. Ajustar si se cambia la
# fecha de arranque.
ASISTENCIA_FECHA_ARRANQUE = _date_const(2026, 7, 27)  # lunes 27/07/2026

JORNADA_HORA_INGRESO = '07:00'
JORNADA_HORA_SALIDA = '16:24'
JORNADA_DESCANSOS_MIN = '20 + 40'
DIAS_RETROACTIVOS_ASISTENCIA = 7


def _tiene_vacacion_aprobada_en(empleado, fecha):
    """True si el empleado tiene una SolicitudVacacion tipo=tiempo aprobada
    que cubra la fecha dada (fecha_inicio <= fecha <= fecha_fin).
    """
    return SolicitudVacacion.objects.filter(
        empleado=empleado,
        tipo='tiempo',
        estado_local='aprobada_rrhh',
        fecha_inicio__lte=fecha,
        fecha_fin__gte=fecha,
    ).exists()


def _fecha_desde_query(request, default=None):
    """Parsea ?fecha=YYYY-MM-DD del query string. Default: hoy."""
    from datetime import date, datetime
    raw = (request.GET.get('fecha') or '').strip()
    if not raw:
        return default or date.today()
    try:
        return datetime.strptime(raw, '%Y-%m-%d').date()
    except ValueError:
        return default or date.today()


@login_required
def asistencia_diaria(request):
    """Tablero del jefe para registrar la asistencia diaria del equipo.

    GET: renderiza el equipo con el estado prellenado (presente por defecto,
         vacaciones aprobadas → readonly).
    POST: procesa el formulario, actualiza los registros del día. Idempotente:
         puede corregirse dentro de la ventana retroactiva (7 días).
    """
    from datetime import date, timedelta
    from django.db import transaction

    fecha = _fecha_desde_query(request)
    hoy = date.today()
    ventana_min = hoy - timedelta(days=DIAS_RETROACTIVOS_ASISTENCIA)

    # Validaciones de fecha
    if fecha > hoy:
        messages.warning(request, 'No puedes registrar asistencia de fechas futuras.')
        return redirect('employees:asistencia_diaria')
    if fecha < ventana_min:
        messages.warning(
            request,
            f'Solo puedes registrar hasta {DIAS_RETROACTIVOS_ASISTENCIA} días atrás. '
            f'Para fechas anteriores solicítalo a RRHH.'
        )
        return redirect('employees:asistencia_diaria')
    es_fin_de_semana = fecha.weekday() >= 5

    equipo = list(_equipo_del_jefe(request.user))
    if not equipo:
        messages.info(request, 'No tienes empleados a cargo para registrar asistencia.')
        return redirect('employees:empleado_perfil')

    # Registros existentes de esa fecha para el equipo
    registros_existentes = {
        r.empleado_id: r
        for r in AsistenciaDiaria.objects.filter(
            empleado__in=equipo, fecha=fecha,
        ).select_related('empleado')
    }

    # Días L-V en la ventana retroactiva SIN ningún registro del equipo actual.
    # Se considera "sin registrar" si no hay ningún AsistenciaDiaria del equipo
    # para esa fecha (independiente de cuántos empleados haya). Excluye la
    # fecha visualizada actualmente para no duplicar el aviso.
    #
    # El corte inferior real es max(ventana_min, ASISTENCIA_FECHA_ARRANQUE):
    # antes de la fecha de arranque el sistema no operaba, así que esos días
    # NO son omisión del jefe y no se reclaman.
    limite_inferior = max(ventana_min, ASISTENCIA_FECHA_ARRANQUE)
    dias_sin_registrar = []
    fechas_con_registro = set(
        AsistenciaDiaria.objects
        .filter(empleado__in=equipo, fecha__gte=limite_inferior, fecha__lte=hoy)
        .values_list('fecha', flat=True)
        .distinct()
    )
    cursor = hoy
    while cursor >= limite_inferior:
        if cursor.weekday() < 5 and cursor != fecha and cursor not in fechas_con_registro:
            dias_sin_registrar.append(cursor)
        cursor -= timedelta(days=1)

    if request.method == 'POST':
        if es_fin_de_semana:
            messages.warning(request, 'No se puede registrar asistencia en sábado/domingo.')
            return redirect('employees:asistencia_diaria')

        try:
            registrado_por = Empleado.objects.get(usuario=request.user)
        except Empleado.DoesNotExist:
            registrado_por = None

        estados_validos = {c[0] for c in AsistenciaDiaria.ESTADO_CHOICES}
        empleado_ids_equipo = {e.id for e in equipo}
        creados, actualizados, errores = 0, 0, []

        with transaction.atomic():
            for emp in equipo:
                # Auto-detección: vacaciones aprobadas mandan sin importar el input
                if _tiene_vacacion_aprobada_en(emp, fecha):
                    estado_final = 'en_vacaciones'
                    motivo_final = 'Vacaciones aprobadas por RRHH.'
                else:
                    estado_final = (request.POST.get(f'estado_{emp.id}') or 'presente').strip()
                    motivo_final = (request.POST.get(f'motivo_{emp.id}') or '').strip()

                if emp.id not in empleado_ids_equipo:
                    continue
                if estado_final not in estados_validos:
                    errores.append(f'{emp.nombre_completo}: estado inválido "{estado_final}"')
                    continue
                if estado_final != 'presente' and estado_final not in AsistenciaDiaria.ESTADOS_AUTOMATICOS:
                    if not motivo_final:
                        errores.append(f'{emp.nombre_completo}: falta motivo para "{estado_final}"')
                        continue

                _, creado = AsistenciaDiaria.objects.update_or_create(
                    empleado=emp, fecha=fecha,
                    defaults={
                        'estado': estado_final,
                        'motivo': motivo_final,
                        'registrado_por': registrado_por,
                        'creado_por': request.user,
                    },
                )
                if creado:
                    creados += 1
                else:
                    actualizados += 1

        if errores:
            for err in errores:
                messages.error(request, err)
        if creados or actualizados:
            messages.success(
                request,
                f'Asistencia guardada: {creados} nuevos, {actualizados} actualizados.',
            )
        return redirect(f'{reverse("employees:asistencia_diaria")}?fecha={fecha.isoformat()}')

    # GET — armar filas con auto-detección
    filas = []
    for emp in equipo:
        reg = registros_existentes.get(emp.id)
        en_vacaciones = _tiene_vacacion_aprobada_en(emp, fecha)
        if en_vacaciones:
            estado, motivo, readonly = 'en_vacaciones', 'Vacaciones aprobadas por RRHH', True
        elif reg:
            estado, motivo, readonly = reg.estado, reg.motivo, False
        else:
            estado, motivo, readonly = 'presente', '', False
        filas.append({
            'empleado': emp,
            'estado': estado,
            'motivo': motivo,
            'readonly': readonly,
            'registrado_previamente': reg is not None,
        })

    fecha_anterior = (fecha - timedelta(days=1)) if fecha > ventana_min else None
    fecha_siguiente = (fecha + timedelta(days=1)) if fecha < hoy else None

    return render(request, 'employees/asistencia/diaria.html', {
        'fecha': fecha,
        'hoy': hoy,
        'es_fin_de_semana': es_fin_de_semana,
        'es_hoy': fecha == hoy,
        'filas': filas,
        'estados_editables': [
            c for c in AsistenciaDiaria.ESTADO_CHOICES
            if c[0] not in AsistenciaDiaria.ESTADOS_AUTOMATICOS
        ],
        'fecha_anterior': fecha_anterior,
        'fecha_siguiente': fecha_siguiente,
        'jornada_ingreso': JORNADA_HORA_INGRESO,
        'jornada_salida': JORNADA_HORA_SALIDA,
        'jornada_descansos': JORNADA_DESCANSOS_MIN,
        'dias_sin_registrar': dias_sin_registrar,
    })


@login_required
def asistencia_historial(request):
    """Historial de asistencia del equipo del jefe (filtro por rango)."""
    from datetime import date

    equipo = list(_equipo_del_jefe(request.user))
    if not equipo:
        messages.info(request, 'No tienes empleados a cargo.')
        return redirect('employees:empleado_perfil')

    hoy = date.today()
    fecha_desde = _fecha_desde_query(request, default=hoy.replace(day=1))
    fecha_hasta_raw = (request.GET.get('hasta') or '').strip()
    try:
        from datetime import datetime
        fecha_hasta = datetime.strptime(fecha_hasta_raw, '%Y-%m-%d').date() if fecha_hasta_raw else hoy
    except ValueError:
        fecha_hasta = hoy

    if fecha_desde > fecha_hasta:
        fecha_desde, fecha_hasta = fecha_hasta, fecha_desde

    registros = (
        AsistenciaDiaria.objects
        .filter(empleado__in=equipo, fecha__gte=fecha_desde, fecha__lte=fecha_hasta)
        .select_related('empleado', 'registrado_por')
        .order_by('-fecha', 'empleado__apellidos')
    )

    # Resumen por empleado
    from django.db.models import Count, Q
    resumen = (
        AsistenciaDiaria.objects
        .filter(empleado__in=equipo, fecha__gte=fecha_desde, fecha__lte=fecha_hasta)
        .values('empleado__id', 'empleado__nombres', 'empleado__apellidos')
        .annotate(
            total=Count('id'),
            presente=Count('id', filter=Q(estado='presente')),
            ausente=Count('id', filter=Q(estado='ausente')),
            retardo=Count('id', filter=Q(estado='retardo')),
            permiso=Count('id', filter=Q(estado='permiso')),
            permiso_no_remunerado=Count('id', filter=Q(estado='permiso_no_remunerado')),
            incapacidad=Count('id', filter=Q(estado='incapacidad')),
            en_vacaciones=Count('id', filter=Q(estado='en_vacaciones')),
        )
        .order_by('empleado__apellidos')
    )

    return render(request, 'employees/asistencia/historial.html', {
        'registros': registros,
        'resumen': resumen,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    })


@login_required
@require_POST
def descargar_carta_vacaciones(request, pk):
    """Genera y devuelve el PDF de la carta de vacaciones (solo POST con consentimiento).

    Requiere:
    - Empleado autenticado dueño de la solicitud.
    - Solicitud en estado aprobada_rrhh.
    - Checkbox 'consentimiento' marcado en el POST.

    Registra la primera descarga con un hash SHA-256 del consentimiento. Descargas
    posteriores permiten re-descargar el PDF pero no sobreescriben la constancia.
    """
    import hashlib
    from django.http import HttpResponse
    from django.utils import timezone
    from apps.employees.vacaciones_carta import generar_carta_vacaciones

    solicitud = get_object_or_404(SolicitudVacacion, pk=pk)

    # Verificar propiedad
    try:
        empleado = Empleado.objects.get(usuario=request.user)
    except Empleado.DoesNotExist:
        messages.error(request, 'Tu usuario no está vinculado a un empleado.')
        return redirect('core:dashboard')

    if solicitud.empleado_id != empleado.id:
        messages.error(request, 'No tienes permiso para descargar esta carta.')
        return redirect('employees:mis_vacaciones')

    if solicitud.estado_local != 'aprobada_rrhh':
        messages.warning(
            request,
            'Solo puedes descargar la carta cuando la solicitud está aprobada por RRHH.'
        )
        return redirect('employees:mis_vacaciones')

    if request.POST.get('consentimiento') != 'si':
        messages.warning(
            request,
            'Debes aceptar el consentimiento para descargar la carta.'
        )
        return redirect('employees:mis_vacaciones')

    # Registrar constancia (solo la primera vez)
    if not solicitud.carta_descargada_fecha:
        ahora = timezone.now()
        hash_input = f'{empleado.id}|{solicitud.id}|{ahora.isoformat()}'.encode()
        SolicitudVacacion.objects.filter(pk=solicitud.pk).update(
            carta_descargada_fecha=ahora,
            carta_confirmada_hash=hashlib.sha256(hash_input).hexdigest(),
        )
        solicitud.refresh_from_db()

    # Generar y devolver PDF
    pdf_bytes = generar_carta_vacaciones(solicitud)
    filename = (
        f'carta_vacaciones_{empleado.numero_documento}_'
        f'{solicitud.fecha_inicio.strftime("%Y%m%d") if solicitud.fecha_inicio else solicitud.pk}.pdf'
    )
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def vacaciones_equipo(request):
    """Panel del jefe: equipo + TODAS las solicitudes de vacaciones del equipo.

    Incluye las creadas por el mismo jefe, por otros roles y las importadas
    desde Odoo por RRHH (jefe_solicitante=NULL). Así el jefe tiene visibilidad
    completa para planear la operación.
    """
    equipo = list(_equipo_del_jefe(request.user))

    # Anotar en cada empleado su solicitud vigente (si existe) para que la UI
    # deshabilite el botón "Solicitar vacaciones" en lugar de dejar al jefe
    # avanzar al formulario y chocarse con el bloqueo.
    for emp in equipo:
        emp.solicitud_vigente = _solicitud_vigente_de(emp)

    solicitudes = SolicitudVacacion.objects.select_related(
        'empleado', 'jefe_solicitante',
    )
    solicitudes = solicitudes.filter(empleado__in=equipo)
    solicitudes = solicitudes.order_by('-fecha_creacion')[:100]

    return render(request, 'employees/vacaciones/equipo.html', {
        'equipo': equipo,
        'solicitudes': solicitudes,
    })


def _solicitud_vigente_de(empleado):
    """Retorna la SolicitudVacacion vigente del empleado o None.

    Se considera vigente si bloquea la creación de una nueva: tipo=tiempo, en
    un estado activo (borrador, pendiente Odoo o aprobada) y con fecha_fin
    todavía no pasada. Las compensaciones en dinero (tipo=pago_dinero) no
    bloquean porque son eventos puntuales de nómina, no ausencias.
    """
    from datetime import date
    return (
        SolicitudVacacion.objects
        .filter(
            empleado=empleado,
            tipo='tiempo',
            estado_local__in=['borrador', 'enviada_pendiente_rrhh', 'aprobada_rrhh'],
            fecha_fin__gte=date.today(),
        )
        .order_by('-fecha_creacion')
        .first()
    )


@login_required
def vacacion_nueva(request, empleado_id):
    """Form para crear y enviar una vacación de un empleado del equipo."""
    empleado = get_object_or_404(Empleado, id=empleado_id)
    if not _puede_solicitar_vacacion_para(request.user, empleado):
        messages.error(request, 'No tienes permiso para solicitar vacaciones de este empleado.')
        return redirect('employees:vacaciones_equipo')

    # Bloqueo: un empleado no puede tener dos solicitudes de tiempo vigentes.
    vigente = _solicitud_vigente_de(empleado)
    if vigente:
        estado_hum = vigente.get_estado_local_display()
        messages.warning(
            request,
            f'{empleado.nombre_completo} ya tiene una solicitud vigente del '
            f'{vigente.fecha_inicio.strftime("%d/%m/%Y")} al '
            f'{vigente.fecha_fin.strftime("%d/%m/%Y")} ({estado_hum}). '
            f'Podrás crear una nueva cuando termine ese periodo o si RRHH la '
            f'rechaza/cancela.'
        )
        return redirect('employees:vacaciones_equipo')

    if request.method == 'POST':
        form = SolicitudVacacionForm(request.POST)
        if form.is_valid():
            from apps.integraciones.odoo.services import enviar_vacacion_a_odoo
            try:
                jefe = Empleado.objects.get(usuario=request.user)
            except Empleado.DoesNotExist:
                jefe = None

            with transaction.atomic():
                solicitud = form.save(commit=False)
                solicitud.empleado = empleado
                solicitud.jefe_solicitante = jefe
                solicitud.creado_por = request.user
                solicitud.estado_local = 'borrador'
                solicitud.save()

            ok, data = enviar_vacacion_a_odoo(solicitud)
            solicitud.fecha_envio_odoo = timezone.now()
            solicitud.respuesta_odoo = data
            if ok:
                solicitud.estado_local = 'enviada_pendiente_rrhh'
                solicitud.leave_id_odoo = data.get('leave_id')
                solicitud.save()
                messages.success(
                    request,
                    f'Solicitud enviada a Odoo (leave_id={data.get("leave_id")}, '
                    f'{data.get("dias")} día(s)). Pendiente de aprobación de RRHH.'
                )
                return redirect('employees:vacaciones_equipo')
            else:
                motivo = data.get('motivo', 'Error desconocido')
                # Rechazo de negocio vs error técnico
                if 'No se pudo contactar' in motivo or motivo.startswith('Error de transporte') or motivo.startswith('HTTP'):
                    solicitud.estado_local = 'error_envio'
                else:
                    solicitud.estado_local = 'rechazada_odoo'
                solicitud.motivo_rechazo = motivo
                solicitud.save()
                messages.error(request, f'Rechazada/Error: {motivo}')
                return redirect('employees:vacaciones_equipo')
    else:
        form = SolicitudVacacionForm()

    return render(request, 'employees/vacaciones/nueva.html', {
        'empleado': empleado,
        'form': form,
    })


@staff_member_required
def vacaciones_admin_panel(request):
    """Panel RRHH: buscar empleado para solicitar + historial global con filtros."""
    # --- Búsqueda de empleado para iniciar una solicitud ---
    q = request.GET.get('q', '').strip()
    empleados_buscados = Empleado.objects.none()
    if q:
        empleados_buscados = Empleado.objects.filter(
            estado__codigo__in=['999', 'p-prue']
        ).filter(
            Q(nombres__icontains=q) | Q(apellidos__icontains=q) |
            Q(numero_documento__icontains=q)
        ).order_by('apellidos', 'nombres')[:20]

    # --- Historial global de solicitudes con filtros ---
    solicitudes = SolicitudVacacion.objects.select_related(
        'empleado', 'jefe_solicitante'
    ).order_by('-fecha_creacion')

    jefe_id = request.GET.get('jefe', '').strip()
    estado = request.GET.get('estado', '').strip()
    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()

    if jefe_id:
        solicitudes = solicitudes.filter(jefe_solicitante_id=jefe_id)
    if estado:
        solicitudes = solicitudes.filter(estado_local=estado)
    if fecha_desde:
        solicitudes = solicitudes.filter(fecha_inicio__gte=fecha_desde)
    if fecha_hasta:
        solicitudes = solicitudes.filter(fecha_fin__lte=fecha_hasta)

    paginator = Paginator(solicitudes, 50)
    page_obj = paginator.get_page(request.GET.get('page'))

    # Dropdown de jefes: solo empleados que efectivamente han solicitado alguna
    jefes_con_solicitudes = Empleado.objects.filter(
        vacaciones_solicitadas__isnull=False
    ).distinct().order_by('apellidos', 'nombres')

    return render(request, 'employees/vacaciones/admin.html', {
        'q': q,
        'empleados_buscados': empleados_buscados,
        'page_obj': page_obj,
        'jefes_con_solicitudes': jefes_con_solicitudes,
        'estado_choices': SolicitudVacacion.ESTADO_CHOICES,
        'filtros': request.GET,
    })