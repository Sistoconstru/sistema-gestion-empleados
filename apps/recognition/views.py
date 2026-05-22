from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import TemplateView, ListView, DetailView
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.db.models import Sum, Count, Q, Avg, Max, Value
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
from datetime import datetime, timedelta

# Importaciones para reportes (solo se usan cuando se necesitan)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .models import (
    HistorialPuntos, TipoActividad, Reconocimiento, TipoReconocimiento,
    InsigniaEmpleado, TipoInsignia, TipoBeneficio, CanjeoBeneficio
)
from apps.employees.models import Empleado


class DashboardView(LoginRequiredMixin, TemplateView):
    """Dashboard principal del módulo de reconocimientos"""
    
    def get_template_names(self):
        if self.request.user.is_staff:
            return ['recognition/admin_dashboard.html']
        return ['recognition/index.html']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if self.request.user.is_staff:
            # Vista de administrador
            context.update(self.get_admin_context())
        else:
            # Vista de empleado
            context.update(self.get_employee_context())
        
        return context
    
    def get_admin_context(self):
        """Contexto para administradores"""
        context = {}
        
        # Métricas generales del sistema
        context['total_reconocimientos'] = Reconocimiento.objects.count()
        context['total_insignias_otorgadas'] = InsigniaEmpleado.objects.count()
        context['total_empleados'] = Empleado.objects.exclude(estado__codigo='inactivo').count()
        context['beneficios_disponibles'] = TipoBeneficio.objects.filter(disponible=True).count()
        
        # Top empleados por puntos
        context['top_empleados'] = self.get_top_ranking()
        
        # Reconocimientos recientes
        context['reconocimientos_recientes'] = Reconocimiento.objects.select_related(
            'empleado', 'tipo_reconocimiento'
        ).order_by('-fecha_otorgamiento')[:10]
        
        # Canjes pendientes
        context['canjes_pendientes'] = CanjeoBeneficio.objects.filter(
    estado='solicitado'
).count()
        
        # Estadísticas de participación
        context['stats_participacion'] = self.get_stats_participacion()
        
        # Lista de empleados para selector
        empleados_queryset = Empleado.objects.exclude(
            estado__codigo='inactivo'
        ).prefetch_related(
            'historialcargo_set__cargo__area'
        ).annotate(
            puntos_totales=Sum(
                'historialpuntos__puntos',
                filter=Q(historialpuntos__validado=True)
            )
        ).order_by('nombres', 'apellidos')  # Removido límite para mostrar todos los empleados
        
        # Convertir a lista para poder agregar campos calculados
        empleados_lista = []
        for empleado in empleados_queryset:
            empleado.puntos_totales = empleado.puntos_totales or 0
            empleados_lista.append(empleado)
        
        context['empleados_lista'] = empleados_lista
        
        # Lista de áreas para filtros
        from apps.organizational.models import AreaEmpresa
        context['areas'] = AreaEmpresa.objects.all().order_by('nombre')
        
        # Insignias disponibles
        context['insignias_disponibles'] = TipoInsignia.objects.filter(
            activa=True
        ).order_by('nombre')
        
        # Tipos de actividad disponibles para asignación manual
        context['tipos_actividad'] = TipoActividad.objects.filter(
            activo=True
        ).order_by('nombre')
        
        return context
    
    def get_employee_context(self):
        """Contexto para empleados"""
        context = {}
        
        # Obtener empleado actual
        empleado = None
        if hasattr(self.request.user, 'empleado'):
            empleado = self.request.user.empleado
        
        # Métricas personales del usuario
        if empleado:
            context['mis_puntos_totales'] = self.get_puntos_totales(empleado)
            context['mis_puntos_mes'] = self.get_puntos_mes_actual(empleado)
            context['mis_insignias'] = self.get_mis_insignias(empleado)
            context['mi_posicion_ranking'] = self.get_posicion_ranking(empleado)
            
            # Calcular puntos disponibles para canjear
            puntos_totales = context['mis_puntos_totales']
            puntos_utilizados = CanjeoBeneficio.objects.filter(
                empleado=empleado,
                estado__in=['solicitado', 'aprobado', 'entregado']
            ).aggregate(total=Sum('puntos_utilizados'))['total'] or 0
            context['mis_puntos_disponibles'] = puntos_totales - puntos_utilizados
        else:
            context['mis_puntos_totales'] = 0
            context['mis_puntos_mes'] = 0
            context['mis_insignias'] = []
            context['mi_posicion_ranking'] = None
            context['mis_puntos_disponibles'] = 0
        
        # Métricas generales del sistema (para referencia)
        context['total_reconocimientos'] = Reconocimiento.objects.count()
        context['total_insignias_otorgadas'] = InsigniaEmpleado.objects.count()
        context['empleados_activos'] = Empleado.objects.exclude(estado__codigo='inactivo').count()
        context['beneficios_disponibles'] = TipoBeneficio.objects.filter(disponible=True).count()
        
        # Top 5 ranking general
        context['top_ranking'] = self.get_top_ranking()
        
        # Reconocimientos recientes
        context['reconocimientos_recientes'] = Reconocimiento.objects.select_related(
            'empleado', 'tipo_reconocimiento'
        ).order_by('-fecha_otorgamiento')[:5]
        
        # Insignias recientes
        context['insignias_recientes'] = InsigniaEmpleado.objects.select_related(
            'empleado', 'tipo_insignia'
        ).order_by('-fecha_otorgamiento')[:5]
        
        # Beneficios populares
        context['beneficios_populares'] = self.get_beneficios_populares()
        
        # Estadísticas de participación
        context['stats_participacion'] = self.get_stats_participacion()
        
        return context
    
    def get_puntos_totales(self, empleado):
        """Puntos totales acumulados por el empleado"""
        return HistorialPuntos.objects.filter(
            empleado=empleado, 
            validado=True
        ).aggregate(total=Sum('puntos'))['total'] or 0
    
    def get_puntos_mes_actual(self, empleado):
        """Puntos ganados en el mes actual"""
        inicio_mes = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        return HistorialPuntos.objects.filter(
            empleado=empleado,
            validado=True,
            fecha_obtencion__gte=inicio_mes
        ).aggregate(total=Sum('puntos'))['total'] or 0
    
    def get_mis_insignias(self, empleado):
        """Insignias del empleado ordenadas por fecha"""
        return InsigniaEmpleado.objects.filter(
            empleado=empleado
        ).select_related('tipo_insignia').order_by('-fecha_otorgamiento')[:6]
    
    def get_posicion_ranking(self, empleado):
        """Posición del empleado en el ranking general"""
        # Obtener todos los empleados con sus puntos totales
        ranking = Empleado.objects.exclude(estado__codigo='inactivo').annotate(
            puntos_totales=Coalesce(
                Sum('historialpuntos__puntos', filter=Q(historialpuntos__validado=True)), 
                Value(0)
            ),
            actividades_completadas=Count('historialpuntos', filter=Q(historialpuntos__validado=True)),
            insignias_count=Count('insigniaempleado')
        ).order_by('-puntos_totales', '-insignias_count', '-actividades_completadas')
        
        for posicion, emp in enumerate(ranking, 1):
            if emp.id == empleado.id:
                return posicion
        return None
    
    def get_top_ranking(self):
        """Top 5 empleados con más puntos"""
        return Empleado.objects.exclude(estado__codigo='inactivo').annotate(
            puntos_totales=Coalesce(
                Sum('historialpuntos__puntos', filter=Q(historialpuntos__validado=True)), 
                Value(0)
            ),
            actividades_completadas=Count('historialpuntos', filter=Q(historialpuntos__validado=True)),
            insignias_count=Count('insigniaempleado')
        ).order_by('-puntos_totales', '-insignias_count', '-actividades_completadas')[:5]
    
    def get_beneficios_populares(self):
        """Beneficios más canjeados"""
        return TipoBeneficio.objects.filter(disponible=True).annotate(
            canjes_count=Count('canjeobeneficio')
        ).order_by('-canjes_count')[:4]
    
    def get_stats_participacion(self):
        """Estadísticas de participación en el sistema"""
        total_empleados = Empleado.objects.exclude(estado__codigo='inactivo').count()
        empleados_con_puntos = HistorialPuntos.objects.values('empleado').distinct().count()
        empleados_con_insignias = InsigniaEmpleado.objects.values('empleado').distinct().count()
        
        return {
            'total_empleados': total_empleados,
            'empleados_con_puntos': empleados_con_puntos,
            'empleados_con_insignias': empleados_con_insignias,
            'porcentaje_participacion': round((empleados_con_puntos / total_empleados * 100), 1) if total_empleados > 0 else 0
        }


class MisPuntosView(LoginRequiredMixin, ListView):
    """Vista detallada de puntos del usuario"""
    template_name = 'recognition/mis_puntos.html'
    context_object_name = 'historial_puntos'
    paginate_by = 15
    
    def get_queryset(self):
        if hasattr(self.request.user, 'empleado'):
            return HistorialPuntos.objects.filter(
                empleado=self.request.user.empleado,
                validado=True
            ).select_related('tipo_actividad').order_by('-fecha_obtencion')
        return HistorialPuntos.objects.none()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if hasattr(self.request.user, 'empleado'):
            empleado = self.request.user.empleado
            
            # Estadísticas de puntos
            context['puntos_totales'] = HistorialPuntos.objects.filter(
                empleado=empleado
            ).aggregate(total=Sum('puntos'))['total'] or 0
            
            context['puntos_mes_actual'] = self.get_puntos_mes_actual(empleado)
            context['puntos_mes_anterior'] = self.get_puntos_mes_anterior(empleado)
            
            # Puntos por tipo de actividad
            context['puntos_por_actividad'] = HistorialPuntos.objects.filter(
                empleado=empleado
            ).values('tipo_actividad__nombre').annotate(
                total_puntos=Sum('puntos'),
                total_actividades=Count('id')
            ).order_by('-total_puntos')
            
            # Promedio mensual
            context['promedio_mensual'] = self.get_promedio_mensual(empleado)
            
        return context
    
    def get_puntos_mes_actual(self, empleado):
        inicio_mes = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        return HistorialPuntos.objects.filter(
            empleado=empleado, fecha_obtencion__gte=inicio_mes
        ).aggregate(total=Sum('puntos'))['total'] or 0
    
    def get_puntos_mes_anterior(self, empleado):
        inicio_mes_actual = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        inicio_mes_anterior = (inicio_mes_actual - timedelta(days=1)).replace(day=1)
        return HistorialPuntos.objects.filter(
            empleado=empleado,
            fecha_obtencion__gte=inicio_mes_anterior,
            fecha_obtencion__lt=inicio_mes_actual
        ).aggregate(total=Sum('puntos'))['total'] or 0
    
    def get_promedio_mensual(self, empleado):
        # Calcular promedio de los últimos 6 meses
        hace_6_meses = timezone.now() - timedelta(days=180)
        total_puntos = HistorialPuntos.objects.filter(
            empleado=empleado, fecha_obtencion__gte=hace_6_meses
        ).aggregate(total=Sum('puntos'))['total'] or 0
        return round(total_puntos / 6, 1)


class RankingView(LoginRequiredMixin, ListView):
    """Vista del ranking de empleados"""
    context_object_name = 'empleados_ranking'
    paginate_by = 20
    
    def get_template_names(self):
        if self.request.user.is_staff:
            return ['recognition/admin_ranking.html']
        return ['recognition/ranking.html']
    
    def get_queryset(self):
        queryset = Empleado.objects.exclude(estado__codigo='inactivo').annotate(
            puntos_totales=Coalesce(
                Sum('historialpuntos__puntos', filter=Q(historialpuntos__validado=True)), 
                Value(0)
            ),
            actividades_completadas=Count('historialpuntos', filter=Q(historialpuntos__validado=True)),
            insignias_count=Count('insigniaempleado')
        ).order_by('-puntos_totales', '-insignias_count', '-actividades_completadas')
        
        if not self.request.user.is_staff:
            # Empleados ven solo el top 20
            queryset = queryset[:20]
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if self.request.user.is_staff:
            context.update(self.get_admin_context())
        else:
            context.update(self.get_employee_context())
        
        return context
    
    def get_admin_context(self):
        """Contexto para administradores"""
        context = {}
        
        # Estadísticas generales
        context['total_empleados'] = Empleado.objects.exclude(
            estado__codigo='inactivo'
        ).count()
        
        context['empleados_con_puntos'] = Empleado.objects.exclude(
            estado__codigo='inactivo'
        ).filter(historialpuntos__validado=True).distinct().count()
        
        # Puntos totales distribuidos
        context['puntos_totales'] = HistorialPuntos.objects.filter(
            validado=True
        ).aggregate(total=Sum('puntos'))['total'] or 0
        
        # Promedio de puntos por empleado
        if context['empleados_con_puntos'] > 0:
            context['promedio_puntos'] = context['puntos_totales'] / context['empleados_con_puntos']
        else:
            context['promedio_puntos'] = 0
        
        # Distribución por áreas
        context['ranking_areas'] = Empleado.objects.exclude(
            estado__codigo='inactivo'
        ).values('historialcargo__cargo__area__nombre').annotate(
            total_puntos=Sum('historialpuntos__puntos', 
                           filter=Q(historialpuntos__validado=True)),
            cantidad_empleados=Count('id')
        ).filter(historialcargo__activo=True).order_by('-total_puntos')[:10]
        
        # Actividad reciente
        context['actividad_reciente'] = HistorialPuntos.objects.filter(
            validado=True
        ).select_related('empleado', 'validado_por').order_by('-fecha_obtencion')[:10]
        
        return context
    
    def get_employee_context(self):
        """Contexto para empleados"""
        context = {}
        
        if hasattr(self.request.user, 'empleado'):
            empleado_actual = self.request.user.empleado
            
            # Posición del empleado actual
            ranking = list(self.get_queryset())
            for posicion, empleado_ranking in enumerate(ranking, 1):
                if empleado_ranking.id == empleado_actual.id:
                    context['mi_posicion'] = posicion
                    break
            else:
                context['mi_posicion'] = None
            
            # Mis puntos
            context['mis_puntos'] = HistorialPuntos.objects.filter(
                empleado=empleado_actual, validado=True
            ).aggregate(total=Sum('puntos'))['total'] or 0
            
            # Mi área
            if empleado_actual.cargo_actual and empleado_actual.cargo_actual.cargo.area:
                context['mi_area'] = empleado_actual.cargo_actual.cargo.area.nombre
            else:
                context['mi_area'] = None
        else:
            context['mi_posicion'] = None
            context['mis_puntos'] = 0
            context['mi_area'] = None
        
        return context


class BeneficiosView(LoginRequiredMixin, ListView):
    """Vista de beneficios disponibles"""
    model = TipoBeneficio
    context_object_name = 'beneficios'
    
    def get_template_names(self):
        if self.request.user.is_staff:
            return ['recognition/admin_beneficios.html']
        return ['recognition/beneficios.html']
    
    def get_queryset(self):
        if self.request.user.is_staff:
            # Administradores ven todos los beneficios
            return TipoBeneficio.objects.all().order_by('costo_puntos')
        else:
            # Empleados solo ven beneficios disponibles
            return TipoBeneficio.objects.filter(disponible=True).order_by('costo_puntos')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if self.request.user.is_staff:
            # Contexto para administradores
            context.update(self.get_admin_context())
        else:
            # Contexto para empleados
            context.update(self.get_employee_context())
        
        return context
    
    def get_admin_context(self):
        """Contexto para administradores"""
        context = {}
        
        # Estadísticas generales
        context['total_beneficios'] = TipoBeneficio.objects.count()
        context['beneficios_activos'] = TipoBeneficio.objects.filter(disponible=True).count()
        context['total_canjes'] = CanjeoBeneficio.objects.count()
        context['canjes_pendientes'] = CanjeoBeneficio.objects.filter(estado='solicitado').count()
        
        # Canjes recientes
        context['canjes_recientes'] = CanjeoBeneficio.objects.select_related(
            'empleado', 'tipo_beneficio'
        ).order_by('-fecha_canje')[:10]
        
        # Beneficios más populares
        context['beneficios_populares'] = TipoBeneficio.objects.annotate(
            canjes_count=Count('canjeobeneficio')
        ).order_by('-canjes_count')[:5]
        
        # Lista de empleados para filtros
        context['empleados_lista'] = Empleado.objects.exclude(
            estado__codigo='inactivo'
        ).select_related('posicion', 'posicion__area').order_by('nombres', 'apellidos')[:50]
        
        # Lista de áreas para filtros
        from apps.organizational.models import AreaEmpresa
        context['areas'] = AreaEmpresa.objects.all().order_by('nombre')
        
        # Insignias disponibles
        context['insignias_disponibles'] = TipoInsignia.objects.filter(
            activa=True
        ).order_by('nombre')
        
        # Canjes pendientes para aprobación
        context['canjes_pendientes_lista'] = CanjeoBeneficio.objects.filter(
            estado='solicitado'
        ).select_related('empleado', 'tipo_beneficio').order_by('-fecha_canje')[:10]
        
        return context
    
    def get_employee_context(self):
        """Contexto para empleados"""
        context = {}
        
        if hasattr(self.request.user, 'empleado'):
            empleado = self.request.user.empleado
            
            # Calcular puntos disponibles
            puntos_totales = HistorialPuntos.objects.filter(
                empleado=empleado
            ).aggregate(total=Sum('puntos'))['total'] or 0
            
            puntos_utilizados = CanjeoBeneficio.objects.filter(
                empleado=empleado,
                estado__in=['solicitado', 'aprobado', 'entregado']
            ).aggregate(total=Sum('puntos_utilizados'))['total'] or 0
            
            context['puntos_totales'] = puntos_totales
            context['puntos_utilizados'] = puntos_utilizados
            context['puntos_disponibles'] = puntos_totales - puntos_utilizados
            
            # Mis canjes recientes
            context['mis_canjes'] = CanjeoBeneficio.objects.filter(
                empleado=empleado
            ).select_related('tipo_beneficio').order_by('-fecha_canje')[:5]
            
            # Categorías disponibles para filtros
            context['categorias'] = TipoBeneficio.objects.filter(
                disponible=True
            ).values_list('categoria', flat=True).distinct().exclude(categoria='')
            
        else:
            context['puntos_totales'] = 0
            context['puntos_utilizados'] = 0
            context['puntos_disponibles'] = 0
            context['mis_canjes'] = []
            context['categorias'] = []
        
        return context


class MisInsigniasView(LoginRequiredMixin, ListView):
    """Vista de insignias del usuario"""
    template_name = 'recognition/mis_insignias.html'
    context_object_name = 'mis_insignias'
    paginate_by = 12
    
    def get_queryset(self):
        if hasattr(self.request.user, 'empleado'):
            return InsigniaEmpleado.objects.filter(
                empleado=self.request.user.empleado
            ).select_related('tipo_insignia').order_by('-fecha_otorgamiento')
        return InsigniaEmpleado.objects.none()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Todas las insignias disponibles
        context['insignias_disponibles'] = TipoInsignia.objects.filter(activa=True).order_by('nivel', 'nombre')
        
        if hasattr(self.request.user, 'empleado'):
            empleado = self.request.user.empleado
            insignias_obtenidas = list(self.get_queryset().values_list('tipo_insignia_id', flat=True))
            context['insignias_obtenidas_ids'] = insignias_obtenidas
            context['total_insignias_obtenidas'] = len(insignias_obtenidas)
            context['total_insignias_disponibles'] = TipoInsignia.objects.filter(activa=True).count()
        
        return context


# =============================================================================
# VISTAS ADMINISTRATIVAS
# =============================================================================

class OtorgarPuntosView(LoginRequiredMixin, View):
    """Vista para otorgar puntos a empleados (solo administradores)"""
    
    def post(self, request):
        if not request.user.is_staff:
            return JsonResponse({'error': 'No tienes permisos para realizar esta acción'}, status=403)
        
        try:
            # Intentar obtener el empleado de diferentes fuentes
            empleado_id = (
                request.POST.get('empleado') or 
                request.POST.get('empleado_preseleccionado')
            )
            tipo_actividad_id = request.POST.get('tipo_actividad')
            descripcion = request.POST.get('descripcion', '')
            
            if not empleado_id or not tipo_actividad_id:
                return JsonResponse({'error': 'Datos inválidos'}, status=400)
            
            empleado = Empleado.objects.get(id=empleado_id)
            tipo_actividad = TipoActividad.objects.get(id=tipo_actividad_id)
            
            # Calcular puntos usando puntos_base del tipo de actividad
            puntos = tipo_actividad.puntos_base
            
            # Si es asignación manual, permitir puntos personalizados
            if tipo_actividad.codigo == 'MANUAL_ADMIN':
                puntos_custom = request.POST.get('puntos_custom')
                if puntos_custom:
                    puntos = int(puntos_custom)
                else:
                    return JsonResponse({'error': 'Debe especificar la cantidad de puntos para asignación manual'}, status=400)
            
            # Crear historial de puntos
            historial = HistorialPuntos.objects.create(
                empleado=empleado,
                tipo_actividad=tipo_actividad,
                puntos=puntos,
                descripcion=descripcion,
                validado=True,
                validado_por=request.user
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Se otorgaron {puntos} puntos a {empleado.nombre_completo} por {tipo_actividad.nombre}',
                'empleado': empleado.nombre_completo,
                'puntos': puntos,
                'tipo_actividad': tipo_actividad.nombre
            })
            
        except Empleado.DoesNotExist:
            return JsonResponse({'error': 'Empleado no encontrado'}, status=404)
        except TipoActividad.DoesNotExist:
            return JsonResponse({'error': 'Tipo de actividad no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'error': f'Error interno: {str(e)}'}, status=500)


class AsignarInsigniaView(LoginRequiredMixin, View):
    """Vista para asignar insignias a empleados (solo administradores)"""
    
    def post(self, request):
        if not request.user.is_staff:
            return JsonResponse({'error': 'No tienes permisos para realizar esta acción'}, status=403)
        
        try:
            # Intentar obtener el empleado de diferentes fuentes
            empleado_id = (
                request.POST.get('empleado') or 
                request.POST.get('empleado_preseleccionado')
            )
            insignia_id = request.POST.get('insignia')
            observaciones = request.POST.get('observaciones', '')
            
            if not empleado_id or not insignia_id:
                return JsonResponse({'error': 'Datos inválidos'}, status=400)
            
            empleado = Empleado.objects.get(id=empleado_id)
            tipo_insignia = TipoInsignia.objects.get(id=insignia_id)
            
            # Verificar si ya tiene la insignia
            if InsigniaEmpleado.objects.filter(empleado=empleado, tipo_insignia=tipo_insignia).exists():
                return JsonResponse({'error': f'{empleado.nombre_completo} ya tiene esta insignia'}, status=400)
            
            # Crear insignia del empleado
            insignia_empleado = InsigniaEmpleado.objects.create(
                empleado=empleado,
                tipo_insignia=tipo_insignia,
                justificacion=observaciones,
                otorgado_automaticamente=False,
                otorgado_por=request.user
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Se asignó la insignia "{tipo_insignia.nombre}" a {empleado.nombre_completo}',
                'empleado': empleado.nombre_completo,
                'insignia': tipo_insignia.nombre
            })
            
        except Empleado.DoesNotExist:
            return JsonResponse({'error': 'Empleado no encontrado'}, status=404)
        except TipoInsignia.DoesNotExist:
            return JsonResponse({'error': 'Insignia no encontrada'}, status=404)
        except Exception as e:
            return JsonResponse({'error': f'Error interno: {str(e)}'}, status=500)


class HistorialEmpleadoView(LoginRequiredMixin, View):
    """Vista AJAX para obtener el historial de reconocimientos de un empleado"""
    
    def get(self, request, empleado_id):
        if not request.user.is_staff:
            return JsonResponse({
                'success': False,
                'message': 'No tienes permisos para ver esta información'
            })
        
        try:
            empleado = Empleado.objects.get(id=empleado_id)
            
            # Obtener historial de puntos
            historial_puntos = HistorialPuntos.objects.filter(
                empleado=empleado
            ).select_related('tipo_actividad', 'validado_por').order_by('-fecha_obtencion')
            
            # Obtener historial de insignias
            historial_insignias = InsigniaEmpleado.objects.filter(
                empleado=empleado
            ).select_related('tipo_insignia', 'otorgado_por').order_by('-fecha_otorgamiento')
            
            # Crear contexto para el template
            context = {
                'empleado': empleado,
                'historial_puntos': historial_puntos,
                'historial_insignias': historial_insignias,
                'total_puntos': sum(hp.puntos for hp in historial_puntos),
                'total_insignias': historial_insignias.count()
            }
            
            # Renderizar el template del historial
            html_content = render_to_string(
                'recognition/historial_empleado.html', 
                context,
                request=request
            )
            
            return JsonResponse({
                'success': True,
                'html': html_content
            })
            
        except Empleado.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Empleado no encontrado'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al cargar el historial: {str(e)}'
            })


class CanjearBeneficioView(LoginRequiredMixin, View):
    """Vista para canjear un beneficio específico"""
    
    def post(self, request, beneficio_id):
        try:
            # Obtener el beneficio
            beneficio = TipoBeneficio.objects.get(id=beneficio_id)
            
            # Obtener el empleado
            empleado = None
            if hasattr(request.user, 'empleado'):
                empleado = request.user.empleado
            else:
                return JsonResponse({'error': 'Usuario no asociado a empleado'}, status=400)
            
            # Validar que el beneficio esté disponible
            if not beneficio.esta_disponible():
                return JsonResponse({'error': 'Este beneficio no está disponible actualmente'}, status=400)
            
            # Validar stock
            if not beneficio.tiene_stock_suficiente():
                return JsonResponse({'error': 'No hay stock disponible para este beneficio'}, status=400)
            
            # Calcular puntos totales del empleado
            puntos_totales = HistorialPuntos.objects.filter(empleado=empleado).aggregate(
                total=Sum('puntos')
            )['total'] or 0
            
            # Restar puntos ya utilizados en canjes
            puntos_utilizados = CanjeoBeneficio.objects.filter(
                empleado=empleado,
                estado__in=['solicitado', 'aprobado', 'entregado']
            ).aggregate(total=Sum('puntos_utilizados'))['total'] or 0
            
            puntos_disponibles = puntos_totales - puntos_utilizados
            
            # Validar puntos suficientes
            if puntos_disponibles < beneficio.costo_puntos:
                return JsonResponse({
                    'error': f'Puntos insuficientes. Necesitas {beneficio.costo_puntos} puntos, tienes {puntos_disponibles}'
                }, status=400)
            
            # Crear el canje
            canje = CanjeoBeneficio.objects.create(
                empleado=empleado,
                tipo_beneficio=beneficio,
                puntos_utilizados=beneficio.costo_puntos,
                estado='solicitado'
            )
            
            # Reducir stock si aplica
            if beneficio.stock_actual is not None:
                beneficio.stock_actual -= 1
                beneficio.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Canje solicitado exitosamente. Código: {canje.codigo_canje}',
                'codigo_canje': canje.codigo_canje,
                'puntos_utilizados': beneficio.costo_puntos,
                'puntos_restantes': puntos_disponibles - beneficio.costo_puntos
            })
            
        except TipoBeneficio.DoesNotExist:
            return JsonResponse({'error': 'Beneficio no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'error': f'Error interno: {str(e)}'}, status=500)


class AprobarCanjeView(LoginRequiredMixin, View):
    """Vista para aprobar canjes (solo administradores)"""
    
    def post(self, request, canje_id):
        if not request.user.is_staff:
            return JsonResponse({'error': 'No tienes permisos para realizar esta acción'}, status=403)
        
        try:
            import json
            
            canje = CanjeoBeneficio.objects.get(id=canje_id)
            
            if canje.estado != 'solicitado':
                return JsonResponse({'error': 'Este canje ya fue procesado'}, status=400)
            
            # Obtener datos del formulario
            data = json.loads(request.body.decode('utf-8'))
            comentarios = data.get('comentarios', '')
            fecha_entrega = data.get('fecha_entrega')
            
            # Aprobar el canje
            from django.utils import timezone
            canje.estado = 'aprobado'
            canje.fecha_aprobacion = timezone.now()
            canje.aprobado_por = request.user
            
            # Agregar comentarios si se proporcionaron
            if comentarios:
                canje.observaciones = comentarios
            
            # Programar fecha de reclamo si se proporcionó
            if fecha_entrega:
                try:
                    fecha_entrega_dt = datetime.strptime(fecha_entrega, '%Y-%m-%d').date()
                    canje.fecha_reclamo_programada = fecha_entrega_dt
                except ValueError:
                    pass
            
            canje.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Canje aprobado exitosamente. Código: {canje.codigo_canje}'
            })
            
        except CanjeoBeneficio.DoesNotExist:
            return JsonResponse({'error': 'Canje no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'error': f'Error interno: {str(e)}'}, status=500)


class RechazarCanjeView(LoginRequiredMixin, View):
    """Vista para rechazar canjes (solo administradores)"""
    
    def post(self, request, canje_id):
        if not request.user.is_staff:
            return JsonResponse({'error': 'No tienes permisos para realizar esta acción'}, status=403)
        
        try:
            import json
            
            canje = CanjeoBeneficio.objects.get(id=canje_id)
            
            if canje.estado != 'solicitado':
                return JsonResponse({'error': 'Este canje ya fue procesado'}, status=400)
            
            # Obtener datos del formulario
            data = json.loads(request.body.decode('utf-8'))
            motivo = data.get('motivo', 'Sin motivo especificado')
            
            # Rechazar el canje
            from django.utils import timezone
            canje.estado = 'rechazado'
            canje.aprobado_por = request.user
            canje.observaciones = f'Rechazado: {motivo}'
            canje.save()
            
            # Restaurar stock si aplica
            if canje.tipo_beneficio.stock_actual is not None:
                canje.tipo_beneficio.stock_actual += 1
                canje.tipo_beneficio.save()
            
            # Restaurar puntos al empleado
            HistorialPuntos.objects.create(
                empleado=canje.empleado,
                tipo_actividad=None,
                puntos=canje.puntos_utilizados,
                descripcion=f'Reembolso por canje rechazado: {canje.tipo_beneficio.nombre}',
                validado=True,
                validado_por=request.user
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Canje rechazado. Puntos y stock restaurados.'
            })
            
        except CanjeoBeneficio.DoesNotExist:
            return JsonResponse({'error': 'Canje no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'error': f'Error interno: {str(e)}'}, status=500)


class EntregarCanjeView(LoginRequiredMixin, View):
    """Vista para marcar canjes como entregados"""
    
    def post(self, request, canje_id):
        if not request.user.is_staff:
            return JsonResponse({'error': 'No tienes permisos para realizar esta acción'}, status=403)
        
        try:
            import json
            
            canje = CanjeoBeneficio.objects.get(id=canje_id)
            
            if canje.estado != 'aprobado':
                return JsonResponse({'error': 'Este canje no está aprobado'}, status=400)
            
            # Obtener datos del formulario
            data = json.loads(request.body.decode('utf-8'))
            comentarios_entrega = data.get('comentarios_entrega', '')
            
            # Marcar como entregado
            from django.utils import timezone
            canje.estado = 'entregado'
            canje.fecha_entrega = timezone.now()
            
            # Agregar comentarios de entrega
            if comentarios_entrega:
                if canje.observaciones:
                    canje.observaciones += f'\nEntrega: {comentarios_entrega}'
                else:
                    canje.observaciones = f'Entrega: {comentarios_entrega}'
            
            canje.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Canje marcado como entregado exitosamente.'
            })
            
        except CanjeoBeneficio.DoesNotExist:
            return JsonResponse({'error': 'Canje no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'error': f'Error interno: {str(e)}'}, status=500)


class MisCanjesView(LoginRequiredMixin, ListView):
    """Vista para que empleados vean sus canjes"""
    model = CanjeoBeneficio
    template_name = 'recognition/mis_canjes.html'
    context_object_name = 'canjes'
    paginate_by = 10
    
    def get_queryset(self):
        if hasattr(self.request.user, 'empleado'):
            return CanjeoBeneficio.objects.filter(
                empleado=self.request.user.empleado
            ).select_related('tipo_beneficio').order_by('-fecha_canje')
        return CanjeoBeneficio.objects.none()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if hasattr(self.request.user, 'empleado'):
            empleado = self.request.user.empleado
            
            # Calcular puntos disponibles
            puntos_totales = HistorialPuntos.objects.filter(empleado=empleado).aggregate(
                total=Sum('puntos')
            )['total'] or 0
            
            puntos_utilizados = CanjeoBeneficio.objects.filter(
                empleado=empleado,
                estado__in=['solicitado', 'aprobado', 'entregado']
            ).aggregate(total=Sum('puntos_utilizados'))['total'] or 0
            
            context['puntos_disponibles'] = puntos_totales - puntos_utilizados
            context['puntos_totales'] = puntos_totales
            context['puntos_utilizados'] = puntos_utilizados
            
            # Estadísticas de canjes
            context['total_canjes'] = self.get_queryset().count()
            context['canjes_aprobados'] = self.get_queryset().filter(estado='aprobado').count()
            context['canjes_pendientes'] = self.get_queryset().filter(estado='solicitado').count()
        
        return context


class MisLogrosView(LoginRequiredMixin, TemplateView):
    """Vista del resumen de logros del empleado"""
    template_name = 'recognition/mis_logros.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if hasattr(self.request.user, 'empleado'):
            empleado = self.request.user.empleado
            
            # Estadísticas de puntos
            historial_puntos = HistorialPuntos.objects.filter(empleado=empleado)
            puntos_totales = historial_puntos.aggregate(total=Sum('puntos'))['total'] or 0
            
            puntos_utilizados = CanjeoBeneficio.objects.filter(
                empleado=empleado,
                estado__in=['solicitado', 'aprobado', 'entregado']
            ).aggregate(total=Sum('puntos_utilizados'))['total'] or 0
            
            context.update({
                'puntos_totales': puntos_totales,
                'puntos_utilizados': puntos_utilizados,
                'puntos_disponibles': puntos_totales - puntos_utilizados,
                
                # Historial reciente de puntos (últimos 10)
                'historial_reciente': historial_puntos.select_related('tipo_actividad').order_by('-fecha_obtencion')[:10],
                
                # Estadísticas de insignias
                'insignias_obtenidas': InsigniaEmpleado.objects.filter(empleado=empleado).select_related('tipo_insignia').order_by('-fecha_otorgamiento'),
                'total_insignias': InsigniaEmpleado.objects.filter(empleado=empleado).count(),
                
                # Estadísticas de canjes
                'canjes_recientes': CanjeoBeneficio.objects.filter(empleado=empleado).select_related('tipo_beneficio').order_by('-fecha_canje')[:5],
                'total_canjes': CanjeoBeneficio.objects.filter(empleado=empleado).count(),
                'canjes_exitosos': CanjeoBeneficio.objects.filter(empleado=empleado, estado__in=['aprobado', 'entregado']).count(),
                
                # Actividades más frecuentes
                'actividades_frecuentes': historial_puntos.values('tipo_actividad__nombre', 'tipo_actividad__puntos_base')
                    .annotate(
                        count=Count('id'),
                        total_puntos=Sum('puntos')
                    ).order_by('-count')[:5],
                
                # Progreso mensual (últimos 6 meses)
                'progreso_mensual': self.get_progreso_mensual(empleado),
            })
        else:
            # Usuario sin empleado asociado
            context.update({
                'puntos_totales': 0,
                'puntos_utilizados': 0,
                'puntos_disponibles': 0,
                'historial_reciente': [],
                'insignias_obtenidas': [],
                'total_insignias': 0,
                'canjes_recientes': [],
                'total_canjes': 0,
                'canjes_exitosos': 0,
                'actividades_frecuentes': [],
                'progreso_mensual': [],
            })
        
        return context
    
    def get_progreso_mensual(self, empleado):
        """Obtiene el progreso de puntos de los últimos 6 meses"""
        # Calcular fecha límite (6 meses atrás)
        fecha_limite = timezone.now() - timedelta(days=180)
        
        # Obtener historial de puntos de los últimos 6 meses
        historial = HistorialPuntos.objects.filter(
            empleado=empleado,
            fecha_obtencion__gte=fecha_limite
        ).values('fecha_obtencion__year', 'fecha_obtencion__month').annotate(
            puntos_mes=Sum('puntos'),
            actividades_mes=Count('id')
        ).order_by('fecha_obtencion__year', 'fecha_obtencion__month')
        
        return list(historial)


class AdminCanjesView(LoginRequiredMixin, TemplateView):
    """Vista de administración de canjes para administradores"""
    template_name = 'recognition/admin_canjes.html'
    
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_staff:
            messages.error(request, 'No tienes permisos para acceder a esta sección')
            return redirect('recognition:index')
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Filtros
        estado_filter = self.request.GET.get('estado', 'todos')
        fecha_desde = self.request.GET.get('fecha_desde')
        fecha_hasta = self.request.GET.get('fecha_hasta')
        beneficio_filter = self.request.GET.get('beneficio')
        
        # Query base
        canjes = CanjeoBeneficio.objects.select_related(
            'empleado', 'tipo_beneficio', 'aprobado_por'
        ).order_by('-fecha_canje')
        
        # Aplicar filtros
        if estado_filter and estado_filter != 'todos':
            canjes = canjes.filter(estado=estado_filter)
        
        if fecha_desde:
            try:
                fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
                canjes = canjes.filter(fecha_canje__gte=fecha_desde_dt)
            except ValueError:
                pass
        
        if fecha_hasta:
            try:
                fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
                canjes = canjes.filter(fecha_canje__lte=fecha_hasta_dt)
            except ValueError:
                pass
        
        if beneficio_filter:
            canjes = canjes.filter(tipo_beneficio__id=beneficio_filter)
        
        # Estadísticas
        total_canjes = CanjeoBeneficio.objects.count()
        canjes_pendientes = CanjeoBeneficio.objects.filter(estado='solicitado').count()
        canjes_aprobados = CanjeoBeneficio.objects.filter(estado='aprobado').count()
        canjes_entregados = CanjeoBeneficio.objects.filter(estado='entregado').count()
        canjes_rechazados = CanjeoBeneficio.objects.filter(estado='rechazado').count()
        
        # Lista de beneficios para filtro
        beneficios = TipoBeneficio.objects.all().order_by('nombre')
        
        context.update({
            'canjes': canjes[:50],  # Limitar a 50 resultados
            'total_canjes': total_canjes,
            'canjes_pendientes': canjes_pendientes,
            'canjes_aprobados': canjes_aprobados,
            'canjes_entregados': canjes_entregados,
            'canjes_rechazados': canjes_rechazados,
            'beneficios': beneficios,
            'estado_filter': estado_filter,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'beneficio_filter': beneficio_filter,
            'estados_choices': CanjeoBeneficio.ESTADOS,
        })
        
        return context


class GenerarReporteRankingView(LoginRequiredMixin, View):
    """Vista para generar reporte del ranking en Excel"""
    
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_staff:
            messages.error(request, 'No tienes permisos para generar reportes')
            return redirect('recognition:ranking')
        return super().dispatch(request, *args, **kwargs)
    
    def post(self, request):
        try:
            if not OPENPYXL_AVAILABLE:
                messages.error(request, 'No se puede generar el reporte: librería openpyxl no disponible')
                return redirect('recognition:ranking')
            
            # Obtener datos del ranking
            empleados_ranking = Empleado.objects.filter(
                estado__codigo__in=['999', 'p-prue']
            ).annotate(
                total_puntos=Sum('historial_puntos__puntos', default=0)
            ).order_by('-total_puntos')[:50]  # Top 50
            
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Ranking General"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers
            headers = [
                'Posición', 'Nombres', 'Apellidos', 'Correo', 'Departamento', 
                'Cargo', 'Total Puntos', 'Fecha de Última Actividad'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Datos
            for row, empleado in enumerate(empleados_ranking, 2):
                # Obtener última actividad
                ultima_actividad = HistorialPuntos.objects.filter(
                    empleado=empleado
                ).order_by('-fecha_obtencion').first()
                
                fecha_ultima = ultima_actividad.fecha_obtencion if ultima_actividad else None
                
                data = [
                    row - 1,  # Posición
                    empleado.nombres,
                    empleado.apellidos,
                    empleado.correo_electronico,
                    empleado.departamento.nombre if empleado.departamento else 'N/A',
                    empleado.nombre_cargo_actual if empleado.nombre_cargo_actual else 'N/A',
                    empleado.total_puntos,
                    fecha_ultima.strftime('%d/%m/%Y %H:%M') if fecha_ultima else 'Sin actividad'
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    if col == 1:  # Posición
                        cell.alignment = Alignment(horizontal='center')
                    elif col == 7:  # Puntos
                        cell.alignment = Alignment(horizontal='right')
            
            # Ajustar ancho de columnas
            column_widths = [10, 20, 20, 30, 20, 25, 15, 20]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
            
            # Agregar información adicional
            ws.cell(row=len(empleados_ranking) + 3, column=1, value="Reporte generado:")
            ws.cell(row=len(empleados_ranking) + 3, column=2, value=datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
            ws.cell(row=len(empleados_ranking) + 4, column=1, value="Generado por:")
            ws.cell(row=len(empleados_ranking) + 4, column=2, value=f"{request.user.first_name} {request.user.last_name}")
            
            # Crear respuesta HTTP
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="ranking_general_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            # Guardar workbook en response
            wb.save(response)
            return response
            
        except Exception as e:
            messages.error(request, f'Error al generar el reporte: {str(e)}')
            return redirect('recognition:ranking')


class GenerarReporteView(LoginRequiredMixin, View):
    """Vista para generar diferentes tipos de reportes en Excel"""
    
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_staff:
            messages.error(request, 'No tienes permisos para generar reportes')
            return redirect('recognition:index')
        return super().dispatch(request, *args, **kwargs)
    
    def post(self, request):
        try:
            if not OPENPYXL_AVAILABLE:
                messages.error(request, 'No se puede generar el reporte. Contacta al administrador del sistema.')
                return redirect('recognition:index')
            
            tipo_reporte = request.POST.get('tipo_reporte')
            
            if tipo_reporte == 'ranking':
                return self.generar_reporte_ranking()
            elif tipo_reporte == 'puntos':
                return self.generar_reporte_puntos()
            elif tipo_reporte == 'canjes':
                return self.generar_reporte_canjes()
            elif tipo_reporte == 'insignias':
                return self.generar_reporte_insignias()
            else:
                messages.error(request, 'Tipo de reporte no válido')
                return redirect('recognition:index')
                
        except Exception as e:
            messages.error(request, f'Error al generar el reporte: {str(e)}')
            return redirect('recognition:index')
    
    def generar_reporte_ranking(self):
        """Generar reporte de ranking de empleados"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Ranking de Empleados"
        
        # Headers
        headers = ['Posición', 'Empleado', 'Cédula', 'Cargo', 'Área', 'Puntos Totales', 'Actividades Completadas', 'Insignias']
        ws.append(headers)
        
        # Datos
        empleados = Empleado.objects.exclude(estado__codigo='inactivo').annotate(
            puntos_totales=Coalesce(
                Sum('historialpuntos__puntos', filter=Q(historialpuntos__validado=True)), 
                Value(0)
            ),
            actividades_completadas=Count('historialpuntos', filter=Q(historialpuntos__validado=True)),
            insignias_count=Count('insigniaempleado')
        ).order_by('-puntos_totales', '-insignias_count', '-actividades_completadas')
        
        for i, empleado in enumerate(empleados, 1):
            ws.append([
                i,
                f"{empleado.nombres} {empleado.apellidos}",
                empleado.numero_documento,
                empleado.nombre_cargo_actual or 'Sin cargo',
                empleado.area_actual.nombre if empleado.area_actual else 'Sin área',
                empleado.puntos_totales,
                empleado.actividades_completadas,
                empleado.insignias_count
            ])
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="ranking_empleados_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response
    
    def generar_reporte_puntos(self):
        """Generar reporte del historial de puntos"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial de Puntos"
        
        # Headers
        headers = ['Fecha', 'Empleado', 'Cédula', 'Actividad', 'Puntos', 'Descripción', 'Validado', 'Validado Por']
        ws.append(headers)
        
        # Datos
        historiales = HistorialPuntos.objects.select_related(
            'empleado', 'tipo_actividad', 'validado_por'
        ).order_by('-fecha_obtencion')
        
        for historial in historiales:
            ws.append([
                historial.fecha_obtencion.strftime('%Y-%m-%d %H:%M'),
                f"{historial.empleado.nombres} {historial.empleado.apellidos}",
                historial.empleado.numero_documento,
                historial.tipo_actividad.nombre,
                historial.puntos,
                historial.descripcion,
                'Sí' if historial.validado else 'No',
                f"{historial.validado_por.first_name} {historial.validado_por.last_name}" if historial.validado_por else 'Sistema'
            ])
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="historial_puntos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response
    
    def generar_reporte_canjes(self):
        """Generar reporte de canjes de beneficios"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Canjes de Beneficios"
        
        # Headers
        headers = ['Fecha Canje', 'Empleado', 'Cédula', 'Beneficio', 'Puntos Utilizados', 'Estado', 'Código Canje', 'Fecha Aprobación', 'Fecha Reclamo', 'Aprobado Por']
        ws.append(headers)
        
        # Datos
        canjes = CanjeoBeneficio.objects.select_related(
            'empleado', 'tipo_beneficio', 'aprobado_por'
        ).order_by('-fecha_canje')
        
        for canje in canjes:
            ws.append([
                canje.fecha_canje.strftime('%Y-%m-%d %H:%M'),
                f"{canje.empleado.nombres} {canje.empleado.apellidos}",
                canje.empleado.numero_documento,
                canje.tipo_beneficio.nombre,
                canje.puntos_utilizados,
                canje.get_estado_display(),
                canje.codigo_canje,
                canje.fecha_aprobacion.strftime('%Y-%m-%d %H:%M') if canje.fecha_aprobacion else '',
                canje.fecha_reclamo_programada.strftime('%Y-%m-%d') if canje.fecha_reclamo_programada else '',
                f"{canje.aprobado_por.first_name} {canje.aprobado_por.last_name}" if canje.aprobado_por else ''
            ])
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="canjes_beneficios_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response
    
    def generar_reporte_insignias(self):
        """Generar reporte de insignias otorgadas"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Insignias Otorgadas"
        
        # Headers
        headers = ['Fecha Otorgamiento', 'Empleado', 'Cédula', 'Insignia', 'Descripción', 'Justificación', 'Otorgado Por', 'Automático']
        ws.append(headers)
        
        # Datos
        insignias = InsigniaEmpleado.objects.select_related(
            'empleado', 'tipo_insignia', 'otorgado_por'
        ).order_by('-fecha_otorgamiento')
        
        for insignia in insignias:
            ws.append([
                insignia.fecha_otorgamiento.strftime('%Y-%m-%d %H:%M'),
                f"{insignia.empleado.nombres} {insignia.empleado.apellidos}",
                insignia.empleado.numero_documento,
                insignia.tipo_insignia.nombre,
                insignia.tipo_insignia.descripcion,
                insignia.justificacion or '',
                f"{insignia.otorgado_por.first_name} {insignia.otorgado_por.last_name}" if insignia.otorgado_por else 'Sistema',
                'Sí' if insignia.otorgado_automaticamente else 'No'
            ])
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="insignias_otorgadas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response