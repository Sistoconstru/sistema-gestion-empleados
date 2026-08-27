"""Vistas del módulo de sorteos.

Empleado:
- lista de sorteos abiertos con estado de elegibilidad y opción de
  autoinscribirse si cumple los requisitos (PWA instalada + encuesta
  requisito completada).

Admin (staff/superuser):
- lista, crear, editar, ver inscritos, pantalla de sorteo (ingresar
  número ganador y registrar), export XLSX de inscritos.
"""
from datetime import date
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.db.models import Max
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views import View
from django.views.generic import ListView

from .models import Sorteo, InscripcionSorteo, GanadorSorteo


def _empleado_de(user):
    return getattr(user, 'empleado', None)


def _tiene_pwa(user):
    """Proxy: tiene al menos una PushSubscription activa (indicador de PWA
    instalada / navegador aceptando notificaciones de SIGHU)."""
    from apps.notifications.models import PushSubscription
    return PushSubscription.objects.filter(usuario=user, activa=True).exists()


def _respondio_encuesta(empleado, encuesta):
    from apps.surveys.models import ParticipacionEncuesta
    if not (empleado and encuesta):
        return False
    return ParticipacionEncuesta.objects.filter(
        empleado=empleado, encuesta=encuesta, completada=True,
    ).exists()


def _estado_inscripcion(sorteo, request):
    """Devuelve dict con flags {puede_inscribirse, ya_inscrito, requisitos,
    razon}. Usado tanto en la lista como en el POST de inscripción."""
    empleado = _empleado_de(request.user)
    ya = InscripcionSorteo.objects.filter(sorteo=sorteo, empleado=empleado).first() if empleado else None
    hoy = date.today()

    r = {
        'sorteo': sorteo,
        'empleado': empleado,
        'ya_inscrito': ya is not None,
        'mi_numero': ya.numero if ya else None,
        'abierto': sorteo.inscripciones_abiertas(hoy),
        'tiene_pwa': _tiene_pwa(request.user) if sorteo.require_pwa else True,
        'respondio_encuesta': _respondio_encuesta(empleado, sorteo.encuesta_requisito),
        'url_encuesta': (
            f'/encuestas/responder/{sorteo.encuesta_requisito.pk}/'
            if sorteo.encuesta_requisito else None
        ),
    }
    r['puede_inscribirse'] = (
        empleado is not None and r['abierto'] and not r['ya_inscrito']
        and r['tiene_pwa'] and r['respondio_encuesta']
    )
    return r


# ============================================================================
# EMPLEADO
# ============================================================================

class SorteosIndexView(LoginRequiredMixin, View):
    template_name = 'sorteos/index.html'

    def get(self, request):
        hoy = date.today()
        sorteos = Sorteo.objects.filter(activo=True).order_by('-fecha_creacion')
        estados = [_estado_inscripcion(s, request) for s in sorteos]
        return render(request, self.template_name, {
            'estados': estados, 'hoy': hoy,
        })


class InscribirseSorteoView(LoginRequiredMixin, View):
    def post(self, request, pk):
        sorteo = get_object_or_404(Sorteo, pk=pk, activo=True)
        estado = _estado_inscripcion(sorteo, request)

        if estado['ya_inscrito']:
            messages.info(request,
                f'Ya estabas inscrito con el número {estado["mi_numero"]}.')
            return redirect('sorteos:index')

        if not estado['abierto']:
            messages.error(request, 'Las inscripciones a este sorteo están cerradas.')
            return redirect('sorteos:index')

        if not estado['empleado']:
            messages.error(request, 'No tienes ficha de empleado activa.')
            return redirect('sorteos:index')

        if sorteo.require_pwa and not estado['tiene_pwa']:
            messages.error(request,
                'Debes tener SIGHU instalada como aplicación (PWA) y notificaciones '
                'activas para inscribirte.')
            return redirect('sorteos:index')

        if not estado['respondio_encuesta']:
            if sorteo.encuesta_requisito:
                messages.warning(request,
                    f'Antes de inscribirte debes responder la encuesta '
                    f'"{sorteo.encuesta_requisito.nombre}".')
                return redirect(f'/encuestas/responder/{sorteo.encuesta_requisito.pk}/')
            messages.error(request, 'No cumples los requisitos para inscribirte.')
            return redirect('sorteos:index')

        # Asignar el siguiente número disponible con lock
        with transaction.atomic():
            ultimo = InscripcionSorteo.objects.select_for_update().filter(
                sorteo=sorteo,
            ).aggregate(m=Max('numero'))['m'] or 0
            inscripcion = InscripcionSorteo.objects.create(
                sorteo=sorteo, empleado=estado['empleado'], numero=ultimo + 1,
            )
        messages.success(request,
            f'¡Te inscribiste al sorteo "{sorteo.nombre}" con el número {inscripcion.numero}!')
        return redirect('sorteos:index')


# ============================================================================
# ADMIN
# ============================================================================

class _StaffRequiredMixin:
    def dispatch(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser):
            messages.error(request, 'No tienes permisos.')
            return redirect('sorteos:index')
        return super().dispatch(request, *args, **kwargs)


class SorteoAdminListView(LoginRequiredMixin, _StaffRequiredMixin, ListView):
    model = Sorteo
    template_name = 'sorteos/admin/lista.html'
    context_object_name = 'sorteos'
    paginate_by = 20

    def get_queryset(self):
        return Sorteo.objects.all().order_by('-fecha_creacion')


class SorteoFormView(LoginRequiredMixin, _StaffRequiredMixin, View):
    """Crear o editar un sorteo (mismo formulario)."""
    template_name = 'sorteos/admin/form.html'

    def _contexto(self, sorteo=None):
        from apps.surveys.models import Encuesta
        return {
            'sorteo': sorteo,
            'encuestas': Encuesta.objects.all().order_by('-fecha_creacion'),
            'hoy_iso': date.today().isoformat(),
        }

    def get(self, request, pk=None):
        sorteo = get_object_or_404(Sorteo, pk=pk) if pk else None
        return render(request, self.template_name, self._contexto(sorteo))

    def post(self, request, pk=None):
        sorteo = get_object_or_404(Sorteo, pk=pk) if pk else None
        try:
            codigo = request.POST.get('codigo', '').strip().upper()
            nombre = request.POST.get('nombre', '').strip()
            descripcion = request.POST.get('descripcion', '').strip()
            cantidad = int(request.POST.get('cantidad_premios') or 1)
            encuesta_id = request.POST.get('encuesta_requisito')
            require_pwa = request.POST.get('require_pwa') == 'on'
            fecha_ini = request.POST.get('fecha_inicio_inscripcion')
            fecha_fin = request.POST.get('fecha_fin_inscripcion')
            fecha_sor = request.POST.get('fecha_sorteo')
            activo = request.POST.get('activo') == 'on'

            if not all([codigo, nombre, encuesta_id, fecha_ini, fecha_fin, fecha_sor]):
                messages.error(request, 'Todos los campos marcados con * son obligatorios.')
                return redirect(request.path)

            if sorteo is None:
                if Sorteo.objects.filter(codigo=codigo).exists():
                    messages.error(request, f'El código "{codigo}" ya existe.')
                    return redirect('sorteos:crear')
                sorteo = Sorteo(codigo=codigo, creado_por=request.user)

            sorteo.nombre = nombre
            sorteo.descripcion = descripcion
            sorteo.cantidad_premios = max(1, cantidad)
            sorteo.encuesta_requisito_id = encuesta_id
            sorteo.require_pwa = require_pwa
            sorteo.fecha_inicio_inscripcion = fecha_ini
            sorteo.fecha_fin_inscripcion = fecha_fin
            sorteo.fecha_sorteo = fecha_sor
            sorteo.activo = activo

            imagen = request.FILES.get('imagen')
            if imagen:
                sorteo.imagen = imagen

            sorteo.save()
            messages.success(request, 'Sorteo guardado.')
            return redirect('sorteos:admin_lista')
        except Exception as e:
            messages.error(request, f'Error: {e}')
            return redirect(request.path)


class InscritosSorteoView(LoginRequiredMixin, _StaffRequiredMixin, View):
    template_name = 'sorteos/admin/inscritos.html'

    def get(self, request, pk):
        sorteo = get_object_or_404(Sorteo, pk=pk)
        inscripciones = InscripcionSorteo.objects.filter(sorteo=sorteo)\
            .select_related('empleado', 'empleado__usuario').order_by('numero')
        ganadores = GanadorSorteo.objects.filter(sorteo=sorteo)\
            .select_related('inscripcion__empleado').order_by('orden_premio')
        return render(request, self.template_name, {
            'sorteo': sorteo,
            'inscripciones': inscripciones,
            'ganadores': ganadores,
        })


class ExportarInscritosView(LoginRequiredMixin, _StaffRequiredMixin, View):
    def get(self, request, pk):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        sorteo = get_object_or_404(Sorteo, pk=pk)
        wb = Workbook()
        ws = wb.active
        ws.title = 'Inscritos'

        headers = ['N°', 'Documento', 'Empleado', 'Cargo', 'Área', 'Fecha inscripción']
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', fgColor='305496')

        row = 2
        for ins in InscripcionSorteo.objects.filter(sorteo=sorteo)\
                .select_related('empleado').order_by('numero'):
            emp = ins.empleado
            hc = emp.historialcargo_set.filter(activo=True).select_related('cargo__area').first()
            ws.cell(row=row, column=1, value=ins.numero)
            ws.cell(row=row, column=2, value=emp.numero_documento)
            ws.cell(row=row, column=3, value=emp.nombre_completo)
            ws.cell(row=row, column=4, value=hc.cargo.nombre if hc else '')
            ws.cell(row=row, column=5, value=hc.cargo.area.nombre if (hc and hc.cargo.area) else '')
            ws.cell(row=row, column=6, value=timezone.localtime(ins.fecha_inscripcion)
                    .strftime('%Y-%m-%d %H:%M'))
            row += 1

        for col, width in enumerate([6, 14, 34, 30, 24, 20], 1):
            ws.column_dimensions[chr(64 + col)].width = width
        ws.freeze_panes = 'A2'

        buf = BytesIO()
        wb.save(buf)
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="sorteo_{sorteo.codigo}_inscritos.xlsx"'
        return resp


class RealizarSorteoView(LoginRequiredMixin, _StaffRequiredMixin, View):
    """Pantalla del sorteo: input número ganador → muestra ganador → registrar."""
    template_name = 'sorteos/admin/realizar.html'

    def get(self, request, pk):
        sorteo = get_object_or_404(Sorteo, pk=pk)
        ganadores = GanadorSorteo.objects.filter(sorteo=sorteo)\
            .select_related('inscripcion__empleado').order_by('orden_premio')
        return render(request, self.template_name, {
            'sorteo': sorteo,
            'ganadores': ganadores,
            'total_inscritos': sorteo.total_inscritos,
        })


class BuscarNumeroSorteoView(LoginRequiredMixin, _StaffRequiredMixin, View):
    """AJAX: dado un número, devuelve al empleado si existe y no ha ganado."""

    def post(self, request, pk):
        import json
        sorteo = get_object_or_404(Sorteo, pk=pk)
        try:
            body = json.loads(request.body or '{}')
            numero = int(body.get('numero'))
        except (ValueError, TypeError):
            return JsonResponse({'ok': False, 'error': 'Número inválido.'}, status=400)

        ins = InscripcionSorteo.objects.filter(sorteo=sorteo, numero=numero)\
            .select_related('empleado', 'empleado__usuario', 'ganador').first()
        if not ins:
            return JsonResponse({
                'ok': False,
                'error': f'No hay ningún participante con el número {numero}. Verifica e ingresa otro.',
            })
        if hasattr(ins, 'ganador'):
            return JsonResponse({
                'ok': False,
                'error': f'El número {numero} ya fue seleccionado como premio '
                         f'{ins.ganador.orden_premio}. Extrae otro número.',
            })

        emp = ins.empleado
        hc = emp.historialcargo_set.filter(activo=True).select_related('cargo__area').first()
        foto_url = ''
        if getattr(emp, 'foto', None):
            try:
                foto_url = emp.foto.url
            except Exception:
                foto_url = ''
        return JsonResponse({
            'ok': True,
            'inscripcion_id': str(ins.id),
            'numero': ins.numero,
            'empleado': {
                'nombre': emp.nombre_completo,
                'documento': emp.numero_documento,
                'cargo': hc.cargo.nombre if hc else '—',
                'area': hc.cargo.area.nombre if (hc and hc.cargo.area) else '—',
                'foto_url': foto_url,
            },
        })


class RegistrarGanadorView(LoginRequiredMixin, _StaffRequiredMixin, View):
    """Confirma un ganador: crea el GanadorSorteo con orden_premio siguiente."""

    def post(self, request, pk):
        import json
        sorteo = get_object_or_404(Sorteo, pk=pk)
        try:
            body = json.loads(request.body or '{}')
            inscripcion_id = body.get('inscripcion_id')
        except ValueError:
            return JsonResponse({'ok': False, 'error': 'Datos inválidos.'}, status=400)

        ins = get_object_or_404(InscripcionSorteo, pk=inscripcion_id, sorteo=sorteo)
        if hasattr(ins, 'ganador'):
            return JsonResponse({'ok': False, 'error': 'Ese participante ya está registrado como ganador.'})
        if sorteo.sorteo_completado:
            return JsonResponse({'ok': False, 'error': 'Ya se entregaron todos los premios de este sorteo.'})

        with transaction.atomic():
            ultimo = GanadorSorteo.objects.select_for_update().filter(
                sorteo=sorteo,
            ).aggregate(m=Max('orden_premio'))['m'] or 0
            ganador = GanadorSorteo.objects.create(
                sorteo=sorteo, inscripcion=ins,
                orden_premio=ultimo + 1, seleccionado_por=request.user,
            )
        return JsonResponse({
            'ok': True,
            'orden_premio': ganador.orden_premio,
            'premios_restantes': sorteo.premios_restantes,
            'sorteo_completado': sorteo.sorteo_completado,
        })
