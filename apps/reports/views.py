from django.shortcuts import render
from django.views.generic import TemplateView, View
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.db.models import Avg, Count, Q, F, Case, When, IntegerField
from django.utils import timezone
from django.http import HttpResponse
from datetime import timedelta, datetime

# Importar modelos que ya funcionan
from apps.employees.models import Empleado, AsistenciaDiaria, NovedadNomina
from apps.training.models import Capacitacion
from apps.evaluations.models import AsignacionEvaluacion, PlanMejoraPredefinido, SeguimientoBimensual

# Importar bibliotecas para exportación
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT


@method_decorator(login_required, name='dispatch')
class DashboardView(TemplateView):
    """Dashboard de reportes con datos reales"""
    template_name = 'reports/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        from apps.reports.aprendices_sena import calcular_estado
        # Datos reales de módulos que funcionan
        context.update({
            'total_empleados': Empleado.objects.count(),
            'capacitaciones_activas': Capacitacion.objects.filter(activa=True).count(),
            'evaluaciones_pendientes': AsignacionEvaluacion.objects.filter(estado='pendiente').count(),
            'sena_estado': calcular_estado(),

            # Placeholders para módulos no implementados
            'satisfaccion_general': 0.0,  # Pendiente: módulo surveys
            'reconocimientos_mes': 0,     # Pendiente: módulo recognition
        })

        return context


@method_decorator(login_required, name='dispatch')
class PerformanceReportView(TemplateView):
    """Reporte específico de evaluaciones de desempeño"""
    template_name = 'reports/evaluations_report.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # ============ 1. KPIs PRINCIPALES ============

        # Total de empleados elegibles para evaluación anual
        # Deben cumplir:
        # 1. Estado activo (código "999")
        # 2. Al menos 5 meses de antigüedad como activos (150 días desde activación)
        #    - Activación = fecha_ingreso + 60 días (período de prueba)
        #    - Elegible = activación + 150 días = fecha_ingreso + 210 días
        fecha_limite = timezone.now().date() - timedelta(days=210)
        total_empleados = Empleado.objects.filter(
            estado__codigo='999',  # Solo activos
            fecha_ingreso__lte=fecha_limite  # Con al menos 210 días desde ingreso
        ).count()

        # Evaluaciones completadas (con puntaje calculado)
        # IMPORTANTE: Solo evaluaciones de DESEMPEÑO ANUAL (escala 1-5, puntaje en porcentaje)
        # Excluimos evaluaciones de PERÍODO DE PRUEBA (escala 1-3, puntaje 1-21)
        evaluaciones_completadas = AsignacionEvaluacion.objects.filter(
            estado='completada',
            puntaje_total__isnull=False,
            evaluacion__tipo_evaluacion__codigo__startswith='ANUAL_'  # Solo evaluaciones anuales
        )

        # Promedio General de Desempeño (basado en porcentaje de evaluación)
        promedio_desempeño = evaluaciones_completadas.aggregate(
            promedio=Avg('puntaje_total')
        )['promedio'] or 0

        # Tasa de Evaluaciones Completas (completadas vs total)
        # Solo contar evaluaciones anuales para que coincida con el filtro de completadas
        total_evaluaciones = AsignacionEvaluacion.objects.filter(
            evaluacion__tipo_evaluacion__codigo__startswith='ANUAL_'
        ).count()
        tasa_completadas = (evaluaciones_completadas.count() / total_evaluaciones * 100) if total_evaluaciones > 0 else 0

        # ============ 2. DISTRIBUCIÓN DE NIVELES ============

        # Clasificar evaluaciones por nivel según puntaje
        distribucion_niveles = {
            'muy_alto': evaluaciones_completadas.filter(puntaje_total__gte=90).count(),
            'alto': evaluaciones_completadas.filter(puntaje_total__gte=75, puntaje_total__lt=90).count(),
            'moderado': evaluaciones_completadas.filter(puntaje_total__gte=60, puntaje_total__lt=75).count(),
            'bajo': evaluaciones_completadas.filter(puntaje_total__gte=40, puntaje_total__lt=60).count(),
            'muy_bajo': evaluaciones_completadas.filter(puntaje_total__lt=40).count(),
        }

        # ============ 2.5. ANÁLISIS POR COMPETENCIAS ============

        from apps.evaluations.models import RespuestaEvaluacion
        from collections import defaultdict

        # Definir las categorías/competencias que se evalúan
        categorias_competencias = [
            'Competencias Organizacionales',
            'Objetivos - El Hacer',
            'Competencias Interpersonales - El Ser',
            'Competencias Técnicas - El Saber'
        ]

        # Pesos de cada competencia
        pesos_competencias = {
            'Competencias Organizacionales': 10.0,
            'Objetivos - El Hacer': 40.0,
            'Competencias Interpersonales - El Ser': 25.0,
            'Competencias Técnicas - El Saber': 25.0
        }

        # Función para mapear variantes de nombres a la categoría estándar
        def normalizar_categoria(categoria_original):
            """
            Normaliza variantes de nombres de categorías para agruparlas correctamente.
            Ejemplo: 'Objetivos' -> 'Objetivos - El Hacer'
                     'Competencias Técnicas' -> 'Competencias Técnicas - El Saber'
            """
            if not categoria_original:
                return None

            # Mapeo de variantes
            if 'Organizacionales' in categoria_original:
                return 'Competencias Organizacionales'
            elif 'Objetivos' in categoria_original:
                return 'Objetivos - El Hacer'
            elif 'Interpersonales' in categoria_original:
                return 'Competencias Interpersonales - El Ser'
            elif 'Técnicas' in categoria_original or 'Tecnicas' in categoria_original:
                return 'Competencias Técnicas - El Saber'

            return None  # No es una competencia reconocida

        # Inicializar estructuras para acumular datos
        analisis_competencias = {}

        for categoria in categorias_competencias:
            analisis_competencias[categoria] = {
                'nombre': categoria,
                'peso': pesos_competencias.get(categoria, 0),
                'suma_porcentajes': 0,
                'cantidad_evaluaciones': 0,
                'promedio': 0,
                'nivel_muy_alto': 0,  # ≥90%
                'nivel_alto': 0,       # 75-89%
                'nivel_moderado': 0,   # 60-74%
                'nivel_bajo': 0,       # 40-59%
                'nivel_muy_bajo': 0,   # <40%
            }

        # Procesar cada evaluación completada para calcular promedios por competencia
        for asignacion in evaluaciones_completadas:
            # Obtener todas las respuestas de esta evaluación
            respuestas = RespuestaEvaluacion.objects.filter(
                asignacion=asignacion
            ).select_related('pregunta', 'opcion_seleccionada')

            # Agrupar respuestas por categoría
            puntajes_por_categoria = defaultdict(list)

            for respuesta in respuestas:
                categoria_original = respuesta.pregunta.categoria

                # Normalizar el nombre de la categoría
                categoria_normalizada = normalizar_categoria(categoria_original)

                # Solo procesar si es una competencia reconocida
                if categoria_normalizada:
                    # Obtener el valor numérico de la respuesta (escala 1-5)
                    if respuesta.opcion_seleccionada:
                        valor = float(respuesta.opcion_seleccionada.valor_numerico)
                        puntajes_por_categoria[categoria_normalizada].append(valor)

            # Calcular promedio por categoría para esta evaluación y clasificar
            for categoria, valores in puntajes_por_categoria.items():
                if valores:
                    # Promedio en escala 1-5
                    promedio_escala = sum(valores) / len(valores)
                    # Convertir a porcentaje (1-5 → 0-100%)
                    porcentaje = ((promedio_escala - 1) / 4) * 100

                    # Acumular para el promedio general
                    analisis_competencias[categoria]['suma_porcentajes'] += porcentaje
                    analisis_competencias[categoria]['cantidad_evaluaciones'] += 1

                    # Clasificar por nivel
                    if porcentaje >= 90:
                        analisis_competencias[categoria]['nivel_muy_alto'] += 1
                    elif porcentaje >= 75:
                        analisis_competencias[categoria]['nivel_alto'] += 1
                    elif porcentaje >= 60:
                        analisis_competencias[categoria]['nivel_moderado'] += 1
                    elif porcentaje >= 40:
                        analisis_competencias[categoria]['nivel_bajo'] += 1
                    else:
                        analisis_competencias[categoria]['nivel_muy_bajo'] += 1

        # Calcular promedios finales y brecha respecto al 80%
        for categoria, datos in analisis_competencias.items():
            if datos['cantidad_evaluaciones'] > 0:
                datos['promedio'] = round(datos['suma_porcentajes'] / datos['cantidad_evaluaciones'], 2)
                datos['brecha_80'] = round(datos['promedio'] - 80, 2)
            else:
                datos['promedio'] = 0
                datos['brecha_80'] = -80

        # ============ 2.6. ANÁLISIS DETALLADO POR PREGUNTA EN CADA COMPETENCIA ============

        from apps.evaluations.models import PreguntaEvaluacion

        # Obtener análisis detallado de preguntas por competencia
        for categoria in categorias_competencias:
            # Obtener todas las preguntas de esta competencia (normalizadas)
            todas_preguntas = PreguntaEvaluacion.objects.filter(activa=True)

            preguntas_competencia = []
            for pregunta in todas_preguntas:
                if normalizar_categoria(pregunta.categoria) == categoria:
                    preguntas_competencia.append(pregunta)

            # Calcular promedio por pregunta
            analisis_preguntas = []
            for pregunta in preguntas_competencia:
                # Obtener todas las respuestas a esta pregunta en evaluaciones completadas
                respuestas_pregunta = RespuestaEvaluacion.objects.filter(
                    asignacion__in=evaluaciones_completadas,
                    pregunta=pregunta,
                    opcion_seleccionada__isnull=False
                ).select_related('opcion_seleccionada')

                if respuestas_pregunta.exists():
                    suma_valores = sum(float(r.opcion_seleccionada.valor_numerico) for r in respuestas_pregunta)
                    cantidad = respuestas_pregunta.count()
                    promedio_escala = suma_valores / cantidad
                    promedio_porcentaje = ((promedio_escala - 1) / 4) * 100

                    # Obtener información de la evaluación (tipo/cargo)
                    nombre_evaluacion = pregunta.evaluacion.nombre if pregunta.evaluacion else "N/A"
                    tipo_evaluacion = pregunta.evaluacion.tipo_evaluacion.nombre if pregunta.evaluacion and pregunta.evaluacion.tipo_evaluacion else "N/A"

                    analisis_preguntas.append({
                        'pregunta': pregunta.pregunta,
                        'promedio_escala': round(promedio_escala, 2),
                        'promedio_porcentaje': round(promedio_porcentaje, 2),
                        'cantidad_respuestas': cantidad,
                        'orden': pregunta.orden,
                        'nombre_evaluacion': nombre_evaluacion,
                        'tipo_evaluacion': tipo_evaluacion
                    })

            # Ordenar por orden de la pregunta
            analisis_preguntas.sort(key=lambda x: x['orden'])

            # Agrupar preguntas por tipo de evaluación
            preguntas_por_tipo = defaultdict(list)
            for pregunta_info in analisis_preguntas:
                tipo = pregunta_info['nombre_evaluacion']
                preguntas_por_tipo[tipo].append(pregunta_info)

            # Convertir a lista de diccionarios para el template
            grupos_evaluacion = []
            for tipo_eval, preguntas_grupo in preguntas_por_tipo.items():
                # Calcular promedio del grupo
                if preguntas_grupo:
                    promedio_grupo = sum(p['promedio_porcentaje'] for p in preguntas_grupo) / len(preguntas_grupo)
                    total_respuestas_grupo = sum(p['cantidad_respuestas'] for p in preguntas_grupo)
                else:
                    promedio_grupo = 0
                    total_respuestas_grupo = 0

                grupos_evaluacion.append({
                    'nombre': tipo_eval,
                    'preguntas': preguntas_grupo,
                    'total_preguntas': len(preguntas_grupo),
                    'promedio_grupo': round(promedio_grupo, 2),
                    'total_respuestas': total_respuestas_grupo
                })

            # Ordenar grupos por nombre
            grupos_evaluacion.sort(key=lambda x: x['nombre'])

            # Agregar al análisis de competencias
            analisis_competencias[categoria]['preguntas'] = analisis_preguntas  # Mantener para compatibilidad
            analisis_competencias[categoria]['grupos_evaluacion'] = grupos_evaluacion
            analisis_competencias[categoria]['total_preguntas'] = len(preguntas_competencia)
            analisis_competencias[categoria]['total_grupos'] = len(grupos_evaluacion)

        # Convertir a lista manteniendo el orden lógico de las competencias
        # (no ordenar por promedio para mantener consistencia en el gráfico de radar)
        analisis_competencias_list = [
            analisis_competencias[cat] for cat in categorias_competencias
        ]

        # ============ 3. ALERTAS CRÍTICAS ============

        # Empleados con desempeño bajo (menos de 60% en evaluaciones anuales)
        empleados_bajo_desempeño = AsignacionEvaluacion.objects.filter(
            estado='completada',
            puntaje_total__lt=60,
            evaluacion__tipo_evaluacion__codigo__startswith='ANUAL_'  # Solo evaluaciones anuales
        ).select_related('empleado_evaluado').order_by('puntaje_total')[:10]

        # Evaluaciones vencidas (solo evaluaciones anuales)
        evaluaciones_vencidas = AsignacionEvaluacion.objects.filter(
            estado__in=['pendiente', 'en_progreso'],
            fecha_vencimiento__lt=timezone.now().date(),
            evaluacion__tipo_evaluacion__codigo__startswith='ANUAL_'  # Solo evaluaciones anuales
        ).select_related('empleado_evaluado').order_by('fecha_vencimiento')[:10]

        # Seguimientos atrasados
        seguimientos_atrasados = SeguimientoBimensual.objects.filter(
            estado='atrasado'
        ).select_related('plan_mejora__asignacion_evaluacion__empleado_evaluado').order_by('fecha_limite')[:10]

        # ============ 4. ANÁLISIS POR CARGO ============

        # Obtener evaluaciones con cargo del empleado
        from apps.organizational.models import Cargo, AreaEmpresa

        analisis_por_cargo = []
        cargos = Cargo.objects.all()

        for cargo in cargos:
            # Obtener IDs de empleados con este cargo activo
            empleados_ids = list(Empleado.objects.filter(
                historialcargo__cargo_id=cargo.id,
                historialcargo__activo=True
            ).values_list('id', flat=True).distinct())

            if not empleados_ids:
                continue

            # Evaluaciones de estos empleados
            evals_cargo = evaluaciones_completadas.filter(
                empleado_evaluado_id__in=empleados_ids
            )

            if evals_cargo.exists():
                promedio = evals_cargo.aggregate(promedio=Avg('puntaje_total'))['promedio']
                cantidad = evals_cargo.count()

                analisis_por_cargo.append({
                    'cargo': cargo,
                    'promedio': round(promedio, 2) if promedio else 0,
                    'cantidad_evaluados': cantidad,
                })

        # Ordenar por promedio (mejor → peor)
        analisis_por_cargo = sorted(analisis_por_cargo, key=lambda x: x['promedio'], reverse=True)

        # ============ 5. ANÁLISIS POR ÁREA ============

        analisis_por_area = []
        areas = AreaEmpresa.objects.all()

        for area in areas:
            # Obtener IDs de cargos del área
            cargos_ids = list(Cargo.objects.filter(area_id=area.id).values_list('id', flat=True))

            if not cargos_ids:
                continue

            # IDs de empleados con cargos de esta área
            empleados_ids = list(Empleado.objects.filter(
                historialcargo__cargo_id__in=cargos_ids,
                historialcargo__activo=True
            ).values_list('id', flat=True).distinct())

            if not empleados_ids:
                continue

            # Evaluaciones de estos empleados
            evals_area = evaluaciones_completadas.filter(
                empleado_evaluado_id__in=empleados_ids
            )

            if evals_area.exists():
                promedio = evals_area.aggregate(promedio=Avg('puntaje_total'))['promedio']
                cantidad = evals_area.count()

                analisis_por_area.append({
                    'area': area,
                    'promedio': round(promedio, 2) if promedio else 0,
                    'cantidad_evaluados': cantidad,
                })

        # Ordenar por promedio
        analisis_por_area = sorted(analisis_por_area, key=lambda x: x['promedio'], reverse=True)

        # ============ 6. ANÁLISIS POR SEDE ============

        from apps.organizational.models import Sede

        analisis_por_sede = []
        sedes = Sede.objects.filter(activa=True)

        for sede in sedes:
            # Obtener IDs de empleados de esta sede
            empleados_ids = list(Empleado.objects.filter(
                sede_id=sede.id,
                estado__permite_acceso_sistema=True
            ).values_list('id', flat=True).distinct())

            if not empleados_ids:
                continue

            # Evaluaciones de estos empleados
            evals_sede = evaluaciones_completadas.filter(
                empleado_evaluado_id__in=empleados_ids
            )

            if evals_sede.exists():
                promedio = evals_sede.aggregate(promedio=Avg('puntaje_total'))['promedio']
                cantidad = evals_sede.count()

                analisis_por_sede.append({
                    'sede': sede,
                    'promedio': round(promedio, 2) if promedio else 0,
                    'cantidad_evaluados': cantidad,
                })

        # Ordenar por promedio
        analisis_por_sede = sorted(analisis_por_sede, key=lambda x: x['promedio'], reverse=True)

        # ============ 7. ANÁLISIS DE PLANES DE MEJORA ============

        total_planes = PlanMejoraPredefinido.objects.count()
        planes_por_estado = {
            'pendiente_aprobacion': PlanMejoraPredefinido.objects.filter(estado='pendiente_aprobacion').count(),
            'aprobado': PlanMejoraPredefinido.objects.filter(estado='aprobado').count(),
            'en_seguimiento': PlanMejoraPredefinido.objects.filter(estado='en_seguimiento').count(),
            'completado': PlanMejoraPredefinido.objects.filter(estado='completado').count(),
            'rechazado': PlanMejoraPredefinido.objects.filter(estado='rechazado').count(),
        }

        # Tasa de aceptación de empleados
        planes_aceptados = PlanMejoraPredefinido.objects.filter(aceptado_por_empleado=True).count()
        tasa_aceptacion = (planes_aceptados / total_planes * 100) if total_planes > 0 else 0

        # ============ AGREGAR TODO AL CONTEXT ============

        context.update({
            # KPIs
            'total_empleados': total_empleados,
            'promedio_desempeño': round(promedio_desempeño, 2),
            'tasa_completadas': round(tasa_completadas, 2),
            'total_evaluaciones': total_evaluaciones,
            'evaluaciones_completadas_count': evaluaciones_completadas.count(),

            # Distribución
            'distribucion_niveles': distribucion_niveles,

            # Competencias
            'analisis_competencias': analisis_competencias_list,

            # Alertas
            'empleados_bajo_desempeño': empleados_bajo_desempeño,
            'evaluaciones_vencidas': evaluaciones_vencidas,
            'seguimientos_atrasados': seguimientos_atrasados,

            # Análisis
            'analisis_por_cargo': analisis_por_cargo,
            'analisis_por_area': analisis_por_area,
            'analisis_por_sede': analisis_por_sede,

            # Planes
            'total_planes': total_planes,
            'planes_por_estado': planes_por_estado,
            'tasa_aceptacion': round(tasa_aceptacion, 2),
        })

        return context


# =============================================================================
# VISTAS DE EXPORTACIÓN
# =============================================================================

@method_decorator(login_required, name='dispatch')
class ExportEvaluationsExcelView(View):
    """Exportar reporte de evaluaciones a Excel"""

    def get(self, request, *args, **kwargs):
        # Crear workbook
        wb = Workbook()

        # Obtener datos (reutilizamos la misma lógica)
        total_empleados = Empleado.objects.filter(estado__permite_acceso_sistema=True).count()

        # Solo evaluaciones de DESEMPEÑO ANUAL (puntaje en porcentaje 0-100)
        evaluaciones_completadas = AsignacionEvaluacion.objects.filter(
            estado='completada',
            puntaje_total__isnull=False,
            evaluacion__tipo_evaluacion__codigo__startswith='ANUAL_'
        )

        promedio_desempeño = evaluaciones_completadas.aggregate(
            promedio=Avg('puntaje_total')
        )['promedio'] or 0

        total_evaluaciones = AsignacionEvaluacion.objects.filter(
            evaluacion__tipo_evaluacion__codigo__startswith='ANUAL_'
        ).count()
        tasa_completadas = (evaluaciones_completadas.count() / total_evaluaciones * 100) if total_evaluaciones > 0 else 0

        # Análisis por cargo
        from apps.organizational.models import Cargo, AreaEmpresa

        analisis_por_cargo = []
        cargos = Cargo.objects.all()

        for cargo in cargos:
            empleados_ids = list(Empleado.objects.filter(
                historialcargo__cargo_id=cargo.id,
                historialcargo__activo=True
            ).values_list('id', flat=True).distinct())

            if not empleados_ids:
                continue

            evals_cargo = evaluaciones_completadas.filter(
                empleado_evaluado_id__in=empleados_ids
            )

            if evals_cargo.exists():
                promedio = evals_cargo.aggregate(promedio=Avg('puntaje_total'))['promedio']
                cantidad = evals_cargo.count()

                analisis_por_cargo.append({
                    'cargo': cargo,
                    'promedio': round(promedio, 2) if promedio else 0,
                    'cantidad_evaluados': cantidad,
                })

        analisis_por_cargo = sorted(analisis_por_cargo, key=lambda x: x['promedio'], reverse=True)

        # Análisis por área
        analisis_por_area = []
        areas = AreaEmpresa.objects.all()

        for area in areas:
            cargos_ids = list(Cargo.objects.filter(area_id=area.id).values_list('id', flat=True))

            if not cargos_ids:
                continue

            empleados_ids = list(Empleado.objects.filter(
                historialcargo__cargo_id__in=cargos_ids,
                historialcargo__activo=True
            ).values_list('id', flat=True).distinct())

            if not empleados_ids:
                continue

            evals_area = evaluaciones_completadas.filter(
                empleado_evaluado_id__in=empleados_ids
            )

            if evals_area.exists():
                promedio = evals_area.aggregate(promedio=Avg('puntaje_total'))['promedio']
                cantidad = evals_area.count()

                analisis_por_area.append({
                    'area': area,
                    'promedio': round(promedio, 2) if promedio else 0,
                    'cantidad_evaluados': cantidad,
                })

        analisis_por_area = sorted(analisis_por_area, key=lambda x: x['promedio'], reverse=True)

        # ===== HOJA 1: RESUMEN EJECUTIVO =====
        ws1 = wb.active
        ws1.title = "Resumen Ejecutivo"

        # Estilos
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        title_font = Font(bold=True, size=14, color="1F4E78")

        # Título
        ws1['A1'] = 'REPORTE DE EVALUACIONES DE DESEMPEÑO'
        ws1['A1'].font = Font(bold=True, size=16, color="1F4E78")
        ws1.merge_cells('A1:D1')

        ws1['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws1.merge_cells('A2:D2')

        # KPIs
        row = 4
        ws1[f'A{row}'] = 'INDICADORES CLAVE'
        ws1[f'A{row}'].font = title_font
        row += 1

        kpis = [
            ('Promedio General de Desempeño', f'{round(promedio_desempeño, 2)} / 100'),
            ('Tasa de Completación', f'{round(tasa_completadas, 2)}%'),
            ('Total Empleados Activos', total_empleados),
            ('Evaluaciones Completadas', f'{evaluaciones_completadas.count()} de {total_evaluaciones}'),
        ]

        for indicador, valor in kpis:
            ws1[f'A{row}'] = indicador
            ws1[f'B{row}'] = valor
            ws1[f'A{row}'].font = Font(bold=True)
            row += 1

        # ===== HOJA 2: ANÁLISIS POR CARGO =====
        ws2 = wb.create_sheet("Análisis por Cargo")

        # Headers
        headers = ['Ranking', 'Cargo', 'Área', 'Promedio', 'Cantidad Evaluados', 'Nivel']
        ws2.append(headers)

        for col_num, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Datos
        for idx, item in enumerate(analisis_por_cargo, 1):
            nivel = ''
            if item['promedio'] >= 90:
                nivel = 'Muy Alto'
            elif item['promedio'] >= 75:
                nivel = 'Alto'
            elif item['promedio'] >= 60:
                nivel = 'Moderado'
            elif item['promedio'] >= 40:
                nivel = 'Bajo'
            else:
                nivel = 'Muy Bajo'

            ws2.append([
                idx,
                item['cargo'].nombre,
                item['cargo'].area.nombre,
                item['promedio'],
                item['cantidad_evaluados'],
                nivel
            ])

        # Ajustar anchos de columna
        for col in range(1, 7):
            ws2.column_dimensions[get_column_letter(col)].width = 20

        # ===== HOJA 3: ANÁLISIS POR ÁREA =====
        ws3 = wb.create_sheet("Análisis por Área")

        # Headers
        headers = ['Área', 'Promedio', 'Cantidad Evaluados', 'Nivel']
        ws3.append(headers)

        for col_num, header in enumerate(headers, 1):
            cell = ws3.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Datos
        for item in analisis_por_area:
            nivel = ''
            if item['promedio'] >= 90:
                nivel = 'Muy Alto'
            elif item['promedio'] >= 75:
                nivel = 'Alto'
            elif item['promedio'] >= 60:
                nivel = 'Moderado'
            elif item['promedio'] >= 40:
                nivel = 'Bajo'
            else:
                nivel = 'Muy Bajo'

            ws3.append([
                item['area'].nombre,
                item['promedio'],
                item['cantidad_evaluados'],
                nivel
            ])

        # Ajustar anchos
        for col in range(1, 5):
            ws3.column_dimensions[get_column_letter(col)].width = 25

        # ===== HOJA 4: ANÁLISIS POR SEDE =====

        # Calcular análisis por sede para Excel (reutilizando lógica)
        from apps.organizational.models import Sede

        analisis_por_sede_excel = []
        sedes = Sede.objects.filter(activa=True)

        for sede in sedes:
            empleados_ids = list(Empleado.objects.filter(
                sede_id=sede.id,
                estado__permite_acceso_sistema=True
            ).values_list('id', flat=True).distinct())

            if not empleados_ids:
                continue

            evals_sede = evaluaciones_completadas.filter(
                empleado_evaluado_id__in=empleados_ids
            )

            if evals_sede.exists():
                promedio = evals_sede.aggregate(promedio=Avg('puntaje_total'))['promedio']
                cantidad = evals_sede.count()

                analisis_por_sede_excel.append({
                    'sede': sede,
                    'promedio': round(promedio, 2) if promedio else 0,
                    'cantidad_evaluados': cantidad,
                })

        analisis_por_sede_excel = sorted(analisis_por_sede_excel, key=lambda x: x['promedio'], reverse=True)

        ws4 = wb.create_sheet("Análisis por Sede")

        # Headers
        headers = ['Sede', 'Ciudad', 'Promedio', 'Cantidad Evaluados', 'Nivel']
        ws4.append(headers)

        for col_num, header in enumerate(headers, 1):
            cell = ws4.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Datos
        for item in analisis_por_sede_excel:
            nivel = ''
            if item['promedio'] >= 90:
                nivel = 'Muy Alto'
            elif item['promedio'] >= 75:
                nivel = 'Alto'
            elif item['promedio'] >= 60:
                nivel = 'Moderado'
            elif item['promedio'] >= 40:
                nivel = 'Bajo'
            else:
                nivel = 'Muy Bajo'

            ws4.append([
                item['sede'].nombre,
                item['sede'].ciudad,
                item['promedio'],
                item['cantidad_evaluados'],
                nivel
            ])

        # Ajustar anchos
        for col in range(1, 6):
            ws4.column_dimensions[get_column_letter(col)].width = 25

        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Reporte_Evaluaciones_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'

        wb.save(response)
        return response


@method_decorator(login_required, name='dispatch')
class ExportEvaluationsPDFView(View):
    """Exportar reporte de evaluaciones a PDF"""

    def get(self, request, *args, **kwargs):
        # Crear respuesta HTTP
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=Reporte_Evaluaciones_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf'

        # Crear documento PDF
        doc = SimpleDocTemplate(response, pagesize=letter)
        elementos = []

        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1F4E78'),
            spaceAfter=30,
            alignment=TA_CENTER
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1F4E78'),
            spaceAfter=12,
        )

        # Título
        elementos.append(Paragraph('REPORTE DE EVALUACIONES DE DESEMPEÑO', title_style))
        elementos.append(Paragraph(f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']))
        elementos.append(Spacer(1, 20))

        # Obtener datos
        total_empleados = Empleado.objects.filter(estado__permite_acceso_sistema=True).count()

        # Solo evaluaciones de DESEMPEÑO ANUAL (puntaje en porcentaje 0-100)
        evaluaciones_completadas = AsignacionEvaluacion.objects.filter(
            estado='completada',
            puntaje_total__isnull=False,
            evaluacion__tipo_evaluacion__codigo__startswith='ANUAL_'
        )

        promedio_desempeño = evaluaciones_completadas.aggregate(
            promedio=Avg('puntaje_total')
        )['promedio'] or 0

        total_evaluaciones = AsignacionEvaluacion.objects.filter(
            evaluacion__tipo_evaluacion__codigo__startswith='ANUAL_'
        ).count()
        tasa_completadas = (evaluaciones_completadas.count() / total_evaluaciones * 100) if total_evaluaciones > 0 else 0

        # KPIs
        elementos.append(Paragraph('INDICADORES CLAVE', heading_style))

        kpi_data = [
            ['Indicador', 'Valor'],
            ['Promedio General de Desempeño', f'{round(promedio_desempeño, 2)} / 100'],
            ['Tasa de Completación', f'{round(tasa_completadas, 2)}%'],
            ['Total Empleados Activos', str(total_empleados)],
            ['Evaluaciones Completadas', f'{evaluaciones_completadas.count()} de {total_evaluaciones}'],
        ]

        kpi_table = Table(kpi_data, colWidths=[4*inch, 2*inch])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))

        elementos.append(kpi_table)
        elementos.append(Spacer(1, 20))

        # Análisis por cargo
        from apps.organizational.models import Cargo, AreaEmpresa

        analisis_por_cargo = []
        cargos = Cargo.objects.all()

        for cargo in cargos:
            empleados_ids = list(Empleado.objects.filter(
                historialcargo__cargo_id=cargo.id,
                historialcargo__activo=True
            ).values_list('id', flat=True).distinct())

            if not empleados_ids:
                continue

            evals_cargo = evaluaciones_completadas.filter(
                empleado_evaluado_id__in=empleados_ids
            )

            if evals_cargo.exists():
                promedio = evals_cargo.aggregate(promedio=Avg('puntaje_total'))['promedio']
                cantidad = evals_cargo.count()

                analisis_por_cargo.append({
                    'cargo': cargo,
                    'promedio': round(promedio, 2) if promedio else 0,
                    'cantidad_evaluados': cantidad,
                })

        analisis_por_cargo = sorted(analisis_por_cargo, key=lambda x: x['promedio'], reverse=True)

        if analisis_por_cargo:
            elementos.append(PageBreak())
            elementos.append(Paragraph('ANÁLISIS POR CARGO (Top 10)', heading_style))

            cargo_data = [['Ranking', 'Cargo', 'Promedio', 'Evaluados']]

            for idx, item in enumerate(analisis_por_cargo[:10], 1):
                cargo_data.append([
                    str(idx),
                    item['cargo'].nombre,
                    str(item['promedio']),
                    str(item['cantidad_evaluados'])
                ])

            cargo_table = Table(cargo_data, colWidths=[0.8*inch, 3*inch, 1.2*inch, 1*inch])
            cargo_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))

            elementos.append(cargo_table)

        # Análisis por sede
        from apps.organizational.models import Sede

        analisis_por_sede_pdf = []
        sedes = Sede.objects.filter(activa=True)

        for sede in sedes:
            empleados_ids = list(Empleado.objects.filter(
                sede_id=sede.id,
                estado__permite_acceso_sistema=True
            ).values_list('id', flat=True).distinct())

            if not empleados_ids:
                continue

            evals_sede = evaluaciones_completadas.filter(
                empleado_evaluado_id__in=empleados_ids
            )

            if evals_sede.exists():
                promedio = evals_sede.aggregate(promedio=Avg('puntaje_total'))['promedio']
                cantidad = evals_sede.count()

                analisis_por_sede_pdf.append({
                    'sede': sede,
                    'promedio': round(promedio, 2) if promedio else 0,
                    'cantidad_evaluados': cantidad,
                })

        analisis_por_sede_pdf = sorted(analisis_por_sede_pdf, key=lambda x: x['promedio'], reverse=True)

        if analisis_por_sede_pdf:
            elementos.append(Spacer(1, 20))
            elementos.append(Paragraph('ANÁLISIS POR SEDE', heading_style))

            sede_data = [['Sede', 'Ciudad', 'Promedio', 'Evaluados']]

            for item in analisis_por_sede_pdf:
                sede_data.append([
                    item['sede'].nombre,
                    item['sede'].ciudad,
                    str(item['promedio']),
                    str(item['cantidad_evaluados'])
                ])

            sede_table = Table(sede_data, colWidths=[2*inch, 1.5*inch, 1.2*inch, 1.3*inch])
            sede_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))

            elementos.append(sede_table)

        # Generar PDF
        doc.build(elementos)
        return response


@method_decorator(login_required, name='dispatch')
class AsistenciaReportView(TemplateView):
    """Reporte RRHH de asistencia con KPIs, filtros por rango/área/sede y desglose por empleado."""
    template_name = 'reports/asistencia_report.html'

    def get_context_data(self, **kwargs):
        from datetime import date
        from apps.organizational.models import AreaEmpresa, Sede
        context = super().get_context_data(**kwargs)

        hoy = date.today()

        # === Filtros por query string ===
        fecha_desde_raw = (self.request.GET.get('desde') or '').strip()
        fecha_hasta_raw = (self.request.GET.get('hasta') or '').strip()
        area_id = (self.request.GET.get('area') or '').strip()
        sede_id = (self.request.GET.get('sede') or '').strip()
        # Filtros nuevos:
        # - empleado_id: acota todo el reporte a un solo empleado y activa el
        #   panel de detalle con fechas y motivos de sus ausencias.
        # - tipo_ausencia: 'todos' (default) o un estado específico. Aplica al
        #   listado de por_empleado y al detalle.
        empleado_id = (self.request.GET.get('empleado') or '').strip()
        tipo_ausencia = (self.request.GET.get('tipo_ausencia') or 'todos').strip()
        tipos_validos = {'todos', 'ausente', 'retardo', 'permiso',
                         'permiso_no_remunerado', 'incapacidad', 'en_vacaciones'}
        if tipo_ausencia not in tipos_validos:
            tipo_ausencia = 'todos'

        try:
            fecha_desde = datetime.strptime(fecha_desde_raw, '%Y-%m-%d').date() if fecha_desde_raw else hoy.replace(day=1)
        except ValueError:
            fecha_desde = hoy.replace(day=1)
        try:
            fecha_hasta = datetime.strptime(fecha_hasta_raw, '%Y-%m-%d').date() if fecha_hasta_raw else hoy
        except ValueError:
            fecha_hasta = hoy
        if fecha_desde > fecha_hasta:
            fecha_desde, fecha_hasta = fecha_hasta, fecha_desde

        # === Query base ===
        qs = AsistenciaDiaria.objects.filter(
            fecha__gte=fecha_desde, fecha__lte=fecha_hasta,
        )
        if area_id:
            qs = qs.filter(
                empleado__historialcargo__cargo__area_id=area_id,
                empleado__historialcargo__activo=True,
            ).distinct()
        if sede_id:
            qs = qs.filter(empleado__sede_id=sede_id)
        if empleado_id:
            qs = qs.filter(empleado_id=empleado_id)

        # === KPIs generales ===
        total_registros = qs.count()
        agrupado = qs.aggregate(
            presente=Count('id', filter=Q(estado='presente')),
            retardo=Count('id', filter=Q(estado='retardo')),
            ausente=Count('id', filter=Q(estado='ausente')),
            permiso=Count('id', filter=Q(estado='permiso')),
            permiso_no_remunerado=Count('id', filter=Q(estado='permiso_no_remunerado')),
            incapacidad=Count('id', filter=Q(estado='incapacidad')),
            en_vacaciones=Count('id', filter=Q(estado='en_vacaciones')),
        )
        # % asistencia = presente / (presente + retardo + ausencias reales)
        # (vacaciones no cuenta como ausencia ni como asistencia efectiva)
        base_asistencia = (
            agrupado['presente'] + agrupado['retardo'] + agrupado['ausente']
            + agrupado['permiso'] + agrupado['permiso_no_remunerado']
            + agrupado['incapacidad']
        )
        pct_asistencia = (agrupado['presente'] / base_asistencia * 100) if base_asistencia else 0
        pct_ausentismo = (
            (agrupado['ausente'] + agrupado['permiso'] + agrupado['permiso_no_remunerado'] + agrupado['incapacidad'])
            / base_asistencia * 100
        ) if base_asistencia else 0

        # === Desglose por empleado (top 100 por más ausencias) ===
        # Se anota sobre TODO el queryset (para mantener conteos correctos de
        # 'presente' y 'vacaciones'). El filtro final depende de tipo_ausencia:
        # - 'todos': empleados con al menos 1 ausencia de cualquier tipo
        # - 'ausente'/'retardo'/etc: solo empleados con >=1 de ese tipo específico
        # Vacaciones no cuenta como ausencia por ausentismo (a menos que sea el
        # tipo elegido explícitamente).
        por_empleado_qs = (
            qs.values('empleado__id', 'empleado__nombres', 'empleado__apellidos', 'empleado__numero_documento')
            .annotate(
                total=Count('id'),
                presente=Count('id', filter=Q(estado='presente')),
                retardo=Count('id', filter=Q(estado='retardo')),
                ausente=Count('id', filter=Q(estado='ausente')),
                permiso=Count('id', filter=Q(estado='permiso')),
                permiso_no_remunerado=Count('id', filter=Q(estado='permiso_no_remunerado')),
                incapacidad=Count('id', filter=Q(estado='incapacidad')),
                en_vacaciones=Count('id', filter=Q(estado='en_vacaciones')),
            )
        )
        if tipo_ausencia == 'todos':
            por_empleado_qs = por_empleado_qs.filter(
                Q(ausente__gt=0) | Q(retardo__gt=0) | Q(permiso__gt=0)
                | Q(permiso_no_remunerado__gt=0) | Q(incapacidad__gt=0)
            )
        else:
            por_empleado_qs = por_empleado_qs.filter(**{f'{tipo_ausencia}__gt': 0})
        por_empleado = por_empleado_qs.order_by(
            '-ausente', '-retardo', 'empleado__apellidos',
        )[:100]

        # === Detalle: fechas y motivos cuando hay 1 empleado seleccionado ===
        # Aplicamos el filtro tipo_ausencia también aquí. Excluimos 'presente'
        # por defecto (no es ausencia); si el usuario eligió un tipo específico,
        # solo mostramos ese.
        empleado_detalle = None
        detalle_ausencias = []
        if empleado_id:
            try:
                empleado_detalle = Empleado.objects.select_related('sede').get(pk=empleado_id)
            except (Empleado.DoesNotExist, ValueError):
                empleado_detalle = None
        if empleado_detalle:
            det_qs = qs.filter(empleado=empleado_detalle).select_related('registrado_por')
            if tipo_ausencia == 'todos':
                det_qs = det_qs.exclude(estado='presente')
            else:
                det_qs = det_qs.filter(estado=tipo_ausencia)
            detalle_ausencias = list(det_qs.order_by('-fecha'))

        # === Cumplimiento de registro por jefe ===
        # Un jefe "cumple" un día laboral (L-V) si tiene AL MENOS 1 registro
        # de asistencia de su equipo esa fecha. Ratio = días con registro /
        # días laborales en el rango. El corte inferior es la fecha de
        # arranque del módulo (antes no era omisión del jefe).
        cumplimiento_jefes = self._calcular_cumplimiento_jefes(
            fecha_desde, fecha_hasta, area_id=area_id, sede_id=sede_id,
        )

        # Lista de empleados con al menos un registro en el rango, para el selector.
        empleados_con_registro = (
            Empleado.objects
            .filter(pk__in=qs.values_list('empleado_id', flat=True))
            .order_by('apellidos', 'nombres')
            if not empleado_id else
            Empleado.objects
            .filter(pk__in=AsistenciaDiaria.objects.filter(
                fecha__gte=fecha_desde, fecha__lte=fecha_hasta,
            ).values_list('empleado_id', flat=True))
            .order_by('apellidos', 'nombres')
        )

        context.update({
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'total_registros': total_registros,
            'agrupado': agrupado,
            'pct_asistencia': round(pct_asistencia, 1),
            'pct_ausentismo': round(pct_ausentismo, 1),
            'por_empleado': por_empleado,
            'cumplimiento_jefes': cumplimiento_jefes,
            'areas': AreaEmpresa.objects.filter(activa=True).order_by('nombre'),
            'sedes': Sede.objects.filter(activa=True).order_by('nombre'),
            'area_seleccionada': int(area_id) if area_id.isdigit() else None,
            'sede_seleccionada': sede_id or None,
            # Filtros nuevos:
            'empleado_seleccionado_id': empleado_id or None,
            'tipo_ausencia_seleccionado': tipo_ausencia,
            'empleado_detalle': empleado_detalle,
            'detalle_ausencias': detalle_ausencias,
            'empleados_selector': empleados_con_registro,
            'tipos_ausencia_choices': [
                ('todos', 'Todas las ausencias'),
                ('ausente', 'Ausente sin justificar'),
                ('retardo', 'Retardo'),
                ('permiso', 'Permiso remunerado'),
                ('permiso_no_remunerado', 'Permiso no remunerado'),
                ('incapacidad', 'Incapacidad'),
                ('en_vacaciones', 'En vacaciones'),
            ],
        })
        return context

    def _calcular_cumplimiento_jefes(self, fecha_desde, fecha_hasta, area_id=None, sede_id=None):
        """Cumplimiento de registro de asistencia por jefe en el rango.

        Retorna lista de dicts: {jefe, subordinados_activos, dias_laborales,
        dias_con_registro, dias_faltantes, pct_cumplimiento, fechas_faltantes}.
        Ordenada por peor cumplimiento primero.
        """
        from datetime import timedelta
        from apps.employees.models import HistorialCargo, Empleado
        from apps.employees.views import ASISTENCIA_FECHA_ARRANQUE
        from apps.employees.utils.dias_habiles import es_festivo_oficial

        # Corte inferior: antes de la fecha de arranque el módulo no operaba,
        # así que esos días no cuentan como omisión del jefe.
        fecha_inicio_efectiva = max(fecha_desde, ASISTENCIA_FECHA_ARRANQUE)

        # 1) Días laborales L-V en el rango efectivo — EXCLUYE festivos.
        # Un festivo entre semana no es día laboral, así que no cuenta como
        # omisión del jefe si no registró asistencia esa fecha.
        dias_laborales = []
        cursor = fecha_inicio_efectiva
        while cursor <= fecha_hasta:
            if cursor.weekday() < 5 and not es_festivo_oficial(cursor):
                dias_laborales.append(cursor)
            cursor += timedelta(days=1)

        if not dias_laborales:
            return []

        # 2) Jefes activos: empleados que son jefe_directo de al menos uno
        jefes_ids = HistorialCargo.objects.filter(
            activo=True, jefe_directo__isnull=False,
        ).values_list('jefe_directo', flat=True).distinct()
        jefes = Empleado.objects.filter(pk__in=list(jefes_ids)).select_related('sede')

        resultado = []
        for jefe in jefes:
            # Subordinados activos del jefe
            subs_qs = Empleado.objects.filter(
                historialcargo__activo=True,
                historialcargo__jefe_directo=jefe,
                estado__codigo__in=['999', 'p-prue'],
            ).distinct()
            if area_id:
                subs_qs = subs_qs.filter(
                    historialcargo__cargo__area_id=area_id,
                    historialcargo__activo=True,
                )
            if sede_id:
                subs_qs = subs_qs.filter(sede_id=sede_id)
            subs = list(subs_qs)
            if not subs:
                continue

            # Fechas con al menos 1 registro del equipo (en el rango efectivo)
            fechas_con_registro = set(
                AsistenciaDiaria.objects
                .filter(empleado__in=subs, fecha__gte=fecha_inicio_efectiva, fecha__lte=fecha_hasta)
                .values_list('fecha', flat=True)
                .distinct()
            )
            dias_con_registro = sum(1 for d in dias_laborales if d in fechas_con_registro)
            fechas_faltantes = [d for d in dias_laborales if d not in fechas_con_registro]

            resultado.append({
                'jefe': jefe,
                'subordinados_activos': len(subs),
                'dias_laborales': len(dias_laborales),
                'dias_con_registro': dias_con_registro,
                'dias_faltantes': len(fechas_faltantes),
                'pct_cumplimiento': round(dias_con_registro / len(dias_laborales) * 100, 1),
                'fechas_faltantes': fechas_faltantes[:5],  # primeras 5 para UI
                'faltantes_total': len(fechas_faltantes),
            })

        resultado.sort(key=lambda r: (r['pct_cumplimiento'], -r['dias_faltantes']))
        return resultado


@method_decorator(login_required, name='dispatch')
class NovedadesReportView(TemplateView):
    """Reporte RRHH de novedades de nómina con aprobación batch."""
    template_name = 'reports/novedades_report.html'

    def get_context_data(self, **kwargs):
        from datetime import date, timedelta
        from apps.organizational.models import AreaEmpresa, Sede
        from apps.employees.models import NovedadNomina
        from django.db.models import Sum, Count
        context = super().get_context_data(**kwargs)

        hoy = date.today()
        # Semana actual por defecto (L-D)
        fecha_desde_raw = (self.request.GET.get('desde') or '').strip()
        fecha_hasta_raw = (self.request.GET.get('hasta') or '').strip()
        estado_filtro = (self.request.GET.get('estado') or 'pendiente').strip()
        area_id = (self.request.GET.get('area') or '').strip()
        sede_id = (self.request.GET.get('sede') or '').strip()

        try:
            fecha_desde = datetime.strptime(fecha_desde_raw, '%Y-%m-%d').date() if fecha_desde_raw else (hoy - timedelta(days=hoy.weekday()))
        except ValueError:
            fecha_desde = hoy - timedelta(days=hoy.weekday())
        try:
            fecha_hasta = datetime.strptime(fecha_hasta_raw, '%Y-%m-%d').date() if fecha_hasta_raw else (fecha_desde + timedelta(days=6))
        except ValueError:
            fecha_hasta = fecha_desde + timedelta(days=6)
        if fecha_desde > fecha_hasta:
            fecha_desde, fecha_hasta = fecha_hasta, fecha_desde

        qs = NovedadNomina.objects.filter(
            fecha__gte=fecha_desde, fecha__lte=fecha_hasta,
        ).select_related('empleado', 'registrado_por')

        if estado_filtro in ('pendiente', 'aprobada', 'rechazada'):
            qs = qs.filter(estado_aprobacion=estado_filtro)
        if area_id:
            qs = qs.filter(
                empleado__historialcargo__cargo__area_id=area_id,
                empleado__historialcargo__activo=True,
            ).distinct()
        if sede_id:
            qs = qs.filter(empleado__sede_id=sede_id)

        # KPIs
        totales_por_estado = NovedadNomina.objects.filter(
            fecha__gte=fecha_desde, fecha__lte=fecha_hasta,
        ).values('estado_aprobacion').annotate(
            n=Count('id'), horas=Sum('total_horas'),
        )
        kpis = {'pendiente': 0, 'aprobada': 0, 'rechazada': 0, 'total': 0}
        horas = {'pendiente': 0, 'aprobada': 0, 'rechazada': 0, 'total': 0}
        for row in totales_por_estado:
            k = row['estado_aprobacion']
            kpis[k] = row['n']
            horas[k] = row['horas'] or 0
            kpis['total'] += row['n']
            horas['total'] += row['horas'] or 0

        # Ranking: empleados con más horas extras en el período (aprobadas + pendientes)
        por_empleado = (
            NovedadNomina.objects.filter(fecha__gte=fecha_desde, fecha__lte=fecha_hasta)
            .exclude(estado_aprobacion='rechazada')
            .values('empleado__id', 'empleado__nombres', 'empleado__apellidos', 'empleado__numero_documento')
            .annotate(total_horas=Sum('total_horas'), total_novedades=Count('id'))
            .order_by('-total_horas')[:30]
        )

        context.update({
            'novedades': qs.order_by('-fecha', 'empleado__apellidos'),
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'estado_filtro': estado_filtro,
            'kpis': kpis,
            'horas': horas,
            'por_empleado': por_empleado,
            'areas': AreaEmpresa.objects.filter(activa=True).order_by('nombre'),
            'sedes': Sede.objects.filter(activa=True).order_by('nombre'),
            'area_seleccionada': int(area_id) if area_id.isdigit() else None,
            'sede_seleccionada': sede_id or None,
        })
        return context


@method_decorator(login_required, name='dispatch')
class NovedadesAccionView(View):
    """Aprueba o rechaza novedades en batch (POST). Autorizados:

    - RRHH (is_staff=True): puede accionar sobre CUALQUIER novedad pendiente.
    - Director: puede accionar solo sobre novedades registradas por
      coordinadores/jefes que le reportan directamente.
    """

    def post(self, request):
        from django.contrib import messages as django_messages
        from django.shortcuts import redirect
        from django.utils import timezone
        from apps.employees.models import NovedadNomina, Empleado

        # Verificar que el user tiene algún rol que le permita aprobar
        director = None
        if not request.user.is_staff:
            try:
                director = Empleado.objects.get(usuario=request.user)
            except Empleado.DoesNotExist:
                director = None
            if not director:
                django_messages.error(request, 'Tu usuario no tiene permiso para aprobar novedades.')
                return redirect('reports:novedades_report')

        accion = (request.POST.get('accion') or '').strip()
        ids = request.POST.getlist('novedad_ids')
        motivo_rechazo = (request.POST.get('motivo_rechazo') or '').strip()

        # URL de retorno consistente: siempre a la lista de pendientes con el
        # filtro explícito. Preserva rango de fechas si venía en el referer.
        from urllib.parse import urlparse, parse_qs
        from django.urls import reverse
        base_url = reverse('reports:novedades_report')
        query_extra = ''
        referer = request.META.get('HTTP_REFERER', '')
        if referer:
            qs_parsed = parse_qs(urlparse(referer).query)
            # Preservar desde/hasta/area/sede — pero forzar estado=pendiente
            preservados = {}
            for k in ('desde', 'hasta', 'area', 'sede'):
                if k in qs_parsed and qs_parsed[k]:
                    preservados[k] = qs_parsed[k][0]
            if preservados:
                query_extra = '&' + '&'.join(f'{k}={v}' for k, v in preservados.items())
        redirect_url = f'{base_url}?estado=pendiente{query_extra}'

        if accion not in ('aprobar', 'rechazar'):
            django_messages.error(request, 'Acción inválida.')
            return redirect(redirect_url)
        if not ids:
            django_messages.warning(request, 'No seleccionaste novedades. Marca los checkboxes antes de aprobar o rechazar.')
            return redirect(redirect_url)
        if accion == 'rechazar' and not motivo_rechazo:
            django_messages.error(request, 'Debes indicar el motivo del rechazo.')
            return redirect(redirect_url)

        qs = NovedadNomina.objects.filter(pk__in=ids, estado_aprobacion='pendiente')

        # Restringir al scope del director si no es staff: solo puede accionar
        # sobre novedades registradas por sus subordinados directos.
        if director is not None:
            from django.db.models import Exists, OuterRef
            from apps.employees.models import HistorialCargo
            coord_reporta_a_director = HistorialCargo.objects.filter(
                activo=True,
                empleado=OuterRef('registrado_por'),
                jefe_directo=director,
            )
            qs = qs.annotate(_de_mi_equipo=Exists(coord_reporta_a_director)).filter(_de_mi_equipo=True)

        n_afectadas = qs.count()
        if n_afectadas == 0:
            django_messages.warning(
                request,
                'Ninguna de las novedades seleccionadas te corresponde para aprobar/rechazar.',
            )
            return redirect(request.META.get('HTTP_REFERER') or 'reports:novedades_report')

        if accion == 'aprobar':
            qs.update(
                estado_aprobacion='aprobada',
                aprobado_por_rrhh=request.user,
                fecha_aprobacion=timezone.now(),
                motivo_rechazo='',
            )
            django_messages.success(request, f'{n_afectadas} novedad(es) aprobadas.')
        else:
            qs.update(
                estado_aprobacion='rechazada',
                aprobado_por_rrhh=request.user,
                fecha_aprobacion=timezone.now(),
                motivo_rechazo=motivo_rechazo,
            )
            django_messages.warning(request, f'{n_afectadas} novedad(es) rechazadas.')

        return redirect(request.META.get('HTTP_REFERER') or 'reports:novedades_report')


@method_decorator(login_required, name='dispatch')
class NovedadesExportExcelView(View):
    """Exporta a Excel las novedades globales para RRHH. Requiere is_staff.

    Los jefes usan un endpoint separado en apps.employees que filtra
    automáticamente por su equipo directo (ver employees:novedades_export_excel).
    """

    def get(self, request):
        from datetime import date, timedelta
        from apps.employees.models import NovedadNomina
        from apps.employees.utils.excel_novedades import (
            generar_excel_novedades, nombre_archivo_novedades,
        )
        from django.contrib import messages as django_messages
        from django.shortcuts import redirect

        if not request.user.is_staff:
            django_messages.error(request, 'Solo RRHH puede acceder al reporte global. Usa el botón Exportar en Mis novedades del equipo.')
            return redirect('reports:novedades_report')

        # Filtros
        hoy = date.today()
        fecha_desde_raw = (request.GET.get('desde') or '').strip()
        fecha_hasta_raw = (request.GET.get('hasta') or '').strip()
        estado_filtro = (request.GET.get('estado') or 'todas').strip()
        area_id = (request.GET.get('area') or '').strip()
        sede_id = (request.GET.get('sede') or '').strip()

        try:
            fecha_desde = datetime.strptime(fecha_desde_raw, '%Y-%m-%d').date() if fecha_desde_raw else (hoy - timedelta(days=hoy.weekday()))
        except ValueError:
            fecha_desde = hoy - timedelta(days=hoy.weekday())
        try:
            fecha_hasta = datetime.strptime(fecha_hasta_raw, '%Y-%m-%d').date() if fecha_hasta_raw else (fecha_desde + timedelta(days=6))
        except ValueError:
            fecha_hasta = fecha_desde + timedelta(days=6)
        if fecha_desde > fecha_hasta:
            fecha_desde, fecha_hasta = fecha_hasta, fecha_desde

        qs = NovedadNomina.objects.filter(
            fecha__gte=fecha_desde, fecha__lte=fecha_hasta,
        ).select_related('empleado', 'registrado_por', 'aprobado_por_rrhh')

        filtros_partes = []
        if estado_filtro in ('pendiente', 'aprobada', 'rechazada'):
            qs = qs.filter(estado_aprobacion=estado_filtro)
            filtros_partes.append(f'Estado: {estado_filtro}')
        if area_id:
            qs = qs.filter(
                empleado__historialcargo__cargo__area_id=area_id,
                empleado__historialcargo__activo=True,
            ).distinct()
            try:
                from apps.organizational.models import AreaEmpresa
                area_obj = AreaEmpresa.objects.filter(pk=area_id).first()
                if area_obj:
                    filtros_partes.append(f'Área: {area_obj.nombre}')
            except Exception:
                pass
        if sede_id:
            qs = qs.filter(empleado__sede_id=sede_id)
            try:
                from apps.organizational.models import Sede
                sede_obj = Sede.objects.filter(pk=sede_id).first()
                if sede_obj:
                    filtros_partes.append(f'Sede: {sede_obj.nombre}')
            except Exception:
                pass

        qs = qs.order_by('fecha', 'empleado__apellidos', 'hora_inicio')
        filtros_desc = ' · '.join(filtros_partes) if filtros_partes else 'Sin filtros adicionales'

        contenido = generar_excel_novedades(
            list(qs),
            titulo='Novedades de nómina — Reporte RRHH',
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            filtros_desc=filtros_desc,
            contexto_scope='Todas las áreas y sedes',
        )
        filename = nombre_archivo_novedades('novedades_rrhh', fecha_desde, fecha_hasta)

        response = HttpResponse(
            contenido,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Cache-Control'] = 'no-store'
        return response

@method_decorator(login_required, name='dispatch')
class AsistenciaAusenciasExcelView(View):
    """Exporta a Excel las ausencias del rango con los mismos filtros del
    reporte web: rango, área, sede, empleado, tipo_ausencia. Una fila por
    ausencia con fecha, estado, motivo y jefe que registró.
    """

    def get(self, request):
        hoy = timezone.now().date()

        # Reutilizar la misma lógica de parseo que la vista del reporte
        def _parse(name, default):
            raw = (request.GET.get(name) or '').strip()
            try:
                return datetime.strptime(raw, '%Y-%m-%d').date() if raw else default
            except ValueError:
                return default

        fecha_desde = _parse('desde', hoy.replace(day=1))
        fecha_hasta = _parse('hasta', hoy)
        if fecha_desde > fecha_hasta:
            fecha_desde, fecha_hasta = fecha_hasta, fecha_desde

        area_id = (request.GET.get('area') or '').strip()
        sede_id = (request.GET.get('sede') or '').strip()
        empleado_id = (request.GET.get('empleado') or '').strip()
        tipo_ausencia = (request.GET.get('tipo_ausencia') or 'todos').strip()

        qs = AsistenciaDiaria.objects.filter(
            fecha__gte=fecha_desde, fecha__lte=fecha_hasta,
        ).select_related('empleado', 'registrado_por').order_by('empleado__apellidos', 'fecha')
        if area_id:
            qs = qs.filter(
                empleado__historialcargo__cargo__area_id=area_id,
                empleado__historialcargo__activo=True,
            ).distinct()
        if sede_id:
            qs = qs.filter(empleado__sede_id=sede_id)
        if empleado_id:
            qs = qs.filter(empleado_id=empleado_id)
        # Solo ausencias (excluye presentes) — el reporte se llama "ausencias"
        if tipo_ausencia == 'todos':
            qs = qs.exclude(estado='presente')
        else:
            qs = qs.filter(estado=tipo_ausencia)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Ausencias'

        headers = ['Documento', 'Empleado', 'Sede', 'Fecha', 'Día', 'Estado', 'Motivo', 'Registrado por']
        ws.append(headers)
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='0E5F3F')
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        estados_display = dict(AsistenciaDiaria.ESTADO_CHOICES)
        dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
        for r in qs:
            ws.append([
                r.empleado.numero_documento,
                r.empleado.nombre_completo,
                r.empleado.sede.nombre if r.empleado.sede else '',
                r.fecha.strftime('%d/%m/%Y'),
                dias_semana[r.fecha.weekday()],
                estados_display.get(r.estado, r.estado),
                r.motivo or '',
                r.registrado_por.nombre_completo if r.registrado_por else '',
            ])

        # Anchos
        anchos = [14, 32, 18, 12, 12, 22, 40, 26]
        for i, w in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        sufijo = f'_{empleado_id[:8]}' if empleado_id else ''
        filename = f'ausencias_{fecha_desde:%Y%m%d}_{fecha_hasta:%Y%m%d}{sufijo}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


@method_decorator(login_required, name='dispatch')
class AprendicesSenaReportView(TemplateView):
    """Dashboard de cumplimiento de cuota SENA (aprendices).

    Muestra: cuota vigente vs aprendices actuales, semáforo, sanción estimada
    si hay faltantes, lista de aprendices con fecha_fin estimada, próximos a
    vencer y la resolución vigente con acceso al PDF.
    """
    template_name = 'reports/aprendices_sena.html'

    def get_context_data(self, **kwargs):
        from datetime import date as _date
        from apps.reports.aprendices_sena import calcular_estado, _smmlv_vigente
        from apps.organizational.models import ResolucionSena, SalarioMinimoAnual
        context = super().get_context_data(**kwargs)
        estado = calcular_estado()

        # Porcentaje de cumplimiento (para barra visual). Cap 100.
        if estado.cuota_requerida > 0:
            pct = min(100, int(round(100 * estado.aprendices_actuales / estado.cuota_requerida)))
        else:
            pct = 0

        smmlv_row = SalarioMinimoAnual.objects.filter(year=_date.today().year).first()
        context.update({
            'estado': estado,
            'pct_cumplimiento': pct,
            'smmlv': _smmlv_vigente(),
            'smmlv_actualizado': smmlv_row is not None,
            'smmlv_year_actual': _date.today().year,
            'smmlv_row': smmlv_row,
            'historial_resoluciones': ResolucionSena.objects.order_by('-fecha_vigencia_inicio')[:5],
        })
        return context


@method_decorator(login_required, name='dispatch')
class AprendicesSenaExcelView(View):
    """Descarga XLSX del reporte SENA — pensado para inspección/auditoría.

    Hoja única con 3 secciones apiladas:
    1. Encabezado + resolución vigente.
    2. Resumen de cumplimiento (X/Y, faltantes, sanción estimada).
    3. Tabla completa de aprendices actuales con detalle por columna.
    """

    def get(self, request):
        from datetime import date
        from apps.reports.aprendices_sena import calcular_estado, _smmlv_vigente

        estado = calcular_estado()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Cuota SENA'

        titulo_font = Font(bold=True, size=14, color='FFFFFF')
        titulo_fill = PatternFill('solid', fgColor='0E5F3F')
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='0E5F3F')
        label_font = Font(bold=True)

        row = 1
        # Encabezado
        ws.cell(row=row, column=1, value='REPORTE CUOTA DE APRENDICES SENA').font = titulo_font
        ws.cell(row=row, column=1).fill = titulo_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        ws.row_dimensions[row].height = 24
        row += 1
        ws.cell(row=row, column=1, value=f'Generado el {date.today():%d/%m/%Y}').font = Font(italic=True, size=10)
        row += 2

        # Resolución vigente
        ws.cell(row=row, column=1, value='RESOLUCIÓN VIGENTE').font = label_font
        ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor='DDDDDD')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 1
        if estado.resolucion:
            r = estado.resolucion
            for label, value in [
                ('Número', r.numero),
                ('Expedida', r.fecha_expedicion.strftime('%d/%m/%Y')),
                ('Vigencia desde', r.fecha_vigencia_inicio.strftime('%d/%m/%Y')),
                ('Vigencia hasta', r.fecha_vigencia_fin.strftime('%d/%m/%Y') if r.fecha_vigencia_fin else 'Vigente'),
                ('Cuota (aprendices)', r.cuota_aprendices),
                ('Trabajadores base (informativo)', r.total_trabajadores_base or '—'),
            ]:
                ws.cell(row=row, column=1, value=label).font = label_font
                ws.cell(row=row, column=2, value=value)
                row += 1
        else:
            ws.cell(row=row, column=1, value='(No hay resolución vigente registrada)').font = Font(italic=True, color='990000')
            row += 1
        row += 1

        # Cumplimiento
        ws.cell(row=row, column=1, value='CUMPLIMIENTO ACTUAL').font = label_font
        ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor='DDDDDD')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 1
        smmlv = _smmlv_vigente()
        for label, value in [
            ('Aprendices activos hoy', estado.aprendices_actuales),
            ('Cuota requerida', estado.cuota_requerida),
            ('Faltantes', estado.faltantes),
            ('Cumple', 'SÍ' if estado.cumple else 'NO'),
            ('SMMLV usado (base sanción)', f'${smmlv:,.0f}'.replace(',', '.')),
            ('Sanción mensual estimada', f'${estado.sancion_mensual_estimada:,.0f}'.replace(',', '.')),
        ]:
            ws.cell(row=row, column=1, value=label).font = label_font
            cell = ws.cell(row=row, column=2, value=value)
            if label == 'Cumple' and value == 'NO':
                cell.font = Font(bold=True, color='C62828')
            row += 1
        row += 1

        # Tabla de aprendices
        ws.cell(row=row, column=1, value='APRENDICES ACTIVOS').font = label_font
        ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor='DDDDDD')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 1

        headers = ['Documento', 'Aprendiz', 'Cargo', 'Etapa', 'Inicio', 'Fin estimado', 'Días restantes']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

        for a in estado.aprendices:
            ws.cell(row=row, column=1, value=a.empleado.numero_documento)
            ws.cell(row=row, column=2, value=a.empleado.nombre_completo)
            ws.cell(row=row, column=3, value=a.cargo.nombre)
            ws.cell(row=row, column=4, value=a.etapa.capitalize())
            ws.cell(row=row, column=5, value=a.fecha_inicio.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=6, value=a.fecha_fin_estimada.strftime('%d/%m/%Y'))
            cell_rest = ws.cell(row=row, column=7, value=a.dias_restantes)
            if a.dias_restantes <= 60:
                cell_rest.font = Font(bold=True, color='C62828')
            elif a.dias_restantes <= 180:
                cell_rest.font = Font(color='E65100')
            row += 1

        # Anchos de columna
        anchos = [16, 34, 30, 14, 14, 14, 18]
        for i, w in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f'aprendices_sena_{date.today():%Y%m%d}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
